#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Standalone API server for 1C reconciliation module.

Serves:
  - static files from current directory
  - GET  /api/reconciliation/erp-docs?spec_id=<int>
  - POST /api/reconciliation/compare

Production integration reads ERP from MariaDB and 1C from a read-only REST API
with normalized reconciliation DTOs.
"""

from __future__ import annotations

import json
import os
import re
import subprocess
import tempfile
import time
import threading
import traceback
import uuid
from email.parser import BytesParser
from email.policy import default as EMAIL_POLICY
from io import BytesIO
from datetime import datetime, timedelta
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.request import Request as UrlRequest, urlopen
from urllib.parse import parse_qs, urlparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None

try:
    import psycopg2
    from psycopg2.extras import execute_batch
except Exception:
    psycopg2 = None
    execute_batch = None

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build as build_google_service
except Exception:
    service_account = None
    build_google_service = None

try:
    from recon_erp_1c.infrastructure.onec_rest.client import OneCRestClient, OneCRestError, onec_rest_status
except Exception:
    OneCRestClient = None
    OneCRestError = RuntimeError

    def onec_rest_status() -> dict[str, object]:
        return {"configured": False, "missing": ["onec_rest_client.py"]}


DB_HOST = os.environ.get("PRINT_DB_HOST", "erp.vedagent")
DB_PORT = int(os.environ.get("PRINT_DB_PORT", "3306"))
DB_NAME = os.environ.get("PRINT_DB_NAME", "veda25")
DB_USER = os.environ.get("PRINT_DB_USER", "data")
DB_PASSWORD = os.environ.get("PRINT_DB_PASSWORD", "")

LISTEN_HOST = os.environ.get("RECON_API_HOST", "0.0.0.0")
LISTEN_PORT = int(os.environ.get("RECON_API_PORT", "8780"))
STATIC_ROOT = Path(os.environ.get("RECON_STATIC_ROOT", str(Path(__file__).resolve().parent)))
OPER_URL_TEMPLATE = os.environ.get("RECON_OPERATION_URL_TEMPLATE", "https://vue.vedagent.ru/supplies/operations/{oper_id}")
ONEC_DEFAULT_DIR = Path(os.environ.get("RECON_ONEC_DIR", str(STATIC_ROOT / "akt_sverki" / "1C")))
ONEC_DRIVE_FOLDER_ID = os.environ.get(
    "RECON_ONEC_DRIVE_FOLDER_ID",
    os.environ.get("RECON_GOOGLE_DRIVE_FOLDER_ID", "1iTmAkt2Bs8Oi1Wco-6bJnd-yAbxW_cNt"),
)
CONTROL_SHEET_ID = os.environ.get("RECON_CONTROL_SHEET_ID", "1v50v2vM8Yqf7TeGQ4oY3tRZEvsDHs0lGW8_q0oQqfSE")
CONTROL_SHEET_GID = os.environ.get("RECON_CONTROL_SHEET_GID", "92475489")
GOOGLE_CREDENTIALS_FILE = Path(
    os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", str(STATIC_ROOT / "secrets" / "google-drive-service-account.json"))
)
ONEC_PG_DSN = os.environ.get("RECON_ONEC_PG_DSN", os.environ.get("RECON_PG_DSN", "dbname=reconciliation_cache user=root"))
ONEC_PARSE_CACHE: dict[str, dict[str, object]] = {}
ONEC_REST_CLIENT_MATRIX_MAX_LIMIT = int(os.environ.get("RECON_ONEC_REST_CLIENT_MATRIX_MAX_LIMIT", "10") or "10")
RECON_BACKGROUND_MATRIX_MAX_LIMIT = int(os.environ.get("RECON_BACKGROUND_MATRIX_MAX_LIMIT", "500") or "500")
RECON_JOB_TTL_SECONDS = int(os.environ.get("RECON_JOB_TTL_SECONDS", "3600") or "3600")
RECON_JOB_MAX_ITEMS = int(os.environ.get("RECON_JOB_MAX_ITEMS", "100") or "100")
RECON_JOBS: dict[str, dict[str, object]] = {}
RECON_JOBS_LOCK = threading.Lock()
RECON_AUTH_REQUIRED_RAW = os.environ.get("RECON_AUTH_REQUIRED")
RECON_AUTH_REQUIRED = (
    RECON_AUTH_REQUIRED_RAW.strip().lower()
    if RECON_AUTH_REQUIRED_RAW is not None
    else "0"
) not in {"0", "false", "no", "off"}
RECON_ERP_API_BASE = os.environ.get("RECON_ERP_API_BASE", "http://erp-test.vedagent/veda/api/erp").rstrip("/")
RECON_ERP_TOKEN_VALIDATE_URL = os.environ.get("RECON_ERP_TOKEN_VALIDATE_URL", f"{RECON_ERP_API_BASE}/validatetoken/")
RECON_ERP_TOKEN_AUDIENCE = os.environ.get("RECON_ERP_TOKEN_AUDIENCE", "reconciliation")
RECON_ERP_API_TIMEOUT = int(os.environ.get("RECON_ERP_API_TIMEOUT", "20") or "20")
RECON_ERP_USERS_TABLE = os.environ.get("RECON_ERP_USERS_TABLE", "veda_users")
RECON_ERP_USER_LOGIN_FIELD = os.environ.get("RECON_ERP_USER_LOGIN_FIELD", "f_login")
RECON_ERP_USER_ID_FIELD = os.environ.get("RECON_ERP_USER_ID_FIELD", "f_id")
RECON_ERP_USER_NAME_FIELDS = [
    field.strip()
    for field in os.environ.get("RECON_ERP_USER_NAME_FIELDS", "f_name1,f_name2,f_name3").split(",")
    if field.strip()
]
RECON_ERP_USER_ACTIVE_CLAUSE = os.environ.get("RECON_ERP_USER_ACTIVE_CLAUSE", "").strip()
AUTH_SESSION_TTL_SECONDS = int(os.environ.get("RECON_AUTH_SESSION_TTL_SECONDS", "28800") or "28800")
AUTH_SESSIONS: dict[str, dict[str, object]] = {}
AUTH_SESSIONS_LOCK = threading.Lock()

STATUS_MATCH = "MATCH"
STATUS_NOT_FOUND_IN_1C = "NOT_FOUND_IN_1C"
STATUS_NOT_FOUND_IN_ERP = "NOT_FOUND_IN_ERP"
STATUS_FIELDS_MISMATCH = "FIELDS_MISMATCH"
STATUS_NOT_COMPARABLE = "NOT_COMPARABLE"


def run_mysql_tsv(query: str) -> list[list[str]]:
    cmd = [
        "mysql",
        "-h",
        DB_HOST,
        "-P",
        str(DB_PORT),
        "-u",
        DB_USER,
        f"-p{DB_PASSWORD}",
        DB_NAME,
        "--default-character-set=utf8mb4",
        "-N",
        "-B",
        "-e",
        query,
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or "mysql query failed")
    lines = proc.stdout.splitlines()
    return [line.split("\t") for line in lines if line.strip()]


def first_row(query: str) -> list[str] | None:
    rows = run_mysql_tsv(query)
    return rows[0] if rows else None


def text_at(row: list[str] | None, idx: int, default: str = "") -> str:
    if not row or idx >= len(row):
        return default
    return row[idx]


def int_at(row: list[str] | None, idx: int, default: int = 0) -> int:
    try:
        return int(text_at(row, idx, str(default)).strip() or str(default))
    except Exception:
        return default


def float_at(row: list[str] | None, idx: int, default: float = 0.0) -> float:
    try:
        return float((text_at(row, idx, str(default)) or "").replace(",", ".").strip() or str(default))
    except Exception:
        return default


def sql_int(value: object) -> int:
    try:
        return int(value)
    except Exception:
        return 0


def sql_bool(value: object) -> int:
    return 1 if sql_int(value) == 1 else 0


def sql_decimal(value: object) -> str:
    amount = normalize_sum(value)
    if amount is None:
        return "NULL"
    return f"{amount:.2f}"


def sql_quote(value: object) -> str:
    if value is None:
        return "NULL"
    raw = str(value)
    escaped = raw.replace("\\", "\\\\").replace("'", "''")
    return f"'{escaped}'"


def sql_identifier(value: str) -> str:
    text = normalize_text(value)
    if not re.fullmatch(r"[A-Za-z0-9_]+", text):
        raise ValueError(f"Unsafe SQL identifier: {text}")
    return text


def normalize_email(value: object) -> str:
    return normalize_text(value).lower()


def auth_config_payload() -> dict[str, object]:
    return {
        "required": RECON_AUTH_REQUIRED,
        "mode": "erp_launch_token",
        "erp": {
            "token_validation_configured": bool(RECON_ERP_TOKEN_VALIDATE_URL),
            "token_audience": RECON_ERP_TOKEN_AUDIENCE,
            "users_table": RECON_ERP_USERS_TABLE,
            "login_field": RECON_ERP_USER_LOGIN_FIELD,
            "name_fields": RECON_ERP_USER_NAME_FIELDS,
        },
    }


def fetch_erp_user_by_identity(identity: dict[str, object]) -> dict[str, object]:
    table = sql_identifier(RECON_ERP_USERS_TABLE)
    id_field = sql_identifier(RECON_ERP_USER_ID_FIELD)
    login_field = sql_identifier(RECON_ERP_USER_LOGIN_FIELD)
    name_fields = [sql_identifier(field) for field in RECON_ERP_USER_NAME_FIELDS]
    select_fields = [id_field, login_field, *name_fields]
    active_clause = f" AND ({RECON_ERP_USER_ACTIVE_CLAUSE})" if RECON_ERP_USER_ACTIVE_CLAUSE else ""

    user_id = sql_int(identity.get("id") or identity.get("user_id"))
    login = normalize_email(identity.get("login") or identity.get("email") or identity.get("username"))
    if user_id > 0:
        where = f"{id_field} = {user_id}"
    elif login:
        where = f"LOWER({login_field}) = LOWER({sql_quote(login)})"
    else:
        raise RuntimeError("ERP token validation response does not contain user id or login")

    query = f"""
SELECT {", ".join(select_fields)}
FROM {table}
WHERE {where}
{active_clause}
LIMIT 1;
"""
    row = first_row(query)
    if not row:
        raise RuntimeError("ERP user from launch token was not found in veda_users")
    display_parts = [text_at(row, idx) for idx in range(2, 2 + len(name_fields))]
    display_name = " ".join(part for part in [normalize_text(part) for part in display_parts] if part)
    profile_name = normalize_text(identity.get("display_name") or identity.get("name"))
    email = normalize_email(text_at(row, 1) or login)
    return {
        "id": int_at(row, 0),
        "email": email,
        "display_name": display_name or profile_name or email or f"ERP user {int_at(row, 0)}",
        "roles": identity.get("roles") if isinstance(identity.get("roles"), list) else [],
    }


def validate_erp_launch_token(launch_token: str) -> dict[str, object]:
    token = normalize_text(launch_token)
    if not token:
        raise RuntimeError("ERP launch token is required")
    if not RECON_ERP_TOKEN_VALIDATE_URL:
        raise RuntimeError("ERP token validation URL is not configured")
    body = json.dumps({"token": token, "audience": RECON_ERP_TOKEN_AUDIENCE}, ensure_ascii=False).encode("utf-8")
    request = UrlRequest(
        RECON_ERP_TOKEN_VALIDATE_URL,
        data=body,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    try:
        with urlopen(request, timeout=RECON_ERP_API_TIMEOUT) as response:
            raw = response.read().decode("utf-8", errors="replace")
    except Exception as exc:
        raise RuntimeError(f"ERP token validation failed: {exc}") from exc
    try:
        payload = json.loads(raw)
    except Exception as exc:
        raise RuntimeError("ERP token validation response is not JSON") from exc

    if isinstance(payload, list):
        if len(payload) < 2 or payload[0] is not True:
            raise RuntimeError(normalize_text(payload[1] if len(payload) > 1 else "") or "ERP token validation rejected")
        user_payload = payload[1]
        if not isinstance(user_payload, dict):
            raise RuntimeError("ERP token validation response does not contain user profile")
        return user_payload

    if isinstance(payload, dict):
        if payload.get("ok") is not True and payload.get("success") is not True:
            raise RuntimeError(normalize_text(payload.get("error") or payload.get("message")) or "ERP token validation rejected")
        user_payload = payload.get("user") or payload.get("profile")
        if not isinstance(user_payload, dict):
            raise RuntimeError("ERP token validation response does not contain user profile")
        return user_payload

    raise RuntimeError("ERP token validation response has unexpected format")


def user_initials(display_name: str, email: str = "") -> str:
    words = [word for word in re.split(r"\s+", normalize_text(display_name)) if word]
    if len(words) >= 2:
        return (words[0][:1] + words[1][:1]).upper()
    if words:
        return words[0][:2].upper()
    return normalize_text(email)[:2].upper() or "U"


def prune_auth_sessions(now: float | None = None) -> None:
    now = time.time() if now is None else now
    with AUTH_SESSIONS_LOCK:
        ttl = max(AUTH_SESSION_TTL_SECONDS, 0)
        if ttl:
            for token in list(AUTH_SESSIONS):
                created_at = float(AUTH_SESSIONS.get(token, {}).get("created_at_ts") or 0)
                if now - created_at > ttl:
                    AUTH_SESSIONS.pop(token, None)


def create_auth_session(user: dict[str, object], erp_token: str = "", identity_profile: dict[str, object] | None = None) -> dict[str, object]:
    prune_auth_sessions()
    session_token = uuid.uuid4().hex + uuid.uuid4().hex
    now = time.time()
    identity_profile = identity_profile or {}
    display_name = normalize_text(user.get("display_name") or identity_profile.get("display_name") or identity_profile.get("name") or user.get("email"))
    email = normalize_email(user.get("email") or identity_profile.get("email") or identity_profile.get("login"))
    public_user = {
        "id": sql_int(user.get("id")),
        "email": email,
        "login": normalize_text(identity_profile.get("login") or email),
        "display_name": display_name,
        "initials": user_initials(display_name, email),
        "roles": user.get("roles") if isinstance(user.get("roles"), list) else [],
    }
    session = {
        "token": session_token,
        "created_at_ts": now,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": public_user,
        "erp_token": erp_token,
    }
    with AUTH_SESSIONS_LOCK:
        AUTH_SESSIONS[session_token] = session
    return session


def get_auth_session(session_token: str) -> dict[str, object] | None:
    prune_auth_sessions()
    token = normalize_text(session_token)
    if not token:
        return None
    with AUTH_SESSIONS_LOCK:
        session = AUTH_SESSIONS.get(token)
        return dict(session) if session else None


def public_auth_session(session: dict[str, object], include_erp_token: bool = False) -> dict[str, object]:
    payload = {
        "authenticated": True,
        "session_token": session.get("token"),
        "user": session.get("user"),
        "erp": {
            "token_available": bool(session.get("erp_token")),
        },
    }
    if include_erp_token:
        payload["erp"]["token"] = session.get("erp_token")
    return payload


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def normalize_number(value: object) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""
    return "".join(raw.split()).lower()


def compact_key(value: object) -> str:
    raw = normalize_text(value).lower().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", "", raw)


def money_equal(left: object, right: object, tolerance: float = 0.01) -> bool:
    left_sum = normalize_sum(left)
    right_sum = normalize_sum(right)
    if left_sum is None or right_sum is None:
        return False
    return abs(left_sum - right_sum) <= tolerance


def parse_any_date_to_iso(value: object) -> str:
    raw = normalize_text(value)
    if not raw:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    candidates = ["%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"]
    for fmt in candidates:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue

    if len(raw) >= 10 and raw[4] == "-" and raw[7] == "-":
        return raw[:10]

    return ""


def parse_document_title(value: object) -> dict[str, str]:
    text = normalize_text(value)
    if not text:
        return {"doctype": "", "number": "", "date": "", "date_iso": ""}

    match = re.search(r"^(?P<type>.+?)\s+(?P<number>[^\s]+)\s+от\s+(?P<date>\d{2}\.\d{2}\.\d{4})", text, re.IGNORECASE)
    if match:
        return {
            "doctype": normalize_text(match.group("type")),
            "number": normalize_text(match.group("number")),
            "date": normalize_text(match.group("date")),
            "date_iso": parse_any_date_to_iso(match.group("date")),
        }

    return {"doctype": text, "number": "", "date": "", "date_iso": ""}


def parse_contract_reference(value: object) -> dict[str, str]:
    text = normalize_text(value)
    result = {
        "contract_text": text,
        "request_type": "",
        "spec_number": "",
        "base_contract": "",
        "contract_number": "",
        "contract_date": "",
        "contract_date_iso": "",
    }
    if not text:
        return result

    request_match = re.search(
        r"(заявка|спецификация)\s*№\s*([^,\s]+).*?по\s+договору\s+(.+?)\s+от\s+(\d{2}\.\d{2}\.\d{4})",
        text,
        re.IGNORECASE,
    )
    if request_match:
        result.update(
            {
                "request_type": request_match.group(1).capitalize(),
                "spec_number": normalize_text(request_match.group(2)),
                "base_contract": normalize_text(request_match.group(3)).strip(" №"),
                "contract_number": normalize_text(request_match.group(3)).strip(" №"),
                "contract_date": normalize_text(request_match.group(4)),
                "contract_date_iso": parse_any_date_to_iso(request_match.group(4)),
            }
        )
        return result

    contract_match = re.search(
        r"(?:договор|контракт)\s*№?\s*(.+?)\s+от\s+(\d{2}\.\d{2}\.\d{4}|\d{2}\.\d{2}\.\d{2})",
        text,
        re.IGNORECASE,
    )
    if contract_match:
        number = normalize_text(contract_match.group(1)).strip(" №")
        date = normalize_text(contract_match.group(2))
        result.update(
            {
                "base_contract": result["base_contract"] or number,
                "contract_number": number,
                "contract_date": date,
                "contract_date_iso": parse_any_date_to_iso(date),
            }
        )
        return result

    combined_match = re.search(r"\b([A-ZА-ЯЁ]{1,6}[-/A-ZА-ЯЁ0-9]*?)/(\d{1,6})\b", text, re.IGNORECASE)
    if combined_match:
        result.update(
            {
                "request_type": "Заявка",
                "spec_number": normalize_text(combined_match.group(2)),
                "base_contract": normalize_text(combined_match.group(1)),
                "contract_number": normalize_text(combined_match.group(1)),
            }
        )

    return result


def extract_invoice_number(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""

    patterns = [
        r"(?:счет[ауом]?|сч[её]т)\s*(?:покупателю|на\s+оплату)?\s*(?:№|N)?\s*([A-ZА-ЯЁ]{1,5}-\d{1,8}|\d{3,8})",
        r"(?:оплата|часть|по)\s+счет[ау]?\s*(?:№|N)?\s*([A-ZА-ЯЁ]{1,5}-\d{1,8}|\d{3,8})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return normalize_text(match.group(1)).strip(".,;")
    return ""


def doc_number_keys(value: object) -> set[str]:
    text = normalize_text(value)
    keys = {compact_key(text)} if text else set()
    if not text:
        return keys
    no_leading_zeroes = re.sub(r"(^|[-/])0+(\d)", r"\1\2", text)
    if no_leading_zeroes != text:
        keys.add(compact_key(no_leading_zeroes))
    digits = re.sub(r"\D+", "", text)
    has_letters = bool(re.search(r"[A-Za-zА-Яа-яЁё]", text))
    if len(digits) >= 3 and not has_letters:
        keys.add(digits)
        stripped = digits.lstrip("0")
        if len(stripped) >= 3:
            keys.add(stripped)
    return {key for key in keys if key}


def strong_doc_number_keys(value: object) -> set[str]:
    keys = set()
    for key in doc_number_keys(value):
        has_letters = bool(re.search(r"[a-zа-я]", key, re.IGNORECASE))
        if has_letters or len(key) >= 4:
            keys.add(key)
    return keys


def normalize_sum(value: object) -> float | None:
    raw = normalize_text(value)
    if not raw:
        return None
    cleaned = raw.replace("\u00a0", "").replace(" ", "").replace("₽", "").replace("$", "")
    if cleaned.count(",") > 0 and cleaned.count(".") > 0:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except Exception:
        return None


def sum_values(rows: list[dict[str, object]], key: str) -> float:
    total = 0.0
    for row in rows:
        value = normalize_sum(row.get(key))
        if value is not None:
            total += value
    return round(total, 2)


def display_date_from_iso(value: object) -> str:
    iso = parse_any_date_to_iso(value)
    if not iso:
        return normalize_text(value)
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return normalize_text(value)


def build_spec_operations_source(spec_id: int, alias: str = "oper") -> str:
    return f"""
{alias}.f_parenttype IN (2, 4)
  AND (
    CASE {alias}.f_parenttype
      WHEN 2 THEN {alias}.f_specid
      WHEN 4 THEN CAST({alias}4_specs.f_valstr AS SIGNED)
      ELSE NULL
    END
  ) = {spec_id}
"""


def build_spec_operations_ids_subquery(spec_id: int) -> str:
    return f"""
SELECT oper.f_id
FROM veda_spec_invoices oper
LEFT JOIN veda_categs oper4_specs
    ON oper4_specs.f_objectid = oper.f_id
   AND oper4_specs.f_ctgtype = 24
   AND oper4_specs.f_objecttype = 5
WHERE {build_spec_operations_source(spec_id, "oper")}
"""


def build_operations_query(spec_id: int) -> str:
    return f"""
SELECT DISTINCT
    oper.f_id AS oper_id,
    COALESCE(oper.f_num_oper, 999999) AS oper_num,
    COALESCE(oper_type.f_name, '') AS oper_type_name,
    COALESCE(oper.f_sum, 0) AS oper_sum,
    COALESCE(oper.f_val, 0) AS oper_val_id,
    COALESCE(NULLIF(val.f_dopprstr, ''), NULLIF(val.f_uslstr, ''), NULLIF(val.f_namedop, ''), val.f_name, '') AS oper_val_name_abbr,
    COALESCE(oper.f_invoiceid, 0) AS invoice_id,
    COALESCE(oper.f_parentid, 0) AS parent_oper_id,
    COALESCE(oper.f_dogid, 0) AS dog_id,
    COALESCE(dog.f_kod1c, '') AS dog_code1c,
    COALESCE(dog.f_dogname, '') AS dog_number,
    COALESCE(DATE_FORMAT(dog.f_dogdate, '%d.%m.%Y'), '') AS dog_date,
    COALESCE(oper.f_nds, 0) AS nds_id,
    COALESCE(nds.f_name, '') AS nds_name,
    COALESCE(oper.f_addnds, 0) AS add_nds_flag,
    COALESCE(oper.f_hnds, 0) AS hnds_flag,
    COALESCE(oper.f_isvozm, 0) AS reimbursement_id,
    COALESCE(vozm.f_name, '') AS reimbursement_name,
    COALESCE(oper.f_bdrarticle, 0) AS bdr_article_id,
    COALESCE(bdr.f_name, '') AS bdr_article_name,
    COALESCE(oper_type.f_c1doctype, 0) AS c1_doc_type_id,
    COALESCE(oper_type.f_zdoctype, 0) AS close_doc_type_id,
    COALESCE(get_paidsum(oper.f_id), 0) AS rp_paid_sum,
    COALESCE(get_expensessum(oper.f_id), 0) AS rp_expenses_sum,
    COALESCE(oper.f_outbuhperiod, 0) AS out_buh_period,
    COALESCE(get_realizsum(oper.f_id), 0) AS rp_realiz_sum,
    COALESCE(get_profit(oper.f_id), 0) AS rp_profit_sum
FROM veda_spec_invoices oper
LEFT JOIN veda_typeopers oper_type
    ON oper.f_idoper = oper_type.f_id
LEFT JOIN veda_spr val
    ON val.f_type = 4 AND val.f_num = oper.f_val
LEFT JOIN veda_spr nds
    ON nds.f_type = 10 AND nds.f_num = oper.f_nds
LEFT JOIN veda_spr vozm
    ON vozm.f_type = 2 AND vozm.f_num = oper.f_isvozm
LEFT JOIN veda_spr bdr
    ON bdr.f_type = 86 AND bdr.f_num = oper.f_bdrarticle
LEFT JOIN veda_dogs dog
    ON dog.f_id = oper.f_dogid
LEFT JOIN veda_categs oper4_specs
    ON oper4_specs.f_objectid = oper.f_id
   AND oper4_specs.f_ctgtype = 24
   AND oper4_specs.f_objecttype = 5
WHERE {build_spec_operations_source(spec_id, "oper")}
ORDER BY COALESCE(oper.f_num_oper, 999999), oper.f_id;
"""


def build_act_docs_query(spec_id: int) -> str:
    return f"""
SELECT DISTINCT
    'act' AS doc_kind,
    oper.f_id AS oper_id,
    COALESCE(oper.f_num_oper, 999999) AS oper_num,
    COALESCE(oper_type.f_name, '') AS oper_type_name,
    COALESCE(akt.f_id, 0) AS erp_doc_id,
    COALESCE(akt.f_kod1c, '') AS code1c,
    COALESCE(akt.f_num, '') AS doc_number,
    COALESCE(
        CASE
            WHEN akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(akt.f_dt1c, '%Y-%m-%d')
            ELSE DATE_FORMAT(akt.f_dt, '%Y-%m-%d')
        END,
        ''
    ) AS doc_date_iso,
    COALESCE(
        CASE
            WHEN akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(akt.f_dt1c, '%d.%m.%Y')
            ELSE DATE_FORMAT(akt.f_dt, '%d.%m.%Y')
        END,
        ''
    ) AS doc_date,
    COALESCE(akt.f_sum, 0) AS doc_sum,
    COALESCE(akt.f_val, 0) AS doc_currency_id,
    COALESCE(NULLIF(val.f_dopprstr, ''), NULLIF(val.f_uslstr, ''), NULLIF(val.f_namedop, ''), val.f_name, '') AS doc_currency,
    COALESCE(akt.f_type, 0) AS doc_type_id,
    COALESCE(akt_type.f_name, '') AS doc_type_name,
    COALESCE(akt.f_dogid, 0) AS dog_id,
    COALESCE(dog.f_kod1c, '') AS dog_code1c,
    COALESCE(dog.f_dogname, '') AS dog_number,
    COALESCE(oper.f_invoiceid, 0) AS invoice_id,
    COALESCE(invoice.f_num, '') AS invoice_number,
    COALESCE(DATE_FORMAT(invoice.f_dt, '%d.%m.%Y'), '') AS invoice_date,
    COALESCE(oper.f_isvozm, 0) AS reimbursement_id,
    COALESCE(vozm.f_name, '') AS reimbursement_name,
    COALESCE(oper.f_nds, 0) AS nds_id,
    COALESCE(nds.f_name, '') AS nds_name,
    0 AS detail_id,
    '' AS detail_name,
    0 AS detail_sum,
    COALESCE(main_akt.f_id, 0) AS main_erp_doc_id,
    COALESCE(NULLIF(main_akt.f_kod1c, ''), NULLIF(main_akt.f_num, ''), '') AS main_code1c,
    COALESCE(main_akt.f_num, '') AS main_number,
    COALESCE(
        CASE
            WHEN main_akt.f_id IS NULL THEN ''
            WHEN main_akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(main_akt.f_dt1c, '%Y-%m-%d')
            ELSE DATE_FORMAT(main_akt.f_dt, '%Y-%m-%d')
        END,
        ''
    ) AS main_date_iso,
    COALESCE(
        CASE
            WHEN main_akt.f_id IS NULL THEN ''
            WHEN main_akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(main_akt.f_dt1c, '%d.%m.%Y')
            ELSE DATE_FORMAT(main_akt.f_dt, '%d.%m.%Y')
        END,
        ''
    ) AS main_date,
    COALESCE(main_akt.f_sum, 0) AS main_sum,
    COALESCE(akt.f_status, 0) AS status_id,
    COALESCE(akt_status.f_name, '') AS status_name,
    CASE
        WHEN akt.f_status = 8 THEN 'posted_in_1c'
        WHEN akt.f_status = 7 THEN 'exported_not_posted'
        WHEN akt.f_status = 9 THEN 'deleted'
        WHEN COALESCE(NULLIF(akt.f_kod1c, ''), '') <> '' THEN 'has_1c_key'
        ELSE 'not_exported'
    END AS onec_export_state
FROM veda_spec_invoices oper
LEFT JOIN veda_typeopers oper_type
    ON oper.f_idoper = oper_type.f_id
LEFT JOIN veda_categs oper4_specs
    ON oper4_specs.f_objectid = oper.f_id
   AND oper4_specs.f_ctgtype = 24
   AND oper4_specs.f_objecttype = 5
JOIN veda_akts akt
    ON akt.f_operid = oper.f_id
LEFT JOIN veda_akts main_akt
    ON main_akt.f_id = akt.f_mainakt
   AND akt.f_mainakt > 0
LEFT JOIN veda_spr val
    ON val.f_type = 4 AND val.f_num = akt.f_val
LEFT JOIN veda_spr akt_type
    ON akt_type.f_type = 104 AND akt_type.f_num = akt.f_type
LEFT JOIN veda_spr akt_status
    ON akt_status.f_type = 12 AND akt_status.f_num = akt.f_status
LEFT JOIN veda_dogs dog
    ON dog.f_id = akt.f_dogid
LEFT JOIN veda_schets invoice
    ON invoice.f_id = oper.f_invoiceid
LEFT JOIN veda_spr vozm
    ON vozm.f_type = 2 AND vozm.f_num = oper.f_isvozm
LEFT JOIN veda_spr nds
    ON nds.f_type = 10 AND nds.f_num = oper.f_nds
WHERE {build_spec_operations_source(spec_id, "oper")}

UNION

SELECT DISTINCT
    'act' AS doc_kind,
    oper.f_id AS oper_id,
    COALESCE(oper.f_num_oper, 999999) AS oper_num,
    COALESCE(oper_type.f_name, '') AS oper_type_name,
    COALESCE(akt.f_id, 0) AS erp_doc_id,
    COALESCE(akt.f_kod1c, '') AS code1c,
    COALESCE(akt.f_num, '') AS doc_number,
    COALESCE(
        CASE
            WHEN akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(akt.f_dt1c, '%Y-%m-%d')
            ELSE DATE_FORMAT(akt.f_dt, '%Y-%m-%d')
        END,
        ''
    ) AS doc_date_iso,
    COALESCE(
        CASE
            WHEN akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(akt.f_dt1c, '%d.%m.%Y')
            ELSE DATE_FORMAT(akt.f_dt, '%d.%m.%Y')
        END,
        ''
    ) AS doc_date,
    COALESCE(akt.f_sum, 0) AS doc_sum,
    COALESCE(akt.f_val, 0) AS doc_currency_id,
    COALESCE(NULLIF(val.f_dopprstr, ''), NULLIF(val.f_uslstr, ''), NULLIF(val.f_namedop, ''), val.f_name, '') AS doc_currency,
    COALESCE(akt.f_type, 0) AS doc_type_id,
    COALESCE(akt_type.f_name, '') AS doc_type_name,
    COALESCE(akt.f_dogid, 0) AS dog_id,
    COALESCE(dog.f_kod1c, '') AS dog_code1c,
    COALESCE(dog.f_dogname, '') AS dog_number,
    COALESCE(oper.f_invoiceid, 0) AS invoice_id,
    COALESCE(invoice.f_num, '') AS invoice_number,
    COALESCE(DATE_FORMAT(invoice.f_dt, '%d.%m.%Y'), '') AS invoice_date,
    COALESCE(oper.f_isvozm, 0) AS reimbursement_id,
    COALESCE(vozm.f_name, '') AS reimbursement_name,
    COALESCE(oper.f_nds, 0) AS nds_id,
    COALESCE(nds.f_name, '') AS nds_name,
    COALESCE(akt_detail_oper.f_id, akt_detail.f_id, 0) AS detail_id,
    COALESCE(akt_detail.f_grnd, '') AS detail_name,
    COALESCE(
        CASE
            WHEN COALESCE(akt_detail_oper.f_sum, 0) > 0 THEN akt_detail_oper.f_sum
            WHEN COALESCE(akt_detail.f_nds, 0) <> 3 THEN akt_detail.f_sum + akt_detail.f_ndssum
            ELSE akt_detail.f_sum
        END,
        0
    ) AS detail_sum,
    COALESCE(main_akt.f_id, 0) AS main_erp_doc_id,
    COALESCE(NULLIF(main_akt.f_kod1c, ''), NULLIF(main_akt.f_num, ''), '') AS main_code1c,
    COALESCE(main_akt.f_num, '') AS main_number,
    COALESCE(
        CASE
            WHEN main_akt.f_id IS NULL THEN ''
            WHEN main_akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(main_akt.f_dt1c, '%Y-%m-%d')
            ELSE DATE_FORMAT(main_akt.f_dt, '%Y-%m-%d')
        END,
        ''
    ) AS main_date_iso,
    COALESCE(
        CASE
            WHEN main_akt.f_id IS NULL THEN ''
            WHEN main_akt.f_dt1c <> '0000-00-00' THEN DATE_FORMAT(main_akt.f_dt1c, '%d.%m.%Y')
            ELSE DATE_FORMAT(main_akt.f_dt, '%d.%m.%Y')
        END,
        ''
    ) AS main_date,
    COALESCE(main_akt.f_sum, 0) AS main_sum,
    COALESCE(akt.f_status, 0) AS status_id,
    COALESCE(akt_status.f_name, '') AS status_name,
    CASE
        WHEN akt.f_status = 8 THEN 'posted_in_1c'
        WHEN akt.f_status = 7 THEN 'exported_not_posted'
        WHEN akt.f_status = 9 THEN 'deleted'
        WHEN COALESCE(NULLIF(akt.f_kod1c, ''), '') <> '' THEN 'has_1c_key'
        ELSE 'not_exported'
    END AS onec_export_state
FROM veda_spec_invoices oper
LEFT JOIN veda_typeopers oper_type
    ON oper.f_idoper = oper_type.f_id
LEFT JOIN veda_categs oper4_specs
    ON oper4_specs.f_objectid = oper.f_id
   AND oper4_specs.f_ctgtype = 24
   AND oper4_specs.f_objecttype = 5
JOIN veda_akts_details_opers akt_detail_oper
    ON akt_detail_oper.f_operid = oper.f_id
JOIN veda_akts_details akt_detail
    ON akt_detail.f_id = akt_detail_oper.f_akts_detailsid
JOIN veda_akts akt
    ON akt.f_id = akt_detail.f_aktid
LEFT JOIN veda_akts main_akt
    ON main_akt.f_id = akt.f_mainakt
   AND akt.f_mainakt > 0
LEFT JOIN veda_spr val
    ON val.f_type = 4 AND val.f_num = akt.f_val
LEFT JOIN veda_spr akt_type
    ON akt_type.f_type = 104 AND akt_type.f_num = akt.f_type
LEFT JOIN veda_spr akt_status
    ON akt_status.f_type = 12 AND akt_status.f_num = akt.f_status
LEFT JOIN veda_dogs dog
    ON dog.f_id = akt.f_dogid
LEFT JOIN veda_schets invoice
    ON invoice.f_id = oper.f_invoiceid
LEFT JOIN veda_spr vozm
    ON vozm.f_type = 2 AND vozm.f_num = oper.f_isvozm
LEFT JOIN veda_spr nds
    ON nds.f_type = 10 AND nds.f_num = oper.f_nds
WHERE {build_spec_operations_source(spec_id, "oper")}
ORDER BY oper_num, oper_id, erp_doc_id, detail_id;
"""


def build_schet_docs_query(spec_id: int) -> str:
    oper_ids = build_spec_operations_ids_subquery(spec_id)
    return f"""
SELECT DISTINCT
    'schet' AS doc_kind,
    COALESCE(oper.f_id, linked_oper.f_id, 0) AS oper_id,
    COALESCE(oper.f_num_oper, linked_oper.f_num_oper, 999999) AS oper_num,
    COALESCE(oper_type.f_name, linked_oper_type.f_name, '') AS oper_type_name,
    COALESCE(schet.f_id, 0) AS erp_doc_id,
    COALESCE(schet.f_kod1c, '') AS code1c,
    COALESCE(schet.f_num, '') AS doc_number,
    COALESCE(DATE_FORMAT(schet.f_dt, '%Y-%m-%d'), '') AS doc_date_iso,
    COALESCE(DATE_FORMAT(schet.f_dt, '%d.%m.%Y'), '') AS doc_date,
    COALESCE(schet.f_sum, 0) AS doc_sum,
    COALESCE(schet.f_val, 0) AS doc_currency_id,
    COALESCE(NULLIF(val.f_dopprstr, ''), NULLIF(val.f_uslstr, ''), NULLIF(val.f_namedop, ''), val.f_name, '') AS doc_currency,
    COALESCE(schet.f_type, 0) AS doc_type_id,
    CASE
        WHEN schet.f_dogtype = 2 AND schet.f_dogid = {spec_id} AND schet.f_ismaininv = 1 THEN 'Счет покупателю'
        WHEN schet.f_maininv > 0 THEN 'Строка агрегирующего счета'
        WHEN schet.f_type = 1 THEN 'Счет поставщика'
        WHEN schet.f_type = 2 THEN 'Счет покупателю'
        ELSE 'Счет'
    END AS doc_type_name,
    COALESCE(schet.f_dogid, 0) AS dog_id,
    COALESCE(dog.f_kod1c, '') AS dog_code1c,
    COALESCE(dog.f_dogname, '') AS dog_number,
    COALESCE(schet.f_maininv, 0) AS invoice_id,
    COALESCE(main_invoice.f_num, '') AS invoice_number,
    COALESCE(DATE_FORMAT(main_invoice.f_dt, '%d.%m.%Y'), '') AS invoice_date,
    COALESCE(oper.f_isvozm, linked_oper.f_isvozm, 0) AS reimbursement_id,
    COALESCE(vozm.f_name, linked_vozm.f_name, '') AS reimbursement_name,
    COALESCE(schet.f_nds, oper.f_nds, linked_oper.f_nds, 0) AS nds_id,
    COALESCE(schet_nds.f_name, oper_nds.f_name, linked_oper_nds.f_name, '') AS nds_name,
    0 AS detail_id,
    COALESCE(schet.f_grnd, '') AS detail_name,
    0 AS detail_sum,
    0 AS main_erp_doc_id,
    '' AS main_code1c,
    '' AS main_number,
    '' AS main_date_iso,
    '' AS main_date,
    0 AS main_sum,
    COALESCE(schet.f_status, 0) AS status_id,
    COALESCE(schet_status.f_name, '') AS status_name,
    CASE
        WHEN schet.f_status = 8 THEN 'posted_in_1c'
        WHEN schet.f_status = 7 THEN 'exported_not_posted'
        WHEN schet.f_status = 9 THEN 'deleted'
        WHEN COALESCE(NULLIF(schet.f_kod1c, ''), '') <> '' THEN 'has_1c_key'
        ELSE 'not_exported'
    END AS onec_export_state
FROM veda_schets schet
LEFT JOIN veda_spec_invoices oper
    ON oper.f_id = schet.f_operid
   AND oper.f_id IN ({oper_ids})
LEFT JOIN veda_typeopers oper_type
    ON oper_type.f_id = oper.f_idoper
LEFT JOIN veda_spec_invoices linked_oper
    ON linked_oper.f_invoiceid = schet.f_id
   AND linked_oper.f_id IN ({oper_ids})
LEFT JOIN veda_typeopers linked_oper_type
    ON linked_oper_type.f_id = linked_oper.f_idoper
LEFT JOIN veda_spr val
    ON val.f_type = 4 AND val.f_num = schet.f_val
LEFT JOIN veda_dogs dog
    ON dog.f_id = schet.f_dogid
LEFT JOIN veda_schets main_invoice
    ON main_invoice.f_id = schet.f_maininv
LEFT JOIN veda_spr vozm
    ON vozm.f_type = 2 AND vozm.f_num = oper.f_isvozm
LEFT JOIN veda_spr linked_vozm
    ON linked_vozm.f_type = 2 AND linked_vozm.f_num = linked_oper.f_isvozm
LEFT JOIN veda_spr schet_nds
    ON schet_nds.f_type = 10 AND schet_nds.f_num = schet.f_nds
LEFT JOIN veda_spr oper_nds
    ON oper_nds.f_type = 10 AND oper_nds.f_num = oper.f_nds
LEFT JOIN veda_spr linked_oper_nds
    ON linked_oper_nds.f_type = 10 AND linked_oper_nds.f_num = linked_oper.f_nds
LEFT JOIN veda_spr schet_status
    ON schet_status.f_type = 12 AND schet_status.f_num = schet.f_status
WHERE schet.f_id > 0
  AND (
        schet.f_operid IN ({oper_ids})
        OR schet.f_id IN (
            SELECT si.f_invoiceid
            FROM veda_spec_invoices si
            LEFT JOIN veda_categs si4_specs
                ON si4_specs.f_objectid = si.f_id
               AND si4_specs.f_ctgtype = 24
               AND si4_specs.f_objecttype = 5
            WHERE {build_spec_operations_source(spec_id, "si")}
              AND si.f_invoiceid > 0
        )
        OR (schet.f_dogtype = 2 AND schet.f_dogid = {spec_id})
      )
ORDER BY oper_num, oper_id, erp_doc_id;
"""


def fetch_operations(spec_id: int) -> list[dict[str, object]]:
    operations: list[dict[str, object]] = []
    for row in run_mysql_tsv(build_operations_query(spec_id)):
        oper_id = int_at(row, 0)
        operations.append(
            {
                "oper_id": oper_id,
                "oper_num": int_at(row, 1, 999999),
                "oper_type_name": text_at(row, 2),
                "oper_sum": float_at(row, 3, 0.0),
                "oper_currency_id": int_at(row, 4, 0),
                "oper_currency": text_at(row, 5),
                "invoice_id": int_at(row, 6, 0),
                "parent_oper_id": int_at(row, 7, 0),
                "dog_id": int_at(row, 8, 0),
                "dog_code1c": text_at(row, 9),
                "dog_number": text_at(row, 10),
                "dog_date": text_at(row, 11),
                "nds_id": int_at(row, 12, 0),
                "nds_name": text_at(row, 13),
                "add_nds_flag": int_at(row, 14, 0),
                "hnds_flag": int_at(row, 15, 0),
                "reimbursement_id": int_at(row, 16, 0),
                "reimbursement_name": text_at(row, 17),
                "bdr_article_id": int_at(row, 18, 0),
                "bdr_article_name": text_at(row, 19),
                "c1_doc_type_id": int_at(row, 20, 0),
                "close_doc_type_id": int_at(row, 21, 0),
                "rp_paid_sum": float_at(row, 22, 0.0),
                "rp_expenses_sum": float_at(row, 23, 0.0),
                "out_buh_period": int_at(row, 24, 0),
                "rp_realiz_sum": float_at(row, 25, 0.0),
                "rp_profit_sum": float_at(row, 26, 0.0),
                "operation_title": text_at(row, 2).strip(),
                "operation_url": OPER_URL_TEMPLATE.format(oper_id=oper_id),
            }
        )
    return operations


def rows_to_erp_docs(rows: list[list[str]]) -> list[dict[str, object]]:
    docs: list[dict[str, object]] = []
    seen: set[tuple[str, int, int, int]] = set()
    for row in rows:
        doc_kind = text_at(row, 0)
        oper_id = int_at(row, 1)
        erp_doc_id = int_at(row, 4, 0)
        if erp_doc_id <= 0:
            continue
        detail_id = int_at(row, 24, 0)
        uniq = (doc_kind, oper_id, erp_doc_id, detail_id)
        if uniq in seen:
            continue
        seen.add(uniq)
        docs.append(
            {
                "doc_kind": doc_kind,
                "erp_doc_id": erp_doc_id,
                "oper_id": oper_id,
                "oper_num": int_at(row, 2, 999999),
                "oper_type_name": text_at(row, 3),
                "operation_title": text_at(row, 3).strip(),
                "operation_url": OPER_URL_TEMPLATE.format(oper_id=oper_id),
                "code1c": text_at(row, 5),
                "number": text_at(row, 6),
                "date_iso": text_at(row, 7),
                "date": text_at(row, 8),
                "sum": float_at(row, 9, 0.0),
                "currency_id": int_at(row, 10, 0),
                "currency": text_at(row, 11),
                "type_id": int_at(row, 12, 0),
                "type_name": text_at(row, 13),
                "dog_id": int_at(row, 14, 0),
                "dog_code1c": text_at(row, 15),
                "dog_number": text_at(row, 16),
                "invoice_id": int_at(row, 17, 0),
                "invoice_number": text_at(row, 18),
                "invoice_date": text_at(row, 19),
                "reimbursement_id": int_at(row, 20, 0),
                "reimbursement_name": text_at(row, 21),
                "nds_id": int_at(row, 22, 0),
                "nds_name": text_at(row, 23),
                "detail_id": detail_id,
                "detail_name": text_at(row, 25),
                "detail_sum": float_at(row, 26, 0.0),
                "main_erp_doc_id": int_at(row, 27, 0),
                "main_code1c": text_at(row, 28),
                "main_number": text_at(row, 29),
                "main_date_iso": text_at(row, 30),
                "main_date": text_at(row, 31),
                "main_sum": float_at(row, 32, 0.0),
                "status_id": int_at(row, 33, 0),
                "status_name": text_at(row, 34),
                "onec_export_state": text_at(row, 35),
            }
        )
    return docs


def fetch_erp_docs(spec_id: int) -> list[dict[str, object]]:
    docs: list[dict[str, object]] = []
    docs.extend(rows_to_erp_docs(run_mysql_tsv(build_act_docs_query(spec_id))))
    docs.extend(rows_to_erp_docs(run_mysql_tsv(build_schet_docs_query(spec_id))))
    return docs


def normalize_onec_docs(raw_docs: object) -> list[dict[str, object]]:
    docs = raw_docs if isinstance(raw_docs, list) else []
    normalized: list[dict[str, object]] = []
    for idx, item in enumerate(docs, start=1):
        if not isinstance(item, dict):
            continue
        code = normalize_text(item.get("code1c") or item.get("код1с") or item.get("kod1c"))
        number = normalize_text(item.get("number") or item.get("num") or item.get("номер"))
        date_iso = parse_any_date_to_iso(item.get("date") or item.get("doc_date") or item.get("дата"))
        amount = normalize_sum(item.get("sum") or item.get("amount") or item.get("сумма"))
        type_value = normalize_text(item.get("type") or item.get("doc_type") or item.get("тип"))

        normalized.append(
            {
                "row_num": idx,
                "code1c": code,
                "number": number,
                "number_norm": normalize_number(number),
                "date_iso": date_iso,
                "date": normalize_text(item.get("date") or item.get("doc_date") or item.get("дата")),
                "sum": amount,
                "type": type_value,
                "type_norm": normalize_number(type_value),
                "_used": False,
            }
        )
    return normalized


def load_onec_docs_from_connector(spec_id: int, scope: str = "specification", scope_id: int | None = None) -> tuple[list[dict[str, object]], str]:
    """
    1C connector stub for current stage.
    Future: replace with read-only REST calls by collected 1C codes.
    """
    mock_path = os.environ.get("RECON_ONEC_MOCK_FILE", "").strip()
    if not mock_path:
        return [], "unavailable"

    path = Path(mock_path)
    if not path.exists():
        return [], "unavailable"

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return [], "unavailable"

    rows = []
    if isinstance(raw, list):
        rows = raw
    elif isinstance(raw, dict):
        key_candidates = []
        if scope == "client":
            key_candidates.append(f"client:{sql_int(scope_id)}")
        key_candidates.extend([f"spec:{spec_id}", str(spec_id)])
        for key in key_candidates:
            if isinstance(raw.get(key), list):
                rows = raw[key]
                break
        if not rows and isinstance(raw.get("docs"), list):
            rows = raw["docs"]

    return normalize_onec_docs(rows), "available"


def worksheet_header_map(ws) -> tuple[int, dict[str, int]]:
    best_row = 1
    best_headers: dict[str, int] = {}
    best_score = 0
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True), start=1):
        headers: dict[str, int] = {}
        for idx, value in enumerate(row):
            header = normalize_text(value)
            if header:
                headers[header] = idx
        score = len(headers)
        lower_headers = {key.lower() for key in headers}
        if any(key in headers for key in ["Счет покупателю", "Поступление на расчетный счет", "Реализация (акт, накладная, УПД)", "Поступление (акт, накладная, УПД)", "Договор"]):
            score += 100
        if {"период", "документ", "аналитика дт", "аналитика кт", "дебет", "кредит"}.issubset(lower_headers):
            score += 100
        if score > best_score:
            best_row = ridx
            best_headers = headers
            best_score = score
    return best_row, best_headers


def header_value(row: tuple[object, ...], headers: dict[str, int], aliases: list[str]) -> object:
    for alias in aliases:
        idx = headers.get(alias)
        if idx is not None and idx < len(row):
            return row[idx]
    return ""


def infer_onec_workbook_kind(path: Path, headers: dict[str, int]) -> str:
    filename = path.name.lower()
    header_names = {key.lower() for key in headers}
    if {
        "поступление (акт, накладная, упд)",
        "комитент",
        "договор с комитентом",
        "идентификатор строки",
        "сумма",
    }.issubset(header_names):
        return "purchase_act_line_details"
    if (
        "поступление (акт, накладная, упд)" in header_names
        and "идентификатор строки" in header_names
        and "сумма" in header_names
        and ("счет затрат" in header_names or "субконто 1, договор комитента" in header_names)
    ):
        return "purchase_act_service_line_details"
    if "поступление (акт, накладная, упд)" in header_names and "детал" in filename:
        return "purchase_act_detail_unsupported"
    if (
        "карточка счета" in filename
        or {"период", "документ", "аналитика дт", "аналитика кт", "дебет", "кредит"}.issubset(header_names)
    ):
        return "account_card"
    if "счет покупателю" in header_names or "счета_на_оплату" in filename:
        return "invoice"
    if "поступление на расчетный счет" in header_names or "поступления_1с" in filename:
        return "bank_receipt"
    if "реализация (акт, накладная, упд)" in header_names or "реализация" in filename:
        return "sale_act"
    if "поступление (акт, накладная, упд)" in header_names or "поступление(акты)" in filename:
        return "purchase_act"
    if "договор" in header_names and ("заявки" in filename or "договор" in filename):
        return "contract"
    return "unknown"


def make_onec_doc(
    *,
    kind: str,
    source_file: str,
    source_sheet: str,
    source_row: int,
    title: object,
    number: object = "",
    date: object = "",
    amount: object = None,
    organization: object = "",
    counterparty: object = "",
    contract: object = "",
    invoice_number: object = "",
    comment: object = "",
    code1c: object = "",
    ref_id: object = "",
    extra: dict[str, object] | None = None,
) -> dict[str, object]:
    title_parts = parse_document_title(title)
    contract_parts = parse_contract_reference(contract)
    doc_number = normalize_text(number) or title_parts["number"]
    date_iso = parse_any_date_to_iso(date) or title_parts["date_iso"]
    date_display = display_date_from_iso(date_iso) if date_iso else (normalize_text(date) or title_parts["date"])
    invoice = normalize_text(invoice_number) or extract_invoice_number(comment) or (doc_number if kind == "invoice" else "")
    doc_code = normalize_text(code1c) or title_parts["number"] or doc_number

    row = {
        "row_num": source_row,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "kind": kind,
        "type": title_parts["doctype"] or {
            "invoice": "Счет покупателю",
            "bank_receipt": "Поступление на расчетный счет",
            "sale_act": "Реализация (акт, накладная, УПД)",
            "purchase_act": "Поступление (акт, накладная, УПД)",
            "contract": "Договор",
        }.get(kind, "Документ 1С"),
        "code1c": doc_code,
        "ref_id": normalize_text(ref_id),
        "number": doc_number,
        "number_norm": normalize_number(doc_number),
        "date": date_display,
        "date_iso": date_iso,
        "sum": normalize_sum(amount),
        "organization": normalize_text(organization),
        "counterparty": normalize_text(counterparty),
        "contract": normalize_text(contract),
        "contract_key": compact_key(contract),
        "request_type": contract_parts["request_type"],
        "spec_number": contract_parts["spec_number"],
        "base_contract": contract_parts["base_contract"],
        "contract_number": contract_parts["contract_number"],
        "contract_date": contract_parts["contract_date"],
        "contract_date_iso": contract_parts["contract_date_iso"],
        "invoice_number": invoice,
        "invoice_number_norm": normalize_number(invoice),
        "comment": normalize_text(comment),
        "_used": False,
    }
    if extra:
        row.update(extra)
    return row


def parse_onec_workbook(path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse 1C XLSX files")

    docs: list[dict[str, object]] = []
    summary = {
        "file": path.name,
        "sheets": [],
        "kind": "unknown",
        "rows_read": 0,
        "docs_count": 0,
        "warnings": [],
    }

    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row, headers = worksheet_header_map(ws)
        kind = infer_onec_workbook_kind(path, headers)
        if summary["kind"] == "unknown" and kind != "unknown":
            summary["kind"] = kind
        sheet_summary = {
            "sheet": ws.title,
            "kind": kind,
            "header_row": header_row,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "docs_count": 0,
        }
        summary["sheets"].append(sheet_summary)
        if kind in {"account_card", "purchase_act_line_details", "purchase_act_service_line_details", "purchase_act_detail_unsupported"}:
            if kind == "purchase_act_detail_unsupported":
                summary["warnings"].append(
                    f"{path.name}/{ws.title}: детальная выгрузка поступлений без колонок Комитент/Договор с комитентом; строки не участвуют в сверке"
                )
            continue
        if kind == "unknown":
            summary["warnings"].append(f"{path.name}/{ws.title}: не распознан тип листа")
            continue

        for source_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            summary["rows_read"] = int(summary["rows_read"]) + 1
            first = header_value(
                row,
                headers,
                ["Счет покупателю", "Поступление на расчетный счет", "Реализация (акт, накладная, УПД)", "Поступление (акт, накладная, УПД)", "Договор"],
            )
            if not normalize_text(first):
                continue
            if kind != "contract" and not parse_document_title(first)["number"]:
                continue

            if kind == "invoice":
                docs.append(
                    make_onec_doc(
                        kind=kind,
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=source_row,
                        title=first,
                        amount=header_value(row, headers, ["Сумма"]),
                        organization=header_value(row, headers, ["Организация"]),
                        counterparty=header_value(row, headers, ["Контрагент"]),
                        contract=header_value(row, headers, ["Договор"]),
                        invoice_number="",
                        comment=header_value(row, headers, ["Комментарий", "Документ-основание"]),
                        ref_id=header_value(row, headers, ["Ссылочный идентификатор"]),
                        extra={
                            "amount_role": "invoice",
                            "basis_document": normalize_text(header_value(row, headers, ["Документ-основание"])),
                            "currency": normalize_text(header_value(row, headers, ["Валюта"])),
                        },
                    )
                )
            elif kind == "bank_receipt":
                docs.append(
                    make_onec_doc(
                        kind=kind,
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=source_row,
                        title=first,
                        number=header_value(row, headers, ["Номер входящего документа"]),
                        date=header_value(row, headers, ["Дата входящего документа"]),
                        amount=header_value(row, headers, ["Сумма"]),
                        organization=header_value(row, headers, ["Организация"]),
                        counterparty=header_value(row, headers, ["Плательщик"]),
                        contract=header_value(row, headers, ["Договор", "Субконто Кт1"]),
                        invoice_number=header_value(row, headers, ["Документ основание"]),
                        comment=header_value(row, headers, ["Назначение платежа", "Комментарий"]),
                        extra={
                            "amount_role": "payment",
                            "operation_kind": normalize_text(header_value(row, headers, ["Вид операции"])),
                            "payment_purpose": normalize_text(header_value(row, headers, ["Назначение платежа"])),
                            "settlement_account": normalize_text(header_value(row, headers, ["Счет расчетов"])),
                        },
                    )
                )
            elif kind in {"sale_act", "purchase_act"}:
                docs.append(
                    make_onec_doc(
                        kind=kind,
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=source_row,
                        title=first,
                        amount=header_value(row, headers, ["Сумма"]),
                        organization=header_value(row, headers, ["Организация"]),
                        counterparty=header_value(row, headers, ["Контрагент"]),
                        contract=header_value(row, headers, ["Договор"]),
                        invoice_number=header_value(row, headers, ["Счет на оплату"]),
                        comment=header_value(row, headers, ["Комментарий", "Номер документа сторонней организации"]),
                        extra={
                            "amount_role": "act",
                            "operation_kind": normalize_text(header_value(row, headers, ["Вид операции"])),
                            "external_number": normalize_text(header_value(row, headers, ["Номер документа сторонней организации"])),
                            "external_date": display_date_from_iso(header_value(row, headers, ["Дата документа сторонней организации"])),
                            "settlement_account": normalize_text(header_value(row, headers, ["Счет учета расчетов с контрагентом"])),
                        },
                    )
                )
            elif kind == "contract":
                contract_parts = parse_contract_reference(first)
                docs.append(
                    make_onec_doc(
                        kind=kind,
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=source_row,
                        title="Договор",
                        number=header_value(row, headers, ["Номер договора"]) or contract_parts["contract_number"],
                        date=header_value(row, headers, ["Дата"]) or contract_parts["contract_date"],
                        amount=header_value(row, headers, ["Сумма"]),
                        organization=header_value(row, headers, ["Организация"]),
                        counterparty=header_value(row, headers, ["Контрагент"]),
                        contract=first,
                        comment=header_value(row, headers, ["Комментарий"]),
                        extra={
                            "amount_role": "contract",
                            "contract_kind": normalize_text(header_value(row, headers, ["Вид договора"])),
                            "currency": normalize_text(header_value(row, headers, ["Валюта"])),
                        },
                    )
                )

        sheet_summary["docs_count"] = len([doc for doc in docs if doc.get("source_file") == path.name and doc.get("source_sheet") == ws.title])

    summary["docs_count"] = len(docs)
    return docs, summary


def infer_account_scope_from_card(path: Path, worksheet_title: object = "", first_cell: object = "") -> str:
    for value in [first_cell, worksheet_title, path.name]:
        match = re.search(r"карточка\s+сч[её]та\s+([0-9.]+)", normalize_text(value), re.IGNORECASE)
        if match:
            return normalize_text(match.group(1))
    return ""


def parse_account_analytics(value: object) -> dict[str, object]:
    text = normalize_text(value)
    lines = [normalize_text(line) for line in re.split(r"\r?\n", text) if normalize_text(line)]
    contract_line = ""
    contract_idx = -1
    for idx, line in enumerate(lines):
        if re.search(r"(?:заявка|спецификация)\s*№", line, re.IGNORECASE) or re.search(r"договор\s*№", line, re.IGNORECASE):
            contract_line = line
            contract_idx = idx
            break

    contract_parts = parse_contract_reference(contract_line)
    counterparty = ""
    if contract_idx > 0:
        for line in reversed(lines[:contract_idx]):
            line_key = compact_key(line)
            if line_key and line_key not in {"опнск", "опмск", "опспб"} and not line.lower().startswith("оп "):
                counterparty = line
                break
    if not counterparty:
        for line in lines:
            line_key = compact_key(line)
            if line_key and line_key not in {"опнск", "опмск", "опспб"} and not line.lower().startswith("оп "):
                counterparty = line
                break

    analytic_doc = {}
    for line in lines:
        parsed = parse_document_title(line)
        if parsed.get("number"):
            analytic_doc = parsed
            break

    return {
        "text": text,
        "lines": lines,
        "counterparty": counterparty,
        "contract": contract_line,
        "contract_key": compact_key(contract_line),
        "request_type": contract_parts["request_type"],
        "spec_number": contract_parts["spec_number"],
        "base_contract": contract_parts["base_contract"],
        "contract_number": contract_parts["contract_number"],
        "contract_date": contract_parts["contract_date"],
        "contract_date_iso": contract_parts["contract_date_iso"],
        "analytic_doc_type": analytic_doc.get("doctype", ""),
        "analytic_doc_number": analytic_doc.get("number", ""),
        "analytic_doc_date_iso": analytic_doc.get("date_iso", ""),
    }


def make_account_movement(
    *,
    source_file: str,
    source_sheet: str,
    source_row: int,
    account_scope: str,
    period: object,
    registrar_text: object,
    debit_analytics: object,
    credit_analytics: object,
    debit_account: object,
    debit_amount: object,
    credit_account: object,
    credit_amount: object,
    balance_side: object,
    balance_amount: object,
) -> dict[str, object]:
    registrar = parse_document_title(registrar_text)
    debit_parts = parse_account_analytics(debit_analytics)
    credit_parts = parse_account_analytics(credit_analytics)
    period_iso = parse_any_date_to_iso(period) or registrar.get("date_iso", "")
    return {
        "row_num": source_row,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "kind": "account_movement",
        "account_scope": account_scope,
        "period": normalize_text(period),
        "period_iso": period_iso,
        "registrar_text": normalize_text(registrar_text),
        "registrar_type": registrar.get("doctype", ""),
        "registrar_number": registrar.get("number", ""),
        "registrar_date": registrar.get("date", ""),
        "registrar_date_iso": registrar.get("date_iso", "") or period_iso,
        "debit_account": normalize_text(debit_account),
        "debit_amount": normalize_sum(debit_amount),
        "debit_analytics": normalize_text(debit_analytics),
        "debit_counterparty": debit_parts["counterparty"],
        "debit_contract": debit_parts["contract"],
        "debit_spec_number": debit_parts["spec_number"],
        "debit_base_contract": debit_parts["base_contract"],
        "debit_contract_number": debit_parts["contract_number"],
        "debit_contract_date_iso": debit_parts["contract_date_iso"],
        "debit_analytic_doc_number": debit_parts["analytic_doc_number"],
        "debit_analytic_doc_date_iso": debit_parts["analytic_doc_date_iso"],
        "credit_account": normalize_text(credit_account),
        "credit_amount": normalize_sum(credit_amount),
        "credit_analytics": normalize_text(credit_analytics),
        "credit_counterparty": credit_parts["counterparty"],
        "credit_contract": credit_parts["contract"],
        "credit_spec_number": credit_parts["spec_number"],
        "credit_base_contract": credit_parts["base_contract"],
        "credit_contract_number": credit_parts["contract_number"],
        "credit_contract_date_iso": credit_parts["contract_date_iso"],
        "credit_analytic_doc_number": credit_parts["analytic_doc_number"],
        "credit_analytic_doc_date_iso": credit_parts["analytic_doc_date_iso"],
        "balance_side": normalize_text(balance_side),
        "balance_amount": normalize_sum(balance_amount),
    }


def parse_onec_account_movements(path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse 1C XLSX files")

    movements: list[dict[str, object]] = []
    summary = {
        "file": path.name,
        "sheets": [],
        "kind": "account_card",
        "rows_read": 0,
        "account_movements_count": 0,
        "warnings": [],
    }

    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row, headers = worksheet_header_map(ws)
        kind = infer_onec_workbook_kind(path, headers)
        first_cell = ""
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            first_cell = row[0] if row else ""
            break
        if kind != "account_card":
            continue

        debit_idx = headers.get("Дебет", 4)
        credit_idx = headers.get("Кредит", 7)
        balance_idx = headers.get("Текущее сальдо", 10)
        account_scope = infer_account_scope_from_card(path, ws.title, first_cell)
        sheet_count = 0
        for source_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            summary["rows_read"] = int(summary["rows_read"]) + 1
            period = row[0] if len(row) > 0 else ""
            registrar_text = row[1] if len(row) > 1 else ""
            if normalize_text(period).lower().startswith("сальдо"):
                continue
            debit_amount = row[debit_idx + 1] if len(row) > debit_idx + 1 else None
            credit_amount = row[credit_idx + 1] if len(row) > credit_idx + 1 else None
            if not normalize_text(registrar_text) and normalize_sum(debit_amount) is None and normalize_sum(credit_amount) is None:
                continue

            movement = make_account_movement(
                source_file=path.name,
                source_sheet=ws.title,
                source_row=source_row,
                account_scope=account_scope,
                period=period,
                registrar_text=registrar_text,
                debit_analytics=row[2] if len(row) > 2 else "",
                credit_analytics=row[3] if len(row) > 3 else "",
                debit_account=row[debit_idx] if len(row) > debit_idx else "",
                debit_amount=debit_amount,
                credit_account=row[credit_idx] if len(row) > credit_idx else "",
                credit_amount=credit_amount,
                balance_side=row[balance_idx] if len(row) > balance_idx else "",
                balance_amount=row[balance_idx + 1] if len(row) > balance_idx + 1 else None,
            )
            if not movement.get("registrar_number"):
                continue
            movements.append(movement)
            sheet_count += 1

        summary["sheets"].append(
            {
                "sheet": ws.title,
                "kind": kind,
                "header_row": header_row,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "account_scope": account_scope,
                "account_movements_count": sheet_count,
            }
        )

    summary["account_movements_count"] = len(movements)
    return movements, summary


def make_onec_document_line(
    *,
    source_file: str,
    source_sheet: str,
    source_row: int,
    title: object,
    nomenclature: object = "",
    content: object = "",
    vat_rate: object = "",
    commissioner: object = "",
    commissioner_contract: object = "",
    settlement_account: object = "",
    line_id: object = "",
    quantity: object = None,
    price: object = None,
    amount: object = None,
    vat_amount: object = None,
    detail_kind: str = "purchase_act_line_details",
    extra: dict[str, object] | None = None,
) -> dict[str, object]:
    title_parts = parse_document_title(title)
    contract_parts = parse_contract_reference(commissioner_contract)
    row = {
        "row_num": source_row,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "kind": detail_kind,
        "document_type": title_parts.get("doctype", ""),
        "document_code1c": title_parts.get("number", ""),
        "document_number": title_parts.get("number", ""),
        "document_date": title_parts.get("date", ""),
        "document_date_iso": title_parts.get("date_iso", ""),
        "nomenclature": normalize_text(nomenclature),
        "content": normalize_text(content),
        "vat_rate": normalize_text(vat_rate),
        "commissioner": normalize_text(commissioner),
        "commissioner_contract": normalize_text(commissioner_contract),
        "commissioner_contract_key": compact_key(commissioner_contract),
        "request_type": contract_parts["request_type"],
        "spec_number": contract_parts["spec_number"],
        "base_contract": contract_parts["base_contract"],
        "contract_number": contract_parts["contract_number"],
        "contract_date": contract_parts["contract_date"],
        "contract_date_iso": contract_parts["contract_date_iso"],
        "settlement_account": normalize_text(settlement_account),
        "line_id": normalize_text(line_id),
        "quantity": normalize_sum(quantity),
        "price": normalize_sum(price),
        "sum": normalize_sum(amount),
        "vat_amount": normalize_sum(vat_amount),
        "_used": False,
    }
    if extra:
        row.update(extra)
    return row


def parse_onec_document_lines(path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse 1C XLSX files")

    lines: list[dict[str, object]] = []
    summary = {
        "file": path.name,
        "sheets": [],
        "kind": "purchase_act_line_details",
        "rows_read": 0,
        "document_lines_count": 0,
        "warnings": [],
    }

    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row, headers = worksheet_header_map(ws)
        kind = infer_onec_workbook_kind(path, headers)
        if kind not in {"purchase_act_line_details", "purchase_act_service_line_details"}:
            continue
        if summary["kind"] == "purchase_act_line_details" and kind != "purchase_act_line_details":
            summary["kind"] = kind

        sheet_count = 0
        for source_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            summary["rows_read"] = int(summary["rows_read"]) + 1
            title = header_value(row, headers, ["Поступление (акт, накладная, УПД)"])
            title_parts = parse_document_title(title)
            if not title_parts.get("number"):
                continue
            line = make_onec_document_line(
                source_file=path.name,
                source_sheet=ws.title,
                source_row=source_row,
                title=title,
                nomenclature=header_value(row, headers, ["Номенклатура"]),
                content=header_value(row, headers, ["Содержание услуги", "Содержание"]),
                vat_rate=header_value(row, headers, ["% НДС", "НДС, %"]),
                commissioner=header_value(row, headers, ["Комитент"]),
                commissioner_contract=header_value(row, headers, ["Договор с комитентом"]),
                settlement_account=header_value(row, headers, ["Счет расчетов", "Счет учета расчетов"]),
                line_id=header_value(row, headers, ["Идентификатор строки", "ИД строки", "Строка"]),
                quantity=header_value(row, headers, ["Количество"]),
                price=header_value(row, headers, ["Цена"]),
                amount=header_value(row, headers, ["Сумма"]),
                vat_amount=header_value(row, headers, ["НДС", "НДС, сумма"]),
                detail_kind=kind,
                extra={
                    "cost_account": normalize_text(header_value(row, headers, ["Счет затрат"])),
                    "cost_department": normalize_text(header_value(row, headers, ["Подразделение затрат"])),
                    "subconto1": normalize_text(header_value(row, headers, ["Субконто 1, Договор комитента", "Субконто 1"])),
                    "subconto2": normalize_text(header_value(row, headers, ["Субконто 2, Договор комитента", "Субконто 2"])),
                    "subconto3": normalize_text(header_value(row, headers, ["Субконто 3"])),
                    "tax_cost_account": normalize_text(header_value(row, headers, ["Счет затрат НУ"])),
                    "tax_subconto1": normalize_text(header_value(row, headers, ["Субконто НУ 1"])),
                    "tax_subconto2": normalize_text(header_value(row, headers, ["Субконто НУ 2"])),
                    "tax_subconto3": normalize_text(header_value(row, headers, ["Субконто НУ 3"])),
                    "vat_account": normalize_text(header_value(row, headers, ["Счет НДС"])),
                    "vat_accounting_method": normalize_text(header_value(row, headers, ["Способ учета НДС"])),
                    "usn_expenses": normalize_text(header_value(row, headers, ["Расходы УСН"])),
                },
            )
            if line.get("sum") is None:
                continue
            lines.append(line)
            sheet_count += 1

        summary["sheets"].append(
            {
                "sheet": ws.title,
                "kind": kind,
                "header_row": header_row,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "document_lines_count": sheet_count,
            }
        )

    summary["document_lines_count"] = len(lines)
    return lines, summary


def discover_onec_files(paths: list[Path]) -> list[Path]:
    discovered: list[Path] = []
    for path in paths:
        if path.is_dir():
            discovered.extend(sorted([item for item in path.iterdir() if item.suffix.lower() == ".xlsx" and not item.name.startswith("~$")]))
        elif path.suffix.lower() == ".xlsx":
            discovered.append(path)
    return discovered


def onec_cache_key(files: list[Path]) -> str:
    parts = []
    for path in files:
        try:
            stat = path.stat()
            parts.append([str(path.resolve()), stat.st_size, int(stat.st_mtime)])
        except Exception:
            parts.append([str(path), 0, 0])
    return json.dumps(parts, ensure_ascii=False, sort_keys=True)


def build_onec_filter_context(snapshot: dict[str, object]) -> dict[str, object]:
    delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
    contracts = snapshot.get("contracts") if isinstance(snapshot.get("contracts"), list) else []
    schets = snapshot.get("schets") if isinstance(snapshot.get("schets"), list) else []
    akts = snapshot.get("akts") if isinstance(snapshot.get("akts"), list) else []
    payments = snapshot.get("payments") if isinstance(snapshot.get("payments"), list) else []

    spec_numbers = {normalize_text(delivery.get("spec_number"))}
    main_dog_number = normalize_text(delivery.get("main_dog_number"))
    spec_number = normalize_text(delivery.get("spec_number"))
    spec_contract_number = f"{main_dog_number}/{spec_number}" if main_dog_number and spec_number else ""
    base_contract_keys = {compact_key(main_dog_number)}
    delivery_contract_keys = {
        compact_key(spec_contract_number),
        compact_key(delivery.get("spec_buyer_code1c")),
        compact_key(delivery.get("spec_principal_code1c")),
    }
    contract_keys = {
        compact_key(main_dog_number),
        *delivery_contract_keys,
    }
    for contract in contracts:
        if isinstance(contract, dict):
            contract_keys.add(compact_key(contract.get("dog_number")))
            contract_keys.add(compact_key(contract.get("dog_code1c")))
            delivery_contract_keys.add(compact_key(contract.get("dog_number")))
            delivery_contract_keys.add(compact_key(contract.get("dog_code1c")))

    doc_keys: set[str] = set()
    doc_key_index: dict[str, set[tuple[str, str]]] = {}
    invoice_keys: set[str] = set()

    def remember_doc_keys(row: dict[str, object], erp_kind: str, *value_names: str) -> None:
        date_iso = parse_any_date_to_iso(
            row.get("date_iso")
            or row.get("date")
            or row.get("payment_date_iso")
            or row.get("payment_date")
            or row.get("main_date_iso")
            or row.get("main_date")
        )
        for value_name in value_names:
            for key in strong_doc_number_keys(row.get(value_name)):
                doc_keys.add(key)
                if date_iso:
                    doc_key_index.setdefault(key, set()).add((date_iso, erp_kind))

    for row in schets:
        if not isinstance(row, dict):
            continue
        remember_doc_keys(row, "invoice", "number", "code1c")
        for key in strong_doc_number_keys(row.get("invoice_number")):
            invoice_keys.add(key)
    for row in akts:
        if not isinstance(row, dict):
            continue
        remember_doc_keys(row, "act", "number", "code1c")
        remember_doc_keys(row, "act", "main_number", "main_code1c")
        for key in strong_doc_number_keys(row.get("invoice_number")):
            invoice_keys.add(key)
    for row in payments:
        if not isinstance(row, dict):
            continue
        remember_doc_keys(row, "payment", "code1c", "pp_number")
        for key in strong_doc_number_keys(row.get("invoice_number")):
            invoice_keys.add(key)

    return {
        "spec_numbers": {value for value in spec_numbers if value},
        "base_contract_keys": {value for value in base_contract_keys if value},
        "contract_keys": {value for value in contract_keys if value},
        "delivery_contract_keys": {value for value in delivery_contract_keys if value},
        "doc_keys": {value for value in doc_keys if value},
        "doc_key_index": {key: value for key, value in doc_key_index.items() if key and value},
        "invoice_keys": {value for value in invoice_keys if value},
    }


def onec_doc_match_reasons(doc: dict[str, object], context: dict[str, object]) -> list[str]:
    reasons: list[str] = []
    spec_numbers = context.get("spec_numbers") if isinstance(context.get("spec_numbers"), set) else set()
    base_contract_keys = context.get("base_contract_keys") if isinstance(context.get("base_contract_keys"), set) else set()
    delivery_contract_keys = context.get("delivery_contract_keys") if isinstance(context.get("delivery_contract_keys"), set) else set()
    doc_keys = context.get("doc_keys") if isinstance(context.get("doc_keys"), set) else set()
    doc_key_index = context.get("doc_key_index") if isinstance(context.get("doc_key_index"), dict) else {}
    invoice_keys = context.get("invoice_keys") if isinstance(context.get("invoice_keys"), set) else set()

    is_contract = normalize_text(doc.get("kind")) == "contract"
    doc_base_contract_key = compact_key(doc.get("base_contract"))
    doc_contract_number_key = compact_key(doc.get("contract_number"))
    doc_spec_number = normalize_text(doc.get("spec_number"))
    base_contract_matched = doc_base_contract_key in base_contract_keys or doc_contract_number_key in base_contract_keys
    delivery_contract_matched = doc_base_contract_key in delivery_contract_keys or doc_contract_number_key in delivery_contract_keys
    spec_contract_matched = doc_spec_number in spec_numbers and base_contract_matched
    contract_text_key = normalize_text(doc.get("contract_key"))
    contract_text_matched = contract_text_key and any(key and key in contract_text_key for key in delivery_contract_keys)
    has_contract_context = bool(
        doc_spec_number
        or normalize_text(doc.get("base_contract"))
        or normalize_text(doc.get("contract_number"))
        or contract_text_key
    )
    has_structured_contract_context = bool(
        doc_spec_number
        or normalize_text(doc.get("base_contract"))
        or normalize_text(doc.get("contract_number"))
    )
    if doc_spec_number:
        if spec_contract_matched or delivery_contract_matched:
            reasons.append("contract")
            reasons.append("spec_number")
    elif delivery_contract_matched:
        reasons.append("contract")
    elif is_contract and base_contract_matched:
        reasons.append("base_contract")

    if contract_text_matched and (not doc_spec_number or doc_spec_number in spec_numbers):
        reasons.append("contract_text")

    context_matches_contract = any(reason in reasons for reason in ("contract", "spec_number", "contract_text"))
    conflicting_spec_context = bool(doc_spec_number and doc_spec_number not in spec_numbers)
    allow_document_key_match = not conflicting_spec_context and (not has_contract_context or context_matches_contract or not is_contract)
    onec_doc_keys = strong_doc_number_keys(doc.get("code1c")) | strong_doc_number_keys(doc.get("number"))
    onec_date = parse_any_date_to_iso(doc.get("date_iso") or doc.get("date"))
    onec_kind = normalize_text(doc.get("kind"))
    indexed_document_match = False
    if onec_date and onec_kind:
        for key in onec_doc_keys:
            for erp_date, erp_kind in doc_key_index.get(key, set()):
                if erp_date == onec_date and onec_kind_compatible(erp_kind, onec_kind):
                    indexed_document_match = True
                    break
            if indexed_document_match:
                break
    fallback_document_match = not doc_key_index and bool(doc_keys.intersection(onec_doc_keys))
    if (
        allow_document_key_match
        and doc.get("kind") != "contract"
        and (indexed_document_match or fallback_document_match)
    ):
        reasons.append("document_number")
    if (
        not conflicting_spec_context
        and (not has_structured_contract_context or context_matches_contract)
        and doc.get("kind") != "contract"
        and invoice_keys.intersection(strong_doc_number_keys(doc.get("invoice_number")))
    ):
        reasons.append("invoice_number")

    return list(dict.fromkeys(reasons))


def filter_onec_docs_for_snapshot(docs: list[dict[str, object]], snapshot: dict[str, object]) -> tuple[list[dict[str, object]], dict[str, object]]:
    context = build_onec_filter_context(snapshot)
    filtered: list[dict[str, object]] = []
    for doc in docs:
        reasons = onec_doc_match_reasons(doc, context)
        if reasons:
            copied = dict(doc)
            copied["match_reasons"] = reasons
            filtered.append(copied)

    if not any(context.values()):
        filtered = [dict(doc, match_reasons=["no_filter_context"]) for doc in docs]

    return filtered, {
        "spec_numbers": sorted(context.get("spec_numbers", [])),
        "contract_keys_count": len(context.get("contract_keys", [])),
        "doc_keys_count": len(context.get("doc_keys", [])),
        "doc_key_signatures_count": sum(len(value) for value in context.get("doc_key_index", {}).values())
        if isinstance(context.get("doc_key_index"), dict)
        else 0,
        "invoice_keys_count": len(context.get("invoice_keys", [])),
        "filtered_count": len(filtered),
        "total_count": len(docs),
    }


def account_movement_matches_context(movement: dict[str, object], context: dict[str, object]) -> bool:
    spec_numbers = context.get("spec_numbers") if isinstance(context.get("spec_numbers"), set) else set()
    base_contract_keys = context.get("base_contract_keys") if isinstance(context.get("base_contract_keys"), set) else set()
    delivery_contract_keys = context.get("delivery_contract_keys") if isinstance(context.get("delivery_contract_keys"), set) else set()
    doc_key_index = context.get("doc_key_index") if isinstance(context.get("doc_key_index"), dict) else {}

    for side in ["debit", "credit"]:
        side_spec = normalize_text(movement.get(f"{side}_spec_number"))
        side_base_key = compact_key(movement.get(f"{side}_base_contract") or movement.get(f"{side}_contract_number"))
        side_contract_key = compact_key(movement.get(f"{side}_contract"))
        if side_spec and side_spec in spec_numbers and side_base_key in base_contract_keys:
            return True
        if side_contract_key and any(key and key in side_contract_key for key in delivery_contract_keys):
            return True

    movement_date = parse_any_date_to_iso(movement.get("registrar_date_iso") or movement.get("period_iso"))
    movement_keys = strong_doc_number_keys(movement.get("registrar_number")) | doc_number_keys(movement.get("registrar_text"))
    if movement_date and movement_keys:
        for key in movement_keys:
            for erp_date, erp_kind in doc_key_index.get(key, set()):
                if erp_date == movement_date and erp_kind == "payment":
                    return True

    return False


def filter_onec_account_movements_for_snapshot(movements: list[dict[str, object]], snapshot: dict[str, object]) -> list[dict[str, object]]:
    if not movements:
        return []
    context = build_onec_filter_context(snapshot)
    return [dict(movement) for movement in movements if account_movement_matches_context(movement, context)]


def document_line_matches_context(line: dict[str, object], context: dict[str, object]) -> bool:
    spec_numbers = context.get("spec_numbers") if isinstance(context.get("spec_numbers"), set) else set()
    base_contract_keys = context.get("base_contract_keys") if isinstance(context.get("base_contract_keys"), set) else set()
    delivery_contract_keys = context.get("delivery_contract_keys") if isinstance(context.get("delivery_contract_keys"), set) else set()
    doc_key_index = context.get("doc_key_index") if isinstance(context.get("doc_key_index"), dict) else {}

    line_spec = normalize_text(line.get("spec_number"))
    line_base_key = compact_key(line.get("base_contract") or line.get("contract_number"))
    line_contract_key = compact_key(line.get("commissioner_contract"))
    if line_spec and line_spec in spec_numbers and line_base_key in base_contract_keys:
        return True
    if line_contract_key and any(key and key in line_contract_key for key in delivery_contract_keys):
        return True

    line_date = parse_any_date_to_iso(line.get("document_date_iso") or line.get("document_date"))
    line_keys = strong_doc_number_keys(line.get("document_code1c")) | strong_doc_number_keys(line.get("document_number"))
    if line_date and line_keys:
        for key in line_keys:
            for erp_date, erp_kind in doc_key_index.get(key, set()):
                if erp_date == line_date and erp_kind == "act":
                    return True
    return False


def filter_onec_document_lines_for_snapshot(lines: list[dict[str, object]], snapshot: dict[str, object]) -> list[dict[str, object]]:
    if not lines:
        return []
    context = build_onec_filter_context(snapshot)
    return [dict(line) for line in lines if document_line_matches_context(line, context)]


def parse_onec_sources(paths: list[Path], snapshot: dict[str, object] | None = None) -> dict[str, object]:
    files = discover_onec_files(paths)
    cache_key = onec_cache_key(files)
    cached = ONEC_PARSE_CACHE.get(cache_key)
    if cached:
        docs = cached.get("docs", []) if isinstance(cached.get("docs"), list) else []
        account_movements = cached.get("account_movements", []) if isinstance(cached.get("account_movements"), list) else []
        document_lines = cached.get("document_lines", []) if isinstance(cached.get("document_lines"), list) else []
        file_summaries = cached.get("files", []) if isinstance(cached.get("files"), list) else []
        warnings = cached.get("warnings", []) if isinstance(cached.get("warnings"), list) else []
        cache_hit = True
    else:
        docs = []
        account_movements = []
        document_lines = []
        file_summaries = []
        warnings = []
        for path in files:
            try:
                parsed_docs, summary = parse_onec_workbook(path)
                docs.extend(parsed_docs)
                if summary.get("kind") == "account_card":
                    parsed_movements, movement_summary = parse_onec_account_movements(path)
                    account_movements.extend(parsed_movements)
                    summary["account_movements_count"] = len(parsed_movements)
                    summary["movement_sheets"] = movement_summary.get("sheets", [])
                    warnings.extend(movement_summary.get("warnings", []))
                if summary.get("kind") in {"purchase_act_line_details", "purchase_act_service_line_details"}:
                    parsed_lines, line_summary = parse_onec_document_lines(path)
                    document_lines.extend(parsed_lines)
                    summary["document_lines_count"] = len(parsed_lines)
                    summary["line_sheets"] = line_summary.get("sheets", [])
                    warnings.extend(line_summary.get("warnings", []))
                file_summaries.append(summary)
                warnings.extend(summary.get("warnings", []))
            except Exception as exc:
                warnings.append(f"{path.name}: {exc}")
        ONEC_PARSE_CACHE.clear()
        ONEC_PARSE_CACHE[cache_key] = {
            "docs": docs,
            "account_movements": account_movements,
            "document_lines": document_lines,
            "files": file_summaries,
            "warnings": warnings,
        }
        cache_hit = False

    filtered_docs = docs
    filtered_movements = account_movements
    filtered_document_lines = document_lines
    filter_summary: dict[str, object] = {"filtered_count": len(docs), "total_count": len(docs)}
    if snapshot is not None:
        filtered_docs, filter_summary = filter_onec_docs_for_snapshot(docs, snapshot)
        filtered_movements = filter_onec_account_movements_for_snapshot(account_movements, snapshot)
        filtered_document_lines = filter_onec_document_lines_for_snapshot(document_lines, snapshot)

    by_kind: dict[str, int] = {}
    for doc in filtered_docs:
        kind = normalize_text(doc.get("kind")) or "unknown"
        by_kind[kind] = by_kind.get(kind, 0) + 1

    contract_pairs = sorted(
        {
            f"{doc.get('base_contract') or doc.get('contract_number')}/{doc.get('spec_number')}".strip("/")
            for doc in filtered_docs
            if doc.get("base_contract") or doc.get("contract_number") or doc.get("spec_number")
        }
    )

    return {
        "files": file_summaries,
        "warnings": warnings,
        "all_docs_count": len(docs),
        "docs_count": len(filtered_docs),
        "all_account_movements_count": len(account_movements),
        "account_movements_count": len(filtered_movements),
        "all_document_lines_count": len(document_lines),
        "document_lines_count": len(filtered_document_lines),
        "cache_hit": cache_hit,
        "by_kind": by_kind,
        "contract_pairs": contract_pairs[:200],
        "filter": filter_summary,
        "docs": filtered_docs,
        "account_movements": filtered_movements,
        "document_lines": filtered_document_lines,
    }


def postgres_connect():
    if psycopg2 is None:
        raise RuntimeError("psycopg2 is not installed; PostgreSQL cache is unavailable")
    if not ONEC_PG_DSN:
        raise RuntimeError("RECON_ONEC_PG_DSN/RECON_PG_DSN is not configured")
    return psycopg2.connect(ONEC_PG_DSN)


def ensure_onec_pg_schema(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
CREATE TABLE IF NOT EXISTS onec_source_files (
    drive_file_id text PRIMARY KEY,
    drive_file_name text NOT NULL,
    drive_path text NOT NULL DEFAULT '',
    mime_type text NOT NULL DEFAULT '',
    modified_time timestamptz NULL,
    file_size bigint NULL,
    docs_count integer NOT NULL DEFAULT 0,
    account_movements_count integer NOT NULL DEFAULT 0,
    document_lines_count integer NOT NULL DEFAULT 0,
    summary_json jsonb NOT NULL DEFAULT '{}'::jsonb,
    warnings_json jsonb NOT NULL DEFAULT '[]'::jsonb,
    synced_at timestamptz NOT NULL DEFAULT now()
);

CREATE TABLE IF NOT EXISTS onec_documents (
    id bigserial PRIMARY KEY,
    drive_file_id text NOT NULL REFERENCES onec_source_files(drive_file_id) ON DELETE CASCADE,
    drive_file_name text NOT NULL,
    drive_path text NOT NULL DEFAULT '',
    source_sheet text NOT NULL DEFAULT '',
    source_row integer NOT NULL DEFAULT 0,
    kind text NOT NULL DEFAULT '',
    code1c text NOT NULL DEFAULT '',
    number text NOT NULL DEFAULT '',
    number_norm text NOT NULL DEFAULT '',
    date_iso date NULL,
    amount numeric(18, 2) NULL,
    doc_type text NOT NULL DEFAULT '',
    organization text NOT NULL DEFAULT '',
    counterparty text NOT NULL DEFAULT '',
    contract text NOT NULL DEFAULT '',
    contract_key text NOT NULL DEFAULT '',
    spec_number text NOT NULL DEFAULT '',
    base_contract text NOT NULL DEFAULT '',
    contract_number text NOT NULL DEFAULT '',
    invoice_number text NOT NULL DEFAULT '',
    invoice_number_norm text NOT NULL DEFAULT '',
    doc_json jsonb NOT NULL,
    synced_at timestamptz NOT NULL DEFAULT now()
);

CREATE INDEX IF NOT EXISTS idx_onec_documents_kind_code_date
    ON onec_documents (kind, code1c, date_iso);
CREATE INDEX IF NOT EXISTS idx_onec_documents_contract
    ON onec_documents (base_contract, spec_number, contract_number);
CREATE INDEX IF NOT EXISTS idx_onec_documents_invoice
    ON onec_documents (invoice_number_norm);
CREATE INDEX IF NOT EXISTS idx_onec_documents_doc_json_gin
    ON onec_documents USING gin (doc_json);

CREATE TABLE IF NOT EXISTS onec_account_movements (
    id bigserial PRIMARY KEY,
    drive_file_id text NOT NULL REFERENCES onec_source_files(drive_file_id) ON DELETE CASCADE,
    drive_file_name text NOT NULL,
    drive_path text NOT NULL DEFAULT '',
    source_sheet text NOT NULL DEFAULT '',
    source_row integer NOT NULL DEFAULT 0,
    account_scope text NOT NULL DEFAULT '',
    period_iso date NULL,
    period_text text NOT NULL DEFAULT '',
    registrar_type text NOT NULL DEFAULT '',
    registrar_number text NOT NULL DEFAULT '',
    registrar_date_iso date NULL,
    registrar_text text NOT NULL DEFAULT '',
    debit_account text NOT NULL DEFAULT '',
    debit_amount numeric(18, 2) NULL,
    debit_analytics text NOT NULL DEFAULT '',
    debit_counterparty text NOT NULL DEFAULT '',
    debit_contract text NOT NULL DEFAULT '',
    debit_spec_number text NOT NULL DEFAULT '',
    debit_base_contract text NOT NULL DEFAULT '',
    debit_contract_number text NOT NULL DEFAULT '',
    credit_account text NOT NULL DEFAULT '',
    credit_amount numeric(18, 2) NULL,
    credit_analytics text NOT NULL DEFAULT '',
    credit_counterparty text NOT NULL DEFAULT '',
    credit_contract text NOT NULL DEFAULT '',
    credit_spec_number text NOT NULL DEFAULT '',
    credit_base_contract text NOT NULL DEFAULT '',
    credit_contract_number text NOT NULL DEFAULT '',
    balance_side text NOT NULL DEFAULT '',
    balance_amount numeric(18, 2) NULL,
    movement_json jsonb NOT NULL,
    synced_at timestamptz NOT NULL DEFAULT now()
);

CREATE INDEX IF NOT EXISTS idx_onec_account_movements_registrar
    ON onec_account_movements (registrar_number, registrar_date_iso);
CREATE INDEX IF NOT EXISTS idx_onec_account_movements_debit_contract
    ON onec_account_movements (debit_base_contract, debit_spec_number);
CREATE INDEX IF NOT EXISTS idx_onec_account_movements_credit_contract
    ON onec_account_movements (credit_base_contract, credit_spec_number);
CREATE INDEX IF NOT EXISTS idx_onec_account_movements_json_gin
    ON onec_account_movements USING gin (movement_json);

CREATE TABLE IF NOT EXISTS onec_document_lines (
    id bigserial PRIMARY KEY,
    drive_file_id text NOT NULL REFERENCES onec_source_files(drive_file_id) ON DELETE CASCADE,
    drive_file_name text NOT NULL,
    drive_path text NOT NULL DEFAULT '',
    source_sheet text NOT NULL DEFAULT '',
    source_row integer NOT NULL DEFAULT 0,
    kind text NOT NULL DEFAULT '',
    document_type text NOT NULL DEFAULT '',
    document_code1c text NOT NULL DEFAULT '',
    document_number text NOT NULL DEFAULT '',
    document_date_iso date NULL,
    nomenclature text NOT NULL DEFAULT '',
    content text NOT NULL DEFAULT '',
    vat_rate text NOT NULL DEFAULT '',
    commissioner text NOT NULL DEFAULT '',
    commissioner_contract text NOT NULL DEFAULT '',
    commissioner_contract_key text NOT NULL DEFAULT '',
    spec_number text NOT NULL DEFAULT '',
    base_contract text NOT NULL DEFAULT '',
    contract_number text NOT NULL DEFAULT '',
    contract_date_iso date NULL,
    settlement_account text NOT NULL DEFAULT '',
    line_id text NOT NULL DEFAULT '',
    amount numeric(18, 2) NULL,
    vat_amount numeric(18, 2) NULL,
    line_json jsonb NOT NULL,
    synced_at timestamptz NOT NULL DEFAULT now()
);

CREATE INDEX IF NOT EXISTS idx_onec_document_lines_doc
    ON onec_document_lines (document_code1c, document_date_iso);
CREATE INDEX IF NOT EXISTS idx_onec_document_lines_contract
    ON onec_document_lines (base_contract, spec_number, contract_number);
CREATE INDEX IF NOT EXISTS idx_onec_document_lines_line_id
    ON onec_document_lines (line_id);
CREATE INDEX IF NOT EXISTS idx_onec_document_lines_json_gin
    ON onec_document_lines USING gin (line_json);
"""
        )
        cur.execute("ALTER TABLE onec_source_files ADD COLUMN IF NOT EXISTS account_movements_count integer NOT NULL DEFAULT 0;")
        cur.execute("ALTER TABLE onec_source_files ADD COLUMN IF NOT EXISTS document_lines_count integer NOT NULL DEFAULT 0;")


def pg_date_or_none(value: object) -> str | None:
    iso = parse_any_date_to_iso(value)
    return iso if iso else None


def pg_int_or_zero(value: object) -> int:
    try:
        return int(value or 0)
    except Exception:
        return 0


def clean_onec_doc_for_storage(doc: dict[str, object]) -> dict[str, object]:
    stored = dict(doc)
    stored.pop("_used", None)
    return stored


def insert_onec_file_to_postgres(
    conn,
    file_meta: dict[str, object],
    docs: list[dict[str, object]],
    summary: dict[str, object],
    account_movements: list[dict[str, object]] | None = None,
    document_lines: list[dict[str, object]] | None = None,
) -> None:
    account_movements = account_movements or []
    document_lines = document_lines or []
    warnings = summary.get("warnings") if isinstance(summary.get("warnings"), list) else []
    file_id = normalize_text(file_meta.get("id"))
    file_name = normalize_text(file_meta.get("name"))
    drive_path = normalize_text(file_meta.get("drive_path") or file_name)
    mime_type = normalize_text(file_meta.get("mimeType"))
    modified_time = normalize_text(file_meta.get("modifiedTime")) or None
    file_size = pg_int_or_zero(file_meta.get("size")) or None

    with conn.cursor() as cur:
        cur.execute(
            """
INSERT INTO onec_source_files
    (drive_file_id, drive_file_name, drive_path, mime_type, modified_time, file_size, docs_count, account_movements_count, document_lines_count, summary_json, warnings_json, synced_at)
VALUES
    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, now())
ON CONFLICT (drive_file_id) DO UPDATE SET
    drive_file_name = EXCLUDED.drive_file_name,
    drive_path = EXCLUDED.drive_path,
    mime_type = EXCLUDED.mime_type,
    modified_time = EXCLUDED.modified_time,
    file_size = EXCLUDED.file_size,
    docs_count = EXCLUDED.docs_count,
    account_movements_count = EXCLUDED.account_movements_count,
    document_lines_count = EXCLUDED.document_lines_count,
    summary_json = EXCLUDED.summary_json,
    warnings_json = EXCLUDED.warnings_json,
    synced_at = now();
""",
            (
                file_id,
                file_name,
                drive_path,
                mime_type,
                modified_time,
                file_size,
                len(docs),
                len(account_movements),
                len(document_lines),
                json.dumps(summary, ensure_ascii=False, default=str),
                json.dumps(warnings, ensure_ascii=False, default=str),
            ),
        )
        cur.execute("DELETE FROM onec_documents WHERE drive_file_id = %s", (file_id,))
        cur.execute("DELETE FROM onec_account_movements WHERE drive_file_id = %s", (file_id,))
        cur.execute("DELETE FROM onec_document_lines WHERE drive_file_id = %s", (file_id,))

        insert_sql = """
INSERT INTO onec_documents
    (drive_file_id, drive_file_name, drive_path, source_sheet, source_row, kind, code1c, number,
     number_norm, date_iso, amount, doc_type, organization, counterparty, contract, contract_key,
     spec_number, base_contract, contract_number, invoice_number, invoice_number_norm, doc_json, synced_at)
VALUES
    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, now());
"""
        for doc in docs:
            stored = clean_onec_doc_for_storage(doc)
            cur.execute(
                insert_sql,
                (
                    file_id,
                    file_name,
                    drive_path,
                    normalize_text(stored.get("source_sheet")),
                    pg_int_or_zero(stored.get("source_row") or stored.get("row_num")),
                    normalize_text(stored.get("kind")),
                    normalize_text(stored.get("code1c")),
                    normalize_text(stored.get("number")),
                    normalize_text(stored.get("number_norm")),
                    pg_date_or_none(stored.get("date_iso") or stored.get("date")),
                    normalize_sum(stored.get("sum")),
                    normalize_text(stored.get("type")),
                    normalize_text(stored.get("organization")),
                    normalize_text(stored.get("counterparty")),
                    normalize_text(stored.get("contract")),
                    normalize_text(stored.get("contract_key")),
                    normalize_text(stored.get("spec_number")),
                    normalize_text(stored.get("base_contract")),
                    normalize_text(stored.get("contract_number")),
                    normalize_text(stored.get("invoice_number")),
                    normalize_text(stored.get("invoice_number_norm")),
                    json.dumps(stored, ensure_ascii=False, default=str),
                ),
            )

        movement_insert_sql = """
INSERT INTO onec_account_movements
    (drive_file_id, drive_file_name, drive_path, source_sheet, source_row, account_scope,
     period_iso, period_text, registrar_type, registrar_number, registrar_date_iso, registrar_text,
     debit_account, debit_amount, debit_analytics, debit_counterparty, debit_contract,
     debit_spec_number, debit_base_contract, debit_contract_number,
     credit_account, credit_amount, credit_analytics, credit_counterparty, credit_contract,
     credit_spec_number, credit_base_contract, credit_contract_number,
     balance_side, balance_amount, movement_json, synced_at)
VALUES
    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
     %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, now());
"""
        movement_batch = []
        for movement in account_movements:
            stored = dict(movement)
            movement_batch.append(
                (
                    file_id,
                    file_name,
                    drive_path,
                    normalize_text(stored.get("source_sheet")),
                    pg_int_or_zero(stored.get("source_row") or stored.get("row_num")),
                    normalize_text(stored.get("account_scope")),
                    pg_date_or_none(stored.get("period_iso") or stored.get("period")),
                    normalize_text(stored.get("period")),
                    normalize_text(stored.get("registrar_type")),
                    normalize_text(stored.get("registrar_number")),
                    pg_date_or_none(stored.get("registrar_date_iso") or stored.get("registrar_date")),
                    normalize_text(stored.get("registrar_text")),
                    normalize_text(stored.get("debit_account")),
                    normalize_sum(stored.get("debit_amount")),
                    normalize_text(stored.get("debit_analytics")),
                    normalize_text(stored.get("debit_counterparty")),
                    normalize_text(stored.get("debit_contract")),
                    normalize_text(stored.get("debit_spec_number")),
                    normalize_text(stored.get("debit_base_contract")),
                    normalize_text(stored.get("debit_contract_number")),
                    normalize_text(stored.get("credit_account")),
                    normalize_sum(stored.get("credit_amount")),
                    normalize_text(stored.get("credit_analytics")),
                    normalize_text(stored.get("credit_counterparty")),
                    normalize_text(stored.get("credit_contract")),
                    normalize_text(stored.get("credit_spec_number")),
                    normalize_text(stored.get("credit_base_contract")),
                    normalize_text(stored.get("credit_contract_number")),
                    normalize_text(stored.get("balance_side")),
                    normalize_sum(stored.get("balance_amount")),
                    json.dumps(stored, ensure_ascii=False, default=str),
                )
            )
            if len(movement_batch) >= 1000:
                if execute_batch is None:
                    for params in movement_batch:
                        cur.execute(movement_insert_sql, params)
                else:
                    execute_batch(cur, movement_insert_sql, movement_batch, page_size=1000)
                movement_batch.clear()
        if movement_batch:
            if execute_batch is None:
                for params in movement_batch:
                    cur.execute(movement_insert_sql, params)
            else:
                execute_batch(cur, movement_insert_sql, movement_batch, page_size=1000)

        line_insert_sql = """
INSERT INTO onec_document_lines
    (drive_file_id, drive_file_name, drive_path, source_sheet, source_row, kind,
     document_type, document_code1c, document_number, document_date_iso,
     nomenclature, content, vat_rate, commissioner, commissioner_contract, commissioner_contract_key,
     spec_number, base_contract, contract_number, contract_date_iso, settlement_account,
     line_id, amount, vat_amount, line_json, synced_at)
VALUES
    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, now());
"""
        line_batch = []
        for line in document_lines:
            stored = dict(line)
            stored.pop("_used", None)
            line_batch.append(
                (
                    file_id,
                    file_name,
                    drive_path,
                    normalize_text(stored.get("source_sheet")),
                    pg_int_or_zero(stored.get("source_row") or stored.get("row_num")),
                    normalize_text(stored.get("kind")),
                    normalize_text(stored.get("document_type")),
                    normalize_text(stored.get("document_code1c")),
                    normalize_text(stored.get("document_number")),
                    pg_date_or_none(stored.get("document_date_iso") or stored.get("document_date")),
                    normalize_text(stored.get("nomenclature")),
                    normalize_text(stored.get("content")),
                    normalize_text(stored.get("vat_rate")),
                    normalize_text(stored.get("commissioner")),
                    normalize_text(stored.get("commissioner_contract")),
                    normalize_text(stored.get("commissioner_contract_key")),
                    normalize_text(stored.get("spec_number")),
                    normalize_text(stored.get("base_contract")),
                    normalize_text(stored.get("contract_number")),
                    pg_date_or_none(stored.get("contract_date_iso") or stored.get("contract_date")),
                    normalize_text(stored.get("settlement_account")),
                    normalize_text(stored.get("line_id")),
                    normalize_sum(stored.get("sum")),
                    normalize_sum(stored.get("vat_amount")),
                    json.dumps(stored, ensure_ascii=False, default=str),
                )
            )
            if len(line_batch) >= 1000:
                if execute_batch is None:
                    for params in line_batch:
                        cur.execute(line_insert_sql, params)
                else:
                    execute_batch(cur, line_insert_sql, line_batch, page_size=1000)
                line_batch.clear()
        if line_batch:
            if execute_batch is None:
                for params in line_batch:
                    cur.execute(line_insert_sql, params)
            else:
                execute_batch(cur, line_insert_sql, line_batch, page_size=1000)


def onec_postgres_summary() -> dict[str, object]:
    try:
        with postgres_connect() as conn:
            ensure_onec_pg_schema(conn)
            with conn.cursor() as cur:
                cur.execute(
                    """
SELECT
    (SELECT COUNT(*) FROM onec_source_files) AS files_count,
    (SELECT COUNT(*) FROM onec_documents) AS docs_count,
    (SELECT COUNT(*) FROM onec_account_movements) AS account_movements_count,
    (SELECT COUNT(*) FROM onec_document_lines) AS document_lines_count,
    (SELECT MAX(synced_at) FROM onec_source_files) AS last_synced_at;
"""
                )
                row = cur.fetchone()
        return {
            "ok": True,
            "files_count": int(row[0] or 0) if row else 0,
            "docs_count": int(row[1] or 0) if row else 0,
            "account_movements_count": int(row[2] or 0) if row else 0,
            "document_lines_count": int(row[3] or 0) if row else 0,
            "last_synced_at": str(row[4]) if row and row[4] else "",
        }
    except Exception as exc:
        return {"ok": False, "error": str(exc), "files_count": 0, "docs_count": 0, "last_synced_at": ""}


def load_onec_sources_from_postgres(snapshot: dict[str, object] | None = None) -> dict[str, object]:
    with postgres_connect() as conn:
        ensure_onec_pg_schema(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT doc_json FROM onec_documents ORDER BY drive_file_name, source_sheet, source_row, id;")
            docs = []
            for (raw_doc,) in cur.fetchall():
                if isinstance(raw_doc, dict):
                    docs.append(dict(raw_doc))
                elif raw_doc:
                    docs.append(json.loads(raw_doc))

            cur.execute("SELECT COUNT(*) FROM onec_account_movements;")
            movement_count_row = cur.fetchone()
            all_account_movements_count = int(movement_count_row[0] or 0) if movement_count_row else 0
            cur.execute("SELECT COUNT(*) FROM onec_document_lines;")
            line_count_row = cur.fetchone()
            all_document_lines_count = int(line_count_row[0] or 0) if line_count_row else 0

            account_movements = []
            document_lines = []
            if snapshot is not None:
                delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
                spec_number = normalize_text(delivery.get("spec_number"))
                base_contract = normalize_text(delivery.get("main_dog_number"))
                line_seen: set[tuple[str, str, int]] = set()

                def append_document_line(raw_line: object) -> None:
                    if isinstance(raw_line, dict):
                        line = dict(raw_line)
                    elif raw_line:
                        line = json.loads(raw_line)
                    else:
                        return
                    line_key = (
                        normalize_text(line.get("source_file")),
                        normalize_text(line.get("source_sheet")),
                        pg_int_or_zero(line.get("source_row") or line.get("row_num")),
                    )
                    if line_key in line_seen:
                        return
                    line_seen.add(line_key)
                    document_lines.append(line)

                if spec_number and base_contract:
                    cur.execute(
                        """
SELECT movement_json
FROM onec_account_movements
WHERE (debit_spec_number = %s AND debit_base_contract = %s)
   OR (credit_spec_number = %s AND credit_base_contract = %s)
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                        (spec_number, base_contract, spec_number, base_contract),
                    )
                elif spec_number:
                    cur.execute(
                        """
SELECT movement_json
FROM onec_account_movements
WHERE debit_spec_number = %s OR credit_spec_number = %s
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                        (spec_number, spec_number),
                    )
                else:
                    cur.execute("SELECT movement_json FROM onec_account_movements WHERE false;")
                for (raw_movement,) in cur.fetchall():
                    if isinstance(raw_movement, dict):
                        account_movements.append(dict(raw_movement))
                    elif raw_movement:
                        account_movements.append(json.loads(raw_movement))

                if spec_number and base_contract:
                    cur.execute(
                        """
SELECT line_json
FROM onec_document_lines
WHERE spec_number = %s AND base_contract = %s
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                        (spec_number, base_contract),
                    )
                elif spec_number:
                    cur.execute(
                        """
SELECT line_json
FROM onec_document_lines
WHERE spec_number = %s
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                        (spec_number,),
                    )
                else:
                    cur.execute("SELECT line_json FROM onec_document_lines WHERE false;")
                for (raw_line,) in cur.fetchall():
                    append_document_line(raw_line)

                act_line_pairs: list[tuple[str, str]] = []
                act_snapshot_rows = snapshot.get("akts") if isinstance(snapshot.get("akts"), list) else []
                for act_row in act_snapshot_rows:
                    if not isinstance(act_row, dict):
                        continue
                    for code_name, date_name in [("code1c", "date_iso"), ("main_code1c", "main_date_iso")]:
                        code = normalize_text(act_row.get(code_name))
                        date_iso = parse_any_date_to_iso(act_row.get(date_name) or act_row.get("date"))
                        if code and date_iso:
                            pair = (code, date_iso)
                            if pair not in act_line_pairs:
                                act_line_pairs.append(pair)
                if act_line_pairs:
                    clauses = []
                    params: list[object] = []
                    for code, date_iso in act_line_pairs:
                        clauses.append("(document_code1c = %s AND document_date_iso = %s)")
                        params.extend([code, date_iso])
                    cur.execute(
                        f"""
SELECT line_json
FROM onec_document_lines
WHERE {" OR ".join(clauses)}
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                        params,
                    )
                    for (raw_line,) in cur.fetchall():
                        append_document_line(raw_line)

            cur.execute(
                """
SELECT drive_file_name, drive_path, docs_count, account_movements_count, document_lines_count, summary_json, warnings_json, synced_at
FROM onec_source_files
ORDER BY drive_path, drive_file_name;
"""
            )
            file_summaries = []
            warnings = []
            for name, drive_path, docs_count, account_movements_count, document_lines_count, summary_json, warnings_json, synced_at in cur.fetchall():
                summary = summary_json if isinstance(summary_json, dict) else {}
                file_warnings = warnings_json if isinstance(warnings_json, list) else []
                warnings.extend(file_warnings)
                file_summaries.append(
                    {
                        "file": name,
                        "drive_path": drive_path,
                        "docs_count": docs_count,
                        "account_movements_count": account_movements_count,
                        "document_lines_count": document_lines_count,
                        "kind": summary.get("kind", "unknown"),
                        "synced_at": str(synced_at) if synced_at else "",
                        "warnings": file_warnings,
                    }
                )

    filtered_docs = docs
    filter_summary: dict[str, object] = {"filtered_count": len(docs), "total_count": len(docs)}
    if snapshot is not None:
        filtered_docs, filter_summary = filter_onec_docs_for_snapshot(docs, snapshot)
        account_movements = filter_onec_account_movements_for_snapshot(account_movements, snapshot)
        document_lines = filter_onec_document_lines_for_snapshot(document_lines, snapshot)

    by_kind: dict[str, int] = {}
    for doc in filtered_docs:
        kind = normalize_text(doc.get("kind")) or "unknown"
        by_kind[kind] = by_kind.get(kind, 0) + 1

    contract_pairs = sorted(
        {
            f"{doc.get('base_contract') or doc.get('contract_number')}/{doc.get('spec_number')}".strip("/")
            for doc in filtered_docs
            if doc.get("base_contract") or doc.get("contract_number") or doc.get("spec_number")
        }
    )

    return {
        "source_state": "postgresql",
        "storage": "postgresql",
        "files": file_summaries,
        "warnings": warnings,
        "all_docs_count": len(docs),
        "docs_count": len(filtered_docs),
        "all_account_movements_count": all_account_movements_count,
        "account_movements_count": len(account_movements),
        "all_document_lines_count": all_document_lines_count,
        "document_lines_count": len(document_lines),
        "cache_hit": True,
        "by_kind": by_kind,
        "contract_pairs": contract_pairs[:200],
        "filter": filter_summary,
        "docs": filtered_docs,
        "account_movements": account_movements,
        "document_lines": document_lines,
    }


def load_onec_postgres_aux_for_snapshot(snapshot: dict[str, object]) -> dict[str, object]:
    account_movements: list[dict[str, object]] = []
    document_lines: list[dict[str, object]] = []
    with postgres_connect() as conn:
        ensure_onec_pg_schema(conn)
        with conn.cursor() as cur:
            delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
            spec_number = normalize_text(delivery.get("spec_number"))
            base_contract = normalize_text(delivery.get("main_dog_number"))
            line_seen: set[tuple[str, str, int]] = set()

            def append_document_line(raw_line: object) -> None:
                if isinstance(raw_line, dict):
                    line = dict(raw_line)
                elif raw_line:
                    line = json.loads(raw_line)
                else:
                    return
                line_key = (
                    normalize_text(line.get("source_file")),
                    normalize_text(line.get("source_sheet")),
                    pg_int_or_zero(line.get("source_row") or line.get("row_num")),
                )
                if line_key in line_seen:
                    return
                line_seen.add(line_key)
                document_lines.append(line)

            if spec_number and base_contract:
                cur.execute(
                    """
SELECT movement_json
FROM onec_account_movements
WHERE (debit_spec_number = %s AND debit_base_contract = %s)
   OR (credit_spec_number = %s AND credit_base_contract = %s)
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                    (spec_number, base_contract, spec_number, base_contract),
                )
            elif spec_number:
                cur.execute(
                    """
SELECT movement_json
FROM onec_account_movements
WHERE debit_spec_number = %s OR credit_spec_number = %s
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                    (spec_number, spec_number),
                )
            else:
                cur.execute("SELECT movement_json FROM onec_account_movements WHERE false;")
            for (raw_movement,) in cur.fetchall():
                if isinstance(raw_movement, dict):
                    account_movements.append(dict(raw_movement))
                elif raw_movement:
                    account_movements.append(json.loads(raw_movement))

            if spec_number and base_contract:
                cur.execute(
                    """
SELECT line_json
FROM onec_document_lines
WHERE spec_number = %s AND base_contract = %s
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                    (spec_number, base_contract),
                )
            elif spec_number:
                cur.execute(
                    """
SELECT line_json
FROM onec_document_lines
WHERE spec_number = %s
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                    (spec_number,),
                )
            else:
                cur.execute("SELECT line_json FROM onec_document_lines WHERE false;")
            for (raw_line,) in cur.fetchall():
                append_document_line(raw_line)

            act_line_pairs: list[tuple[str, str]] = []
            act_snapshot_rows = snapshot.get("akts") if isinstance(snapshot.get("akts"), list) else []
            for act_row in act_snapshot_rows:
                if not isinstance(act_row, dict):
                    continue
                for code_name, date_name in [("code1c", "date_iso"), ("main_code1c", "main_date_iso")]:
                    code = normalize_text(act_row.get(code_name))
                    date_iso = parse_any_date_to_iso(act_row.get(date_name) or act_row.get("date"))
                    if code and date_iso:
                        pair = (code, date_iso)
                        if pair not in act_line_pairs:
                            act_line_pairs.append(pair)
            if act_line_pairs:
                clauses = []
                params: list[object] = []
                for code, date_iso in act_line_pairs:
                    clauses.append("(document_code1c = %s AND document_date_iso = %s)")
                    params.extend([code, date_iso])
                cur.execute(
                    f"""
SELECT line_json
FROM onec_document_lines
WHERE {" OR ".join(clauses)}
ORDER BY drive_file_name, source_sheet, source_row, id;
""",
                    params,
                )
                for (raw_line,) in cur.fetchall():
                    append_document_line(raw_line)

    account_movements = filter_onec_account_movements_for_snapshot(account_movements, snapshot)
    document_lines = filter_onec_document_lines_for_snapshot(document_lines, snapshot)
    return {"account_movements": account_movements, "document_lines": document_lines}


def filter_onec_postgres_base_source_for_snapshot(base_source: dict[str, object], snapshot: dict[str, object]) -> dict[str, object]:
    docs = base_source.get("docs") if isinstance(base_source.get("docs"), list) else []
    filtered_docs, filter_summary = filter_onec_docs_for_snapshot(docs, snapshot)
    aux = load_onec_postgres_aux_for_snapshot(snapshot)
    account_movements = aux.get("account_movements") if isinstance(aux.get("account_movements"), list) else []
    document_lines = aux.get("document_lines") if isinstance(aux.get("document_lines"), list) else []

    by_kind: dict[str, int] = {}
    for doc in filtered_docs:
        kind = normalize_text(doc.get("kind")) or "unknown"
        by_kind[kind] = by_kind.get(kind, 0) + 1

    contract_pairs = sorted(
        {
            f"{doc.get('base_contract') or doc.get('contract_number')}/{doc.get('spec_number')}".strip("/")
            for doc in filtered_docs
            if doc.get("base_contract") or doc.get("contract_number") or doc.get("spec_number")
        }
    )

    return {
        "source_state": "postgresql",
        "storage": "postgresql",
        "files": base_source.get("files", []),
        "warnings": base_source.get("warnings", []),
        "all_docs_count": base_source.get("all_docs_count", len(docs)),
        "docs_count": len(filtered_docs),
        "all_account_movements_count": base_source.get("all_account_movements_count", 0),
        "account_movements_count": len(account_movements),
        "all_document_lines_count": base_source.get("all_document_lines_count", 0),
        "document_lines_count": len(document_lines),
        "cache_hit": True,
        "by_kind": by_kind,
        "contract_pairs": contract_pairs[:200],
        "filter": filter_summary,
        "docs": filtered_docs,
        "account_movements": account_movements,
        "document_lines": document_lines,
    }


def _first_non_empty(*values: object) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def _safe_rows(value: object) -> list[dict[str, object]]:
    if not isinstance(value, list):
        return []
    return [dict(item) for item in value if isinstance(item, dict)]


def _nested_get(row: dict[str, object], key: str, *fallback_keys: str) -> object:
    for item_key in (key, *fallback_keys):
        if "." in item_key:
            current: object = row
            ok = True
            for part in item_key.split("."):
                if isinstance(current, dict):
                    current = current.get(part)
                else:
                    ok = False
                    break
            if ok and current not in (None, ""):
                return current
        elif row.get(item_key) not in (None, ""):
            return row.get(item_key)
    return ""


def onec_rest_period_for_snapshot(snapshot: dict[str, object]) -> dict[str, str]:
    delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
    values: list[str] = []
    for key in ["spec_date_iso", "spec_date"]:
        parsed = parse_any_date_to_iso(delivery.get(key))
        if parsed:
            values.append(parsed)
    for collection_name in ["schets", "akts", "payments"]:
        collection = snapshot.get(collection_name)
        if not isinstance(collection, list):
            continue
        for row in collection:
            if not isinstance(row, dict):
                continue
            for key in ["date_iso", "date", "main_date_iso", "main_date", "payment_date_iso", "payment_date"]:
                parsed = parse_any_date_to_iso(row.get(key))
                if parsed:
                    values.append(parsed)
    if not values:
        today = datetime.now().date()
        return {"date_from": today.replace(month=1, day=1).isoformat(), "date_to": today.isoformat()}
    dates = [datetime.strptime(value, "%Y-%m-%d").date() for value in sorted(set(values))]
    return {
        "date_from": (min(dates) - timedelta(days=31)).isoformat(),
        "date_to": (max(dates) + timedelta(days=31)).isoformat(),
    }


def build_onec_rest_request(snapshot: dict[str, object]) -> dict[str, object]:
    delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
    contracts = snapshot.get("contracts") if isinstance(snapshot.get("contracts"), list) else []
    schets = snapshot.get("schets") if isinstance(snapshot.get("schets"), list) else []
    akts = snapshot.get("akts") if isinstance(snapshot.get("akts"), list) else []
    payments = snapshot.get("payments") if isinstance(snapshot.get("payments"), list) else []
    request_contracts: list[dict[str, object]] = []
    for contract in contracts:
        if not isinstance(contract, dict):
            continue
        role_label = normalize_text(contract.get("role_label")).lower()
        source_roles = normalize_text(contract.get("source_roles")).lower()
        role = "related"
        if "комитент" in role_label or "principal" in source_roles:
            role = "committent"
        elif "покупател" in role_label or "buyer" in source_roles:
            role = "buyer"
        request_contracts.append(
            {
                "role": role,
                "erp_dog_id": sql_int(contract.get("dog_id")),
                "code1c": normalize_text(contract.get("dog_code1c")),
                "number": normalize_text(contract.get("dog_number")),
                "base_contract": normalize_text(contract.get("base_dog_number")),
                "spec_number": normalize_text(contract.get("spec_number")),
                "date": normalize_text(contract.get("dog_date")),
                "type": normalize_text(contract.get("dog_type_name")),
            }
        )

    def doc_payload(row: dict[str, object], kind: str) -> dict[str, object]:
        if kind == "payment":
            date = parse_any_date_to_iso(row.get("payment_date_iso") or row.get("payment_date"))
            amount = normalize_sum(row.get("classified_sum") or row.get("payment_sum"))
            number = normalize_text(row.get("pp_number"))
            erp_id = sql_int(row.get("payment_id"))
        else:
            date = parse_any_date_to_iso(row.get("date_iso") or row.get("date"))
            amount = normalize_sum(row.get("sum"))
            number = normalize_text(row.get("number"))
            erp_id = sql_int(row.get("erp_doc_id"))
        return {
            "kind": kind,
            "erp_id": erp_id,
            "operation_id": sql_int(row.get("oper_id")),
            "code1c": normalize_text(row.get("code1c")),
            "number": number,
            "date": date,
            "sum": amount,
            "invoice_number": normalize_text(row.get("invoice_number")),
            "contract_code1c": normalize_text(row.get("dog_code1c")),
            "contract_number": normalize_text(row.get("dog_number")),
            "type": normalize_text(row.get("type_name") or row.get("payment_type") or row.get("doc_kind")),
        }

    documents: list[dict[str, object]] = []
    for row in schets:
        if isinstance(row, dict):
            documents.append(doc_payload(row, "customer_invoice"))
    for row in akts:
        if not isinstance(row, dict):
            continue
        documents.append(doc_payload(row, "closing_document"))
        main_code = normalize_text(row.get("main_code1c"))
        if main_code:
            main_row = dict(row)
            main_row["code1c"] = main_code
            main_row["number"] = normalize_text(row.get("main_number"))
            main_row["date_iso"] = normalize_text(row.get("main_date_iso"))
            main_row["date"] = normalize_text(row.get("main_date"))
            main_row["sum"] = row.get("main_sum")
            documents.append(doc_payload(main_row, "closing_document"))
    for row in payments:
        if isinstance(row, dict):
            documents.append(doc_payload(row, "payment"))

    return {
        "request_id": f"spec-{sql_int(snapshot.get('spec_id'))}-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "mode": "delivery_reconciliation",
        "period": onec_rest_period_for_snapshot(snapshot),
        "delivery": {
            "spec_id": sql_int(snapshot.get("spec_id")),
            "spec_number": normalize_text(delivery.get("spec_number")),
            "spec_date": normalize_text(delivery.get("spec_date_iso") or delivery.get("spec_date")),
            "base_contract": normalize_text(delivery.get("main_dog_number")),
            "buyer_contract_code1c": normalize_text(delivery.get("spec_buyer_code1c")),
            "committent_contract_code1c": normalize_text(delivery.get("spec_principal_code1c")),
        },
        "organizations": [
            {
                "erp_id": sql_int(delivery.get("org_id")),
                "name": normalize_text(delivery.get("org_name")),
                "abbr": normalize_text(delivery.get("org_abbr")),
                "inn": normalize_text(delivery.get("org_inn")),
            }
        ],
        "counterparties": [
            {
                "erp_id": sql_int(delivery.get("client_id")),
                "name": normalize_text(delivery.get("client_name")),
                "abbr": normalize_text(delivery.get("client_abbr")),
                "inn": normalize_text(delivery.get("client_inn")),
            }
        ],
        "contracts": request_contracts,
        "documents": documents,
        "include": {
            "contracts": True,
            "customer_invoices": True,
            "payments": True,
            "sales": True,
            "purchases": True,
            "document_lines": True,
            "account_movements": True,
            "balances": True,
        },
        "transport": "rest",
        "contract_version": "reconciliation.v1",
    }


def _onec_contract_extra(row: dict[str, object]) -> dict[str, object]:
    spec_number = _first_non_empty(row.get("spec_number"), row.get("specNum"), row.get("Заявка"), row.get("Спецификация"))
    base_contract = _first_non_empty(row.get("base_contract"), row.get("baseContract"), row.get("dog_number"), row.get("contract_number"))
    contract_number = _first_non_empty(row.get("contract_number"), row.get("contractNumber"), row.get("number"), row.get("dog_number"))
    contract_text = _first_non_empty(row.get("contract"), row.get("contract_text"), contract_number)
    return {
        "contract": contract_text,
        "contract_key": compact_key(contract_text),
        "spec_number": spec_number,
        "base_contract": base_contract,
        "contract_number": contract_number,
        "contract_date": normalize_text(row.get("contract_date") or row.get("contractDate")),
        "contract_date_iso": parse_any_date_to_iso(row.get("contract_date_iso") or row.get("contract_date") or row.get("contractDate")),
    }


def _onec_doc(row: dict[str, object], *, kind: str, source_sheet: str, source_row: int, default_title: str) -> dict[str, object]:
    extra = _onec_contract_extra(row)
    extra.update(
        {
            "source": "rest",
            "guid": normalize_text(row.get("guid") or row.get("c1guid") or row.get("ref")),
            "posted": bool(row.get("posted") or row.get("is_posted") or row.get("Проведен")),
            "deleted": bool(row.get("deleted") or row.get("is_deleted") or row.get("ПометкаУдаления")),
            "vat_rate": normalize_text(row.get("vat_rate") or row.get("nds") or row.get("СтавкаНДС")),
        }
    )
    return make_onec_doc(
        kind=kind,
        source_file="1C REST",
        source_sheet=source_sheet,
        source_row=source_row,
        title=_first_non_empty(row.get("title"), row.get("type"), row.get("doc_type"), default_title),
        number=_first_non_empty(row.get("number"), row.get("num"), row.get("Номер")),
        date=_first_non_empty(row.get("date_iso"), row.get("date"), row.get("doc_date"), row.get("Дата")),
        amount=row.get("sum") if row.get("sum") not in (None, "") else row.get("amount"),
        organization=_first_non_empty(row.get("organization"), row.get("org"), row.get("Организация")),
        counterparty=_first_non_empty(row.get("counterparty"), row.get("contragent"), row.get("Контрагент")),
        contract=extra.get("contract"),
        invoice_number=_first_non_empty(row.get("invoice_number"), row.get("invoiceNumber"), row.get("СчетНаОплату")),
        comment=_first_non_empty(row.get("comment"), row.get("purpose"), row.get("НазначениеПлатежа")),
        code1c=_first_non_empty(row.get("code1c"), row.get("kod1c"), row.get("Код1С"), row.get("number")),
        ref_id=_first_non_empty(row.get("ref_id"), row.get("guid"), row.get("c1guid")),
        extra=extra,
    )


def normalize_onec_rest_snapshot(payload: dict[str, object], snapshot: dict[str, object] | None = None) -> dict[str, object]:
    data = payload.get("snapshot") if isinstance(payload.get("snapshot"), dict) else payload
    docs: list[dict[str, object]] = []
    for block_name, kind, sheet, default_title in [
        ("contracts", "contract", "Договоры", "Договор"),
        ("customer_invoices", "invoice", "Счета покупателю", "Счет покупателю"),
        ("payments", "bank_receipt", "Поступления оплат", "Поступление на расчетный счет"),
        ("sales", "sale_act", "Реализация", "Реализация (акт, накладная, УПД)"),
        ("purchases", "purchase_act", "Поступление", "Поступление (акт, накладная, УПД)"),
    ]:
        for idx, row in enumerate(_safe_rows(data.get(block_name)), start=1):
            docs.append(_onec_doc(row, kind=kind, source_sheet=sheet, source_row=idx, default_title=default_title))

    account_movements: list[dict[str, object]] = []
    for idx, row in enumerate(_safe_rows(data.get("account_movements")), start=1):
        account_movements.append(
            {
                "row_num": idx,
                "source_file": "1C REST",
                "source_sheet": "Карточка счетов",
                "source_row": idx,
                "account_scope": normalize_text(row.get("account_scope") or row.get("account") or row.get("Счет")),
                "period": normalize_text(row.get("period") or row.get("date")),
                "period_iso": parse_any_date_to_iso(row.get("period_iso") or row.get("period") or row.get("date")),
                "registrar_text": normalize_text(row.get("registrar_text") or row.get("registrar") or row.get("document")),
                "registrar_number": normalize_text(row.get("registrar_number") or row.get("number") or row.get("code1c")),
                "registrar_date_iso": parse_any_date_to_iso(row.get("registrar_date_iso") or row.get("registrar_date") or row.get("date")),
                "debit_account": normalize_text(_nested_get(row, "debit.account", "debit_account", "СчетДт")),
                "debit_amount": normalize_sum(_nested_get(row, "debit.amount", "debit_amount", "СуммаДт")) or 0,
                "debit_analytics": normalize_text(_nested_get(row, "debit.analytics", "debit_analytics", "АналитикаДт")),
                "debit_counterparty": normalize_text(_nested_get(row, "debit.counterparty", "debit_counterparty")),
                "debit_contract": normalize_text(_nested_get(row, "debit.contract", "debit_contract")),
                "debit_spec_number": normalize_text(_nested_get(row, "debit.spec_number", "debit_spec_number")),
                "debit_base_contract": normalize_text(_nested_get(row, "debit.base_contract", "debit_base_contract")),
                "debit_contract_number": normalize_text(_nested_get(row, "debit.contract_number", "debit_contract_number")),
                "credit_account": normalize_text(_nested_get(row, "credit.account", "credit_account", "СчетКт")),
                "credit_amount": normalize_sum(_nested_get(row, "credit.amount", "credit_amount", "СуммаКт")) or 0,
                "credit_analytics": normalize_text(_nested_get(row, "credit.analytics", "credit_analytics", "АналитикаКт")),
                "credit_counterparty": normalize_text(_nested_get(row, "credit.counterparty", "credit_counterparty")),
                "credit_contract": normalize_text(_nested_get(row, "credit.contract", "credit_contract")),
                "credit_spec_number": normalize_text(_nested_get(row, "credit.spec_number", "credit_spec_number")),
                "credit_base_contract": normalize_text(_nested_get(row, "credit.base_contract", "credit_base_contract")),
                "credit_contract_number": normalize_text(_nested_get(row, "credit.contract_number", "credit_contract_number")),
            }
        )

    document_lines: list[dict[str, object]] = []
    for idx, row in enumerate(_safe_rows(data.get("document_lines")), start=1):
        contract_extra = _onec_contract_extra(row)
        document_lines.append(
            {
                "row_num": idx,
                "source_file": "1C REST",
                "source_sheet": "Строки документов",
                "source_row": idx,
                "kind": normalize_text(row.get("kind") or row.get("line_kind")) or "purchase_act_line_details",
                "document_type": normalize_text(row.get("document_type") or row.get("type")),
                "document_code1c": normalize_text(row.get("document_code1c") or row.get("code1c") or row.get("number")),
                "document_number": normalize_text(row.get("document_number") or row.get("number")),
                "document_date": normalize_text(row.get("document_date") or row.get("date")),
                "document_date_iso": parse_any_date_to_iso(row.get("document_date_iso") or row.get("document_date") or row.get("date")),
                "line_id": normalize_text(row.get("line_id") or row.get("row_id")),
                "line_number": normalize_text(row.get("line_number") or row.get("row_number")),
                "amount": normalize_sum(row.get("amount") if row.get("amount") not in (None, "") else row.get("sum")) or 0,
                "nomenclature": normalize_text(row.get("nomenclature") or row.get("Номенклатура")),
                "content": normalize_text(row.get("content") or row.get("Содержание")),
                "vat_rate": normalize_text(row.get("vat_rate") or row.get("nds") or row.get("СтавкаНДС")),
                "commissioner": normalize_text(row.get("commissioner") or row.get("committent") or row.get("Комитент")),
                "commissioner_contract": normalize_text(row.get("commissioner_contract") or row.get("committent_contract") or contract_extra.get("contract")),
                "commissioner_contract_key": compact_key(row.get("commissioner_contract") or row.get("committent_contract") or contract_extra.get("contract")),
                "spec_number": contract_extra.get("spec_number"),
                "base_contract": contract_extra.get("base_contract"),
                "contract_number": contract_extra.get("contract_number"),
                "contract_date": contract_extra.get("contract_date"),
                "contract_date_iso": contract_extra.get("contract_date_iso"),
                "settlement_account": normalize_text(row.get("settlement_account") or row.get("account") or row.get("Счет")),
            }
        )

    filtered_docs = docs
    filtered_movements = account_movements
    filtered_document_lines = document_lines
    filter_summary: dict[str, object] = {"filtered_count": len(docs), "total_count": len(docs)}
    if snapshot is not None:
        filtered_docs, filter_summary = filter_onec_docs_for_snapshot(docs, snapshot)
        filtered_movements = filter_onec_account_movements_for_snapshot(account_movements, snapshot)
        filtered_document_lines = filter_onec_document_lines_for_snapshot(document_lines, snapshot)

    by_kind: dict[str, int] = {}
    for doc in filtered_docs:
        kind = normalize_text(doc.get("kind")) or "unknown"
        by_kind[kind] = by_kind.get(kind, 0) + 1
    contract_pairs = sorted(
        {
            f"{doc.get('base_contract') or doc.get('contract_number')}/{doc.get('spec_number')}".strip("/")
            for doc in filtered_docs
            if doc.get("base_contract") or doc.get("contract_number") or doc.get("spec_number")
        }
    )
    warnings = data.get("warnings") if isinstance(data.get("warnings"), list) else []
    return {
        "source_state": "rest",
        "storage": "rest",
        "files": [{"file": "1C REST", "kind": "rest_snapshot"}],
        "warnings": warnings,
        "all_docs_count": len(docs),
        "docs_count": len(filtered_docs),
        "all_account_movements_count": len(account_movements),
        "account_movements_count": len(filtered_movements),
        "all_document_lines_count": len(document_lines),
        "document_lines_count": len(filtered_document_lines),
        "cache_hit": False,
        "by_kind": by_kind,
        "contract_pairs": contract_pairs[:200],
        "filter": filter_summary,
        "docs": filtered_docs,
        "account_movements": filtered_movements,
        "document_lines": filtered_document_lines,
        "rest_request_id": normalize_text(data.get("request_id") or payload.get("request_id")),
    }


def prune_recon_jobs(now: float | None = None) -> None:
    now = time.time() if now is None else now
    with RECON_JOBS_LOCK:
        ttl = max(RECON_JOB_TTL_SECONDS, 0)
        if ttl:
            for job_id in list(RECON_JOBS):
                created_at = float(RECON_JOBS.get(job_id, {}).get("created_at_ts") or 0)
                if now - created_at > ttl:
                    RECON_JOBS.pop(job_id, None)
        max_items = max(RECON_JOB_MAX_ITEMS, 0)
        if max_items and len(RECON_JOBS) > max_items:
            overflow = len(RECON_JOBS) - max_items
            oldest = sorted(RECON_JOBS.items(), key=lambda item: float(item[1].get("created_at_ts") or 0))
            for job_id, _ in oldest[:overflow]:
                RECON_JOBS.pop(job_id, None)


def recon_jobs_summary() -> dict[str, object]:
    prune_recon_jobs()
    with RECON_JOBS_LOCK:
        by_status: dict[str, int] = {}
        for job in RECON_JOBS.values():
            status = normalize_text(job.get("status")) or "unknown"
            by_status[status] = by_status.get(status, 0) + 1
        return {
            "items": len(RECON_JOBS),
            "ttl_seconds": RECON_JOB_TTL_SECONDS,
            "max_items": RECON_JOB_MAX_ITEMS,
            "background_matrix_max_limit": RECON_BACKGROUND_MATRIX_MAX_LIMIT,
            "by_status": by_status,
        }


def start_recon_job(kind: str, params: dict[str, object], runner) -> dict[str, object]:
    prune_recon_jobs()
    job_id = uuid.uuid4().hex
    now = time.time()
    job = {
        "id": job_id,
        "kind": kind,
        "status": "queued",
        "params": params,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "created_at_ts": now,
        "started_at": "",
        "finished_at": "",
        "elapsed_seconds": None,
        "error": "",
        "result": None,
    }
    with RECON_JOBS_LOCK:
        RECON_JOBS[job_id] = job

    def worker() -> None:
        started = time.time()
        with RECON_JOBS_LOCK:
            current = RECON_JOBS.get(job_id)
            if current is not None:
                current["status"] = "running"
                current["started_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            result = runner()
            status = "completed"
            error = ""
        except Exception as exc:
            result = {
                "ok": False,
                "error": str(exc),
                "traceback": traceback.format_exc(limit=8),
            }
            status = "failed"
            error = str(exc)
        finished = time.time()
        with RECON_JOBS_LOCK:
            current = RECON_JOBS.get(job_id)
            if current is not None:
                current["status"] = status
                current["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                current["elapsed_seconds"] = round(finished - started, 3)
                current["error"] = error
                current["result"] = result

    thread = threading.Thread(target=worker, name=f"recon-job-{job_id[:8]}", daemon=True)
    thread.start()
    return public_recon_job(job, include_result=False)


def public_recon_job(job: dict[str, object], include_result: bool = False) -> dict[str, object]:
    result = job.get("result")
    payload = {
        "id": job.get("id"),
        "kind": job.get("kind"),
        "status": job.get("status"),
        "params": job.get("params"),
        "created_at": job.get("created_at"),
        "started_at": job.get("started_at"),
        "finished_at": job.get("finished_at"),
        "elapsed_seconds": job.get("elapsed_seconds"),
        "error": job.get("error"),
        "result_ready": result is not None,
    }
    if include_result:
        payload["result"] = result
    return payload


def get_recon_job(job_id: str, include_result: bool = False) -> dict[str, object] | None:
    prune_recon_jobs()
    with RECON_JOBS_LOCK:
        job = RECON_JOBS.get(job_id)
        if job is None:
            return None
        return public_recon_job(dict(job), include_result=include_result)


def load_onec_sources_from_rest(snapshot: dict[str, object]) -> dict[str, object]:
    if OneCRestClient is None:
        raise RuntimeError("onec_rest_client.py is unavailable")
    request = build_onec_rest_request(snapshot)
    payload = OneCRestClient.from_env().get_reconciliation_snapshot(request)
    source = normalize_onec_rest_snapshot(payload, snapshot=snapshot)
    source["request"] = request
    return source


def list_drive_children(service, folder_id: str, parent_path: str = "") -> list[dict[str, object]]:
    files: list[dict[str, object]] = []
    page_token = None
    while True:
        response = (
            service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="nextPageToken, files(id, name, mimeType, modifiedTime, size, md5Checksum)",
                pageSize=1000,
                pageToken=page_token,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            )
            .execute()
        )
        for item in response.get("files", []):
            item_path = f"{parent_path}/{item.get('name')}".strip("/")
            item["drive_path"] = item_path
            if item.get("mimeType") == "application/vnd.google-apps.folder":
                files.extend(list_drive_children(service, normalize_text(item.get("id")), item_path))
            else:
                files.append(item)
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return files


def is_supported_drive_sheet(file_meta: dict[str, object]) -> bool:
    name = normalize_text(file_meta.get("name")).lower()
    mime_type = normalize_text(file_meta.get("mimeType"))
    return (
        name.endswith(".xlsx")
        or mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or mime_type == "application/vnd.google-apps.spreadsheet"
    )


def download_drive_sheet_as_xlsx(service, file_meta: dict[str, object]) -> bytes:
    mime_type = normalize_text(file_meta.get("mimeType"))
    file_id = normalize_text(file_meta.get("id"))
    if mime_type == "application/vnd.google-apps.spreadsheet":
        request = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    return request.execute()


def google_credentials(scopes: list[str]):
    if service_account is None or build_google_service is None:
        raise RuntimeError("Google API libraries are not installed")
    if not GOOGLE_CREDENTIALS_FILE.exists():
        raise RuntimeError(f"Google service account file not found: {GOOGLE_CREDENTIALS_FILE}")
    return service_account.Credentials.from_service_account_file(str(GOOGLE_CREDENTIALS_FILE), scopes=scopes)


def parse_drive_modified_time(value: object) -> datetime | None:
    raw = normalize_text(value)
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(microsecond=0)
    except Exception:
        return None


def current_onec_file_cache(conn, file_meta: dict[str, object]) -> dict[str, object] | None:
    file_id = normalize_text(file_meta.get("id"))
    if not file_id:
        return None
    with conn.cursor() as cur:
        cur.execute(
            """
SELECT modified_time, file_size, docs_count, account_movements_count, document_lines_count, summary_json
FROM onec_source_files
WHERE drive_file_id = %s;
""",
            (file_id,),
        )
        row = cur.fetchone()
    if not row:
        return None

    stored_modified, stored_size, docs_count, movements_count, document_lines_count, summary_json = row
    expected_size = pg_int_or_zero(file_meta.get("size")) or None
    expected_modified = parse_drive_modified_time(file_meta.get("modifiedTime"))
    same_size = expected_size is None or int(stored_size or 0) == int(expected_size or 0)
    same_modified = True
    if expected_modified and stored_modified:
        same_modified = stored_modified.replace(microsecond=0) == expected_modified
    if not same_size or not same_modified:
        return None

    summary = summary_json if isinstance(summary_json, dict) else {}
    file_name_lower = normalize_text(file_meta.get("name")).lower()
    looks_like_line_details = (
        "агентские" in file_name_lower
        or "детал" in file_name_lower
        or "строк" in file_name_lower
    ) and "поступ" in file_name_lower
    if looks_like_line_details and summary.get("kind") not in {"purchase_act_line_details", "purchase_act_service_line_details", "purchase_act_detail_unsupported"}:
        return None
    return {
        "docs_count": int(docs_count or 0),
        "account_movements_count": int(movements_count or 0),
        "document_lines_count": int(document_lines_count or 0),
        "kind": summary.get("kind", "unknown"),
    }


def sync_onec_drive_to_postgres(folder_id: str | None = None) -> dict[str, object]:
    folder_id = normalize_text(folder_id) or ONEC_DRIVE_FOLDER_ID
    if not folder_id:
        raise RuntimeError("RECON_ONEC_DRIVE_FOLDER_ID is not configured")

    credentials = google_credentials(["https://www.googleapis.com/auth/drive.readonly"])
    service = build_google_service("drive", "v3", credentials=credentials)
    drive_files = list_drive_children(service, folder_id)
    sheet_files = [item for item in drive_files if is_supported_drive_sheet(item)]

    result_files: list[dict[str, object]] = []
    total_docs = 0
    total_account_movements = 0
    total_document_lines = 0
    warnings: list[str] = []

    conn = postgres_connect()
    try:
        ensure_onec_pg_schema(conn)
        conn.commit()
        with tempfile.TemporaryDirectory(prefix="onec_drive_") as tmp_dir:
            for item in sheet_files:
                try:
                    cached_file = current_onec_file_cache(conn, item)
                    if cached_file:
                        result_files.append(
                            {
                                "name": item.get("name"),
                                "drive_path": item.get("drive_path"),
                                "kind": cached_file.get("kind"),
                                "docs_count": cached_file.get("docs_count", 0),
                                "account_movements_count": cached_file.get("account_movements_count", 0),
                                "document_lines_count": cached_file.get("document_lines_count", 0),
                                "warnings_count": 0,
                                "skipped": True,
                            }
                        )
                        total_docs += int(cached_file.get("docs_count") or 0)
                        total_account_movements += int(cached_file.get("account_movements_count") or 0)
                        total_document_lines += int(cached_file.get("document_lines_count") or 0)
                        continue

                    payload = download_drive_sheet_as_xlsx(service, item)
                    safe_name = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._() -]+", "_", normalize_text(item.get("name"))).strip(" .") or "drive.xlsx"
                    if not safe_name.lower().endswith(".xlsx"):
                        safe_name = f"{safe_name}.xlsx"
                    path = Path(tmp_dir) / safe_name
                    path.write_bytes(payload)
                    docs, summary = parse_onec_workbook(path)
                    account_movements: list[dict[str, object]] = []
                    document_lines: list[dict[str, object]] = []
                    if summary.get("kind") == "account_card":
                        account_movements, movement_summary = parse_onec_account_movements(path)
                        summary["account_movements_count"] = len(account_movements)
                        summary["movement_sheets"] = movement_summary.get("sheets", [])
                        movement_warnings = movement_summary.get("warnings") if isinstance(movement_summary.get("warnings"), list) else []
                        summary_warnings = summary.get("warnings") if isinstance(summary.get("warnings"), list) else []
                        summary["warnings"] = summary_warnings + movement_warnings
                    if summary.get("kind") in {"purchase_act_line_details", "purchase_act_service_line_details"}:
                        document_lines, line_summary = parse_onec_document_lines(path)
                        summary["document_lines_count"] = len(document_lines)
                        summary["line_sheets"] = line_summary.get("sheets", [])
                        line_warnings = line_summary.get("warnings") if isinstance(line_summary.get("warnings"), list) else []
                        summary_warnings = summary.get("warnings") if isinstance(summary.get("warnings"), list) else []
                        summary["warnings"] = summary_warnings + line_warnings
                    insert_onec_file_to_postgres(conn, item, docs, summary, account_movements=account_movements, document_lines=document_lines)
                    conn.commit()
                    total_docs += len(docs)
                    total_account_movements += len(account_movements)
                    total_document_lines += len(document_lines)
                    file_warnings = summary.get("warnings") if isinstance(summary.get("warnings"), list) else []
                    warnings.extend([f"{item.get('drive_path')}: {warning}" for warning in file_warnings])
                    result_files.append(
                        {
                            "name": item.get("name"),
                            "drive_path": item.get("drive_path"),
                            "kind": summary.get("kind"),
                            "docs_count": len(docs),
                            "account_movements_count": len(account_movements),
                            "document_lines_count": len(document_lines),
                            "warnings_count": len(file_warnings),
                        }
                    )
                except Exception as exc:
                    conn.rollback()
                    warning = f"{item.get('drive_path') or item.get('name')}: {exc}"
                    warnings.append(warning)
                    result_files.append(
                        {
                            "name": item.get("name"),
                            "drive_path": item.get("drive_path"),
                            "docs_count": 0,
                            "account_movements_count": 0,
                            "document_lines_count": 0,
                            "error": str(exc),
                        }
                    )
    finally:
        conn.close()

    return {
        "folder_id": folder_id,
        "drive_files_count": len(drive_files),
        "xlsx_files_count": len(sheet_files),
        "docs_count": total_docs,
        "account_movements_count": total_account_movements,
        "document_lines_count": total_document_lines,
        "files": result_files,
        "warnings": warnings,
    }


def sheets_service():
    credentials = google_credentials(["https://www.googleapis.com/auth/spreadsheets.readonly"])
    return build_google_service("sheets", "v4", credentials=credentials)


def google_sheet_range_title(title: str) -> str:
    return "'" + title.replace("'", "''") + "'"


def row_cell(row: list[object], idx: int) -> object:
    return row[idx] if idx < len(row) else ""


def parse_control_spec_number(value: object) -> str:
    text = normalize_text(value)
    match = re.search(r"заявка\s*№\s*([0-9A-Za-zА-Яа-яЁё/_-]+)", text, re.IGNORECASE)
    return normalize_text(match.group(1)) if match else ""


def parse_control_upd_refs(value: object) -> list[dict[str, object]]:
    text = normalize_text(value)
    refs: list[dict[str, object]] = []
    if not text:
        return refs
    for match in re.finditer(r"(?:упд|счф|сф|отчет|отч[её]т)[^\d]{0,10}(\d{1,8}).{0,20}?от\s+(\d{2}\.\d{2}\.\d{4})", text, re.IGNORECASE):
        refs.append(
            {
                "number": normalize_text(match.group(1)),
                "date": normalize_text(match.group(2)),
                "date_iso": parse_any_date_to_iso(match.group(2)),
                "text": text,
            }
        )
    if not refs:
        date_match = re.search(r"(\d{2}\.\d{2}\.\d{4})", text)
        num_match = re.search(r"(\d{1,8})", text)
        if date_match or num_match:
            refs.append(
                {
                    "number": normalize_text(num_match.group(1)) if num_match else "",
                    "date": normalize_text(date_match.group(1)) if date_match else "",
                    "date_iso": parse_any_date_to_iso(date_match.group(1)) if date_match else "",
                    "text": text,
                }
            )
    return refs


def load_control_sheet_values(spreadsheet_id: str, gid: str = "", sheet_title: str = "") -> tuple[dict[str, object], list[list[object]]]:
    service = sheets_service()
    metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="properties/title,sheets(properties(title,sheetId,gridProperties(rowCount,columnCount)))").execute()
    sheets = metadata.get("sheets", [])
    selected = None
    if sheet_title:
        selected = next((item for item in sheets if item.get("properties", {}).get("title") == sheet_title), None)
    if selected is None and gid:
        try:
            gid_int = int(gid)
            selected = next((item for item in sheets if int(item.get("properties", {}).get("sheetId", -1)) == gid_int), None)
        except Exception:
            selected = None
    if selected is None:
        selected = sheets[0] if sheets else None
    if selected is None:
        raise RuntimeError("Google Sheet has no worksheets")

    props = selected.get("properties", {})
    title = normalize_text(props.get("title"))
    row_count = int((props.get("gridProperties") or {}).get("rowCount") or 20000)
    response = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range=f"{google_sheet_range_title(title)}!A1:M{row_count}",
            valueRenderOption="UNFORMATTED_VALUE",
            dateTimeRenderOption="FORMATTED_STRING",
        )
        .execute()
    )
    return {
        "spreadsheet_id": spreadsheet_id,
        "spreadsheet_title": metadata.get("properties", {}).get("title", ""),
        "sheet_title": title,
        "sheet_gid": str(props.get("sheetId", "")),
    }, response.get("values", [])


def extract_control_sheet_block(spec_id: int, spreadsheet_id: str = "", gid: str = "", sheet_title: str = "") -> dict[str, object]:
    snapshot = build_erp_snapshot(spec_id)
    delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
    spec_number = normalize_text(delivery.get("spec_number"))
    if not spec_number:
        raise RuntimeError(f"ERP spec {spec_id} has no spec number")

    sheet_meta, values = load_control_sheet_values(
        normalize_text(spreadsheet_id) or CONTROL_SHEET_ID,
        normalize_text(gid) or CONTROL_SHEET_GID,
        normalize_text(sheet_title),
    )

    start_idx = None
    for idx, row in enumerate(values):
        if parse_control_spec_number(row_cell(row, 1)) == spec_number:
            start_idx = idx
            break
    if start_idx is None:
        return {
            "ok": False,
            "spec_id": spec_id,
            "delivery": delivery,
            "sheet": sheet_meta,
            "found": False,
            "message": f"Заявка №{spec_number} не найдена в контрольной Google Sheet",
            "rows": [],
            "totals": {},
            "erp": snapshot.get("settlements", {}),
        }

    end_idx = len(values)
    for idx in range(start_idx + 1, len(values)):
        label = normalize_text(row_cell(values[idx], 1))
        if label:
            end_idx = idx
            break

    block_rows: list[dict[str, object]] = []
    for idx in range(start_idx, end_idx):
        row = values[idx]
        if not any(normalize_text(cell) for cell in row):
            continue
        block_rows.append(
            {
                "row": idx + 1,
                "spec_label": normalize_text(row_cell(row, 1)),
                "invoice_code": normalize_text(row_cell(row, 3)),
                "invoice_sum": normalize_sum(row_cell(row, 4)),
                "payment_sum": normalize_sum(row_cell(row, 5)),
                "reimbursable_sum": normalize_sum(row_cell(row, 7)),
                "non_reimbursable_sum": normalize_sum(row_cell(row, 8)),
                "upd_text": normalize_text(row_cell(row, 10)),
                "delta": normalize_sum(row_cell(row, 11)),
                "mark": normalize_text(row_cell(row, 12)),
                "upd_refs": parse_control_upd_refs(row_cell(row, 10)),
            }
        )

    def total(key: str) -> float:
        return round(sum(float(row.get(key) or 0.0) for row in block_rows), 2)

    totals = {
        "invoice_total": total("invoice_sum"),
        "payment_total": total("payment_sum"),
        "reimbursable_total": total("reimbursable_sum"),
        "non_reimbursable_total": total("non_reimbursable_sum"),
        "delta": total("delta"),
    }
    if totals["delta"] == 0 and block_rows:
        first_delta = normalize_sum(block_rows[0].get("delta"))
        if first_delta is not None:
            totals["delta"] = round(first_delta, 2)

    return {
        "ok": True,
        "spec_id": spec_id,
        "delivery": delivery,
        "sheet": sheet_meta,
        "found": True,
        "start_row": start_idx + 1,
        "end_row": end_idx,
        "rows": block_rows,
        "totals": totals,
        "erp_snapshot": snapshot,
    }


def find_onec_doc_by_code(docs: list[dict[str, object]], code: object, kinds: set[str]) -> dict[str, object] | None:
    code_keys = strong_doc_number_keys(code) or doc_number_keys(code)
    if not code_keys:
        return None
    for doc in docs:
        if normalize_text(doc.get("kind")) not in kinds:
            continue
        doc_keys = strong_doc_number_keys(doc.get("code1c")) | strong_doc_number_keys(doc.get("number")) | doc_number_keys(doc.get("code1c")) | doc_number_keys(doc.get("number"))
        if code_keys.intersection(doc_keys):
            return doc
    return None


def find_onec_act_by_control_ref(docs: list[dict[str, object]], refs: list[dict[str, object]], amount: object) -> dict[str, object] | None:
    amount_value = normalize_sum(amount)
    for doc in docs:
        if normalize_text(doc.get("kind")) not in {"sale_act", "purchase_act"}:
            continue
        doc_date = parse_any_date_to_iso(doc.get("date_iso") or doc.get("date"))
        doc_keys = doc_number_keys(doc.get("code1c")) | doc_number_keys(doc.get("number"))
        for ref in refs:
            ref_date = parse_any_date_to_iso(ref.get("date_iso") or ref.get("date"))
            ref_keys = doc_number_keys(ref.get("number"))
            if ref_date and doc_date and ref_date != doc_date:
                continue
            if ref_keys and doc_keys and ref_keys.intersection(doc_keys):
                return doc
        if amount_value is not None and money_equal(amount_value, doc.get("sum")):
            return doc
    return None


def metric_row(name: str, sheet_value: object, erp_value: object, onec_value: object = None, note: str = "") -> dict[str, object]:
    sheet_sum = normalize_sum(sheet_value)
    erp_sum = normalize_sum(erp_value)
    onec_sum = normalize_sum(onec_value)
    erp_diff = round((erp_sum or 0.0) - (sheet_sum or 0.0), 2) if sheet_sum is not None and erp_sum is not None else None
    onec_diff = round((onec_sum or 0.0) - (sheet_sum or 0.0), 2) if sheet_sum is not None and onec_sum is not None else None
    status = "MATCH"
    if erp_diff not in (None, 0.0):
        status = "FIELDS_MISMATCH"
    if onec_diff not in (None, 0.0):
        status = "FIELDS_MISMATCH"
    return {
        "metric": name,
        "sheet": sheet_sum,
        "erp": erp_sum,
        "onec": onec_sum,
        "erp_diff": erp_diff,
        "onec_diff": onec_diff,
        "status": status,
        "note": note,
    }


def build_three_way_control(spec_id: int, spreadsheet_id: str = "", gid: str = "", sheet_title: str = "") -> dict[str, object]:
    control = extract_control_sheet_block(spec_id, spreadsheet_id, gid, sheet_title)
    if not control.get("ok"):
        return control

    snapshot = control.get("erp_snapshot") if isinstance(control.get("erp_snapshot"), dict) else build_erp_snapshot(spec_id)
    settlements = snapshot.get("settlements") if isinstance(snapshot.get("settlements"), dict) else {}
    rows = control.get("rows") if isinstance(control.get("rows"), list) else []
    totals = control.get("totals") if isinstance(control.get("totals"), dict) else {}

    try:
        onec_source = load_onec_sources_from_postgres(snapshot)
        onec_docs = onec_source.get("docs") if isinstance(onec_source.get("docs"), list) else []
        onec_state = onec_source.get("source_state", "postgresql")
    except Exception as exc:
        onec_docs = []
        onec_state = f"unavailable: {exc}"

    schets = snapshot.get("schets") if isinstance(snapshot.get("schets"), list) else []
    sheet_invoice_codes = [row.get("invoice_code") for row in rows if normalize_text(row.get("invoice_code"))]
    erp_invoice_docs = [doc for doc in schets if any(doc_number_keys(code).intersection(doc_number_keys(doc.get("code1c"))) for code in sheet_invoice_codes)]
    onec_invoice_docs: list[dict[str, object]] = []
    document_rows: list[dict[str, object]] = []
    for row in rows:
        code = normalize_text(row.get("invoice_code"))
        if not code:
            continue
        erp_doc = next((doc for doc in erp_invoice_docs if doc_number_keys(code).intersection(doc_number_keys(doc.get("code1c")))), None)
        onec_doc = find_onec_doc_by_code(onec_docs, code, {"invoice"})
        if onec_doc:
            onec_invoice_docs.append(onec_doc)
        document_rows.append(
            {
                "kind": "invoice",
                "sheet_row": row.get("row"),
                "sheet_code": code,
                "sheet_sum": row.get("invoice_sum"),
                "erp_code1c": erp_doc.get("code1c") if erp_doc else "",
                "erp_sum": erp_doc.get("sum") if erp_doc else None,
                "onec_code1c": onec_doc.get("code1c") if onec_doc else "",
                "onec_sum": onec_doc.get("sum") if onec_doc else None,
                "status": "MATCH" if erp_doc and onec_doc and money_equal(row.get("invoice_sum"), erp_doc.get("sum")) and money_equal(row.get("invoice_sum"), onec_doc.get("sum")) else "FIELDS_MISMATCH",
            }
        )

    upd_refs = []
    for row in rows:
        upd_refs.extend(row.get("upd_refs") if isinstance(row.get("upd_refs"), list) else [])
    onec_upd_doc = find_onec_act_by_control_ref(onec_docs, upd_refs, totals.get("non_reimbursable_total"))
    if onec_upd_doc:
        document_rows.append(
            {
                "kind": "non_reimbursable_upd",
                "sheet_code": "; ".join(ref.get("text", "") for ref in upd_refs if ref.get("text"))[:200],
                "sheet_sum": totals.get("non_reimbursable_total"),
                "erp_code1c": "",
                "erp_sum": settlements.get("control_non_reimbursable_total"),
                "onec_code1c": onec_upd_doc.get("code1c"),
                "onec_sum": onec_upd_doc.get("sum"),
                "status": "MATCH" if money_equal(totals.get("non_reimbursable_total"), onec_upd_doc.get("sum")) else "FIELDS_MISMATCH",
            }
        )

    onec_payment_matches = [
        doc for doc in onec_docs
        if normalize_text(doc.get("kind")) == "bank_receipt" and money_equal(doc.get("sum"), totals.get("payment_total"))
    ]

    erp_invoice_total_by_sheet_codes = sum_values(erp_invoice_docs, "sum")
    onec_invoice_total_by_sheet_codes = sum_values(onec_invoice_docs, "sum")
    onec_payment_total = sum_values(onec_payment_matches, "sum")
    onec_non_reimbursable_total = normalize_sum(onec_upd_doc.get("sum")) if onec_upd_doc else None

    metrics = [
        metric_row("Счета из блока Sheet", totals.get("invoice_total"), erp_invoice_total_by_sheet_codes, onec_invoice_total_by_sheet_codes, "Сравнение только по кодам счетов из ручной таблицы"),
        metric_row("Оплаты покупателя", totals.get("payment_total"), settlements.get("paid_total"), onec_payment_total, "1C берется по поступлению с суммой блока и договорным контекстом"),
        metric_row("Возмещаемые расходы", totals.get("reimbursable_total"), settlements.get("control_reimbursable_total"), None, "ERP берется по get_expensessum без f_addnds"),
        metric_row("Невозмещаемые услуги", totals.get("non_reimbursable_total"), settlements.get("control_non_reimbursable_total"), onec_non_reimbursable_total, "ERP берется по операциям без f_addnds; 1C - по УПД из колонки № счф"),
        metric_row("(+/-) контрольной сверки", totals.get("delta"), settlements.get("control_sheet_delta"), None, "Формула Sheet: оплаты - возмещаемые - невозмещаемые"),
    ]

    return {
        "ok": True,
        "spec_id": spec_id,
        "delivery": control.get("delivery"),
        "sheet": control.get("sheet"),
        "source_state": {"onec": onec_state, "sheet": "google_sheet"},
        "control_sheet": {
            "found": True,
            "start_row": control.get("start_row"),
            "end_row": control.get("end_row"),
            "rows": rows,
            "totals": totals,
        },
        "erp": {
            "settlements": settlements,
            "add_nds_operations": [
                op for op in (snapshot.get("operations") if isinstance(snapshot.get("operations"), list) else [])
                if sql_int(op.get("add_nds_flag")) == 1
            ],
            "control_excluded_cost_operations": [
                op for op in (snapshot.get("operations") if isinstance(snapshot.get("operations"), list) else [])
                if sql_int(op.get("reimbursement_id")) != 1
                and (normalize_sum(op.get("rp_expenses_sum")) or 0.0) > 0
                and (normalize_sum(op.get("rp_profit_sum")) or 0.0) < 0
            ],
        },
        "onec": {
            "docs_count": len(onec_docs),
            "invoice_docs_count": len(onec_invoice_docs),
            "payment_matches_count": len(onec_payment_matches),
        },
        "metrics": metrics,
        "documents": document_rows,
    }


def erp_item_from_doc(row: dict[str, object], kind: str) -> dict[str, object]:
    amount = row.get("sum")
    payment_full_sum = None
    if kind == "payment":
        classified_sum = normalize_sum(row.get("classified_sum"))
        payment_full_sum = normalize_sum(row.get("payment_sum"))
        amount = classified_sum if classified_sum is not None else payment_full_sum
    item = {
        "kind": kind,
        "oper_id": row.get("oper_id"),
        "operation_url": row.get("operation_url") or (OPER_URL_TEMPLATE.format(oper_id=row.get("oper_id")) if row.get("oper_id") else ""),
        "operation_title": row.get("oper_type_name") or row.get("operation_title") or "",
        "erp_doc_id": row.get("erp_doc_id") or row.get("payment_id") or 0,
        "code1c": row.get("code1c") or "",
        "number": row.get("number") or row.get("pp_number") or "",
        "date_iso": parse_any_date_to_iso(row.get("date_iso") or row.get("date") or row.get("payment_date")),
        "date": row.get("date") or row.get("payment_date") or "",
        "sum": normalize_sum(amount),
        "type_id": row.get("type_id") or kind,
        "type_name": "Поступление оплаты" if kind == "payment" else row.get("type_name") or row.get("oper_type_name") or kind,
        "status_id": row.get("status_id") or 0,
        "status_name": row.get("status_name") or "",
        "onec_export_state": row.get("onec_export_state") or "",
        "dog_number": row.get("dog_number") or "",
        "dog_code1c": row.get("dog_code1c") or "",
        "invoice_number": row.get("invoice_number") or row.get("number") if kind == "invoice" else row.get("invoice_number") or "",
        "reimbursement_name": row.get("reimbursement_name") or "",
        "currency": row.get("currency") or row.get("doc_currency") or "",
    }
    if kind == "payment" and payment_full_sum is not None:
        item["payment_full_sum"] = payment_full_sum
    return item


def compact_join(values: set[str], limit: int = 4) -> str:
    cleaned = sorted({normalize_text(value) for value in values if normalize_text(value)})
    if not cleaned:
        return ""
    return "; ".join(cleaned[:limit]) + (f"; +{len(cleaned) - limit}" if len(cleaned) > limit else "")


def is_non_reimbursable(row: dict[str, object]) -> bool:
    reimbursement_id = sql_int(row.get("reimbursement_id"))
    reimbursement_name = normalize_text(row.get("reimbursement_name")).lower()
    return (reimbursement_id > 0 and reimbursement_id != 1) or "невозмещ" in reimbursement_name


def act_group_key(row: dict[str, object]) -> tuple[object, ...]:
    main_code = normalize_text(row.get("main_code1c"))
    main_date_iso = parse_any_date_to_iso(row.get("main_date_iso") or row.get("main_date"))
    main_doc_id = sql_int(row.get("main_erp_doc_id"))
    code = main_code or normalize_text(row.get("code1c"))
    date_iso = main_date_iso or parse_any_date_to_iso(row.get("date_iso") or row.get("date"))
    doc_id = sql_int(row.get("erp_doc_id"))

    if is_non_reimbursable(row):
        return ("act", "non_reimbursable_upd", compact_key(code) or (f"main:{main_doc_id}" if main_doc_id else "no_code"), date_iso)

    if code and date_iso:
        return ("act", "code_date", compact_key(code), date_iso)

    if doc_id > 0:
        return ("act", "erp_doc", doc_id)

    return ("act", "row", row.get("oper_id"), doc_id, date_iso)


def aggregate_act_compare_items(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: dict[tuple[object, ...], dict[str, object]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = act_group_key(row)
        item = groups.get(key)
        if item is None:
            item = erp_item_from_doc(row, "act")
            main_doc_id = sql_int(row.get("main_erp_doc_id"))
            main_code = normalize_text(row.get("main_code1c"))
            main_number = normalize_text(row.get("main_number"))
            main_date_iso = parse_any_date_to_iso(row.get("main_date_iso") or row.get("main_date"))
            main_date = normalize_text(row.get("main_date")) or display_date_from_iso(main_date_iso)
            if main_doc_id > 0:
                item["erp_doc_id"] = main_doc_id
            if main_code:
                item["code1c"] = main_code
            if main_number:
                item["number"] = main_number
            if main_date_iso:
                item["date_iso"] = main_date_iso
                item["date"] = main_date
            item["_doc_ids"] = set()
            item["_main_doc_ids"] = set()
            item["_main_sums"] = {}
            item["_detail_ids"] = set()
            item["_operation_titles"] = set()
            item["_reimbursements"] = set()
            item["_numbers"] = set()
            item["_main_numbers"] = set()
            item["_dog_numbers"] = set()
            item["_dog_codes"] = set()
            item["_sum"] = 0.0
            item["is_aggregated"] = False
            groups[key] = item

        doc_id = sql_int(row.get("erp_doc_id"))
        main_doc_id = sql_int(row.get("main_erp_doc_id"))
        main_sum = normalize_sum(row.get("main_sum"))
        main_number = normalize_text(row.get("main_number"))
        detail_id = sql_int(row.get("detail_id"))
        detail_sum = normalize_sum(row.get("detail_sum"))
        doc_sum = normalize_sum(row.get("sum"))

        if detail_id > 0:
            if detail_id not in item["_detail_ids"]:
                item["_detail_ids"].add(detail_id)
                if detail_sum:
                    item["_sum"] += detail_sum
                elif doc_id > 0 and doc_id not in item["_doc_ids"]:
                    item["_sum"] += doc_sum or 0.0
        elif doc_id > 0 and doc_id not in item["_doc_ids"]:
            item["_sum"] += doc_sum or 0.0
        elif doc_id <= 0:
            item["_sum"] += doc_sum or 0.0

        if doc_id > 0:
            item["_doc_ids"].add(doc_id)
        if main_doc_id > 0:
            item["_main_doc_ids"].add(main_doc_id)
            if main_sum is not None:
                item["_main_sums"][main_doc_id] = main_sum
        if normalize_text(row.get("oper_type_name")):
            item["_operation_titles"].add(normalize_text(row.get("oper_type_name")))
        if normalize_text(row.get("reimbursement_name")):
            item["_reimbursements"].add(normalize_text(row.get("reimbursement_name")))
        if main_number:
            item["_main_numbers"].add(main_number)
        elif normalize_text(row.get("number")):
            item["_numbers"].add(normalize_text(row.get("number")))
        if normalize_text(row.get("dog_number")):
            item["_dog_numbers"].add(normalize_text(row.get("dog_number")))
        if normalize_text(row.get("dog_code1c")):
            item["_dog_codes"].add(normalize_text(row.get("dog_code1c")))

    result: list[dict[str, object]] = []
    for item in groups.values():
        doc_ids = item.pop("_doc_ids", set())
        main_doc_ids = item.pop("_main_doc_ids", set())
        main_sums = item.pop("_main_sums", {})
        detail_ids = item.pop("_detail_ids", set())
        titles = item.pop("_operation_titles", set())
        reimbursements = item.pop("_reimbursements", set())
        numbers = item.pop("_numbers", set())
        main_numbers = item.pop("_main_numbers", set())
        dog_numbers = item.pop("_dog_numbers", set())
        dog_codes = item.pop("_dog_codes", set())
        total = item.pop("_sum", item.get("sum"))

        if main_sums:
            total = sum(float(value or 0.0) for value in main_sums.values())
        item["sum"] = round(float(total or 0.0), 2)
        if titles:
            item["operation_title"] = compact_join(titles)
        if reimbursements:
            item["reimbursement_name"] = compact_join(reimbursements, limit=6)
        if main_numbers:
            item["number"] = compact_join(main_numbers)
            if numbers:
                item["child_numbers"] = compact_join(numbers)
        elif numbers:
            item["number"] = compact_join(numbers)
        if dog_numbers:
            item["dog_number"] = compact_join(dog_numbers, limit=3)
        if dog_codes:
            item["dog_code1c"] = compact_join(dog_codes, limit=3)

        item["erp_doc_ids"] = sorted(main_doc_ids | doc_ids)
        item["erp_child_doc_ids"] = sorted(doc_ids)
        item["erp_detail_ids"] = sorted(detail_ids)
        item["erp_doc_count"] = len(main_doc_ids | doc_ids)
        item["erp_detail_count"] = len(detail_ids)
        item["is_aggregated"] = bool(main_doc_ids) or len(doc_ids) > 1 or len(detail_ids) > 1 or len(titles) > 1
        if item["is_aggregated"]:
            item["aggregation_note"] = (
                "ERP-строки агрегированы для сверки с одним документом 1С; "
                f"основных документов ERP: {len(main_doc_ids)}, дочерних документов ERP: {len(doc_ids)}, деталей: {len(detail_ids)}"
            )
            if item.get("reimbursement_name"):
                item["type_name"] = f"{item.get('reimbursement_name')} (сводно)"
        result.append(item)

    return result


def build_erp_compare_items(snapshot: dict[str, object]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    for row in snapshot.get("schets") if isinstance(snapshot.get("schets"), list) else []:
        if isinstance(row, dict):
            if not normalize_text(row.get("code1c")) and normalize_text(row.get("invoice_number")):
                continue
            items.append(erp_item_from_doc(row, "invoice"))
    payment_items: dict[object, dict[str, object]] = {}
    payment_titles: dict[object, set[str]] = {}
    payment_reimbursements: dict[object, set[str]] = {}
    for row in snapshot.get("payments") if isinstance(snapshot.get("payments"), list) else []:
        if not isinstance(row, dict):
            continue
        if normalize_text(row.get("direction")) != "incoming":
            continue
        key = row.get("payment_id") or (
            row.get("code1c"),
            row.get("pp_number"),
            row.get("payment_date_iso"),
            row.get("direction"),
        )
        if key not in payment_items:
            payment_items[key] = erp_item_from_doc(row, "payment")
            payment_titles[key] = set()
            payment_reimbursements[key] = set()
        else:
            current_sum = normalize_sum(payment_items[key].get("sum")) or 0.0
            classified_sum = normalize_sum(row.get("classified_sum"))
            if classified_sum is not None:
                payment_items[key]["sum"] = round(current_sum + classified_sum, 2)
        if normalize_text(row.get("oper_type_name")):
            payment_titles[key].add(normalize_text(row.get("oper_type_name")))
        if normalize_text(row.get("reimbursement_name")):
            payment_reimbursements[key].add(normalize_text(row.get("reimbursement_name")))
    for key, item in payment_items.items():
        titles = sorted(payment_titles.get(key, set()))
        reimbursements = sorted(payment_reimbursements.get(key, set()))
        if titles:
            item["operation_title"] = "; ".join(titles[:4]) + (f"; +{len(titles) - 4}" if len(titles) > 4 else "")
        if reimbursements:
            item["reimbursement_name"] = "; ".join(reimbursements)
        payment_sum = normalize_sum(item.get("payment_full_sum"))
        spec_sum = normalize_sum(item.get("sum"))
        if payment_sum is not None and spec_sum is not None and abs(payment_sum - spec_sum) > 0.01:
            item["aggregation_note"] = f"Часть платежа по заявке: {spec_sum:.2f} из платежа {payment_sum:.2f}"
        items.append(item)
    act_source = snapshot.get("akts") if isinstance(snapshot.get("akts"), list) else []
    act_rows = [row for row in act_source if isinstance(row, dict)]
    act_items = aggregate_act_compare_items(act_rows)
    coded_non_reimb_dates = {
        parse_any_date_to_iso(item.get("date_iso") or item.get("date"))
        for item in act_items
        if normalize_text(item.get("code1c")) and is_non_reimbursable(item)
    }
    for item in act_items:
        item_date = parse_any_date_to_iso(item.get("date_iso") or item.get("date"))
        if (
            not normalize_text(item.get("code1c"))
            and is_non_reimbursable(item)
            and item_date in coded_non_reimb_dates
        ):
            continue
        items.append(item)
    return items


def erp_item_has_primary_key(erp_item: dict[str, object]) -> bool:
    return bool(
        strong_doc_number_keys(erp_item.get("code1c"))
        and parse_any_date_to_iso(erp_item.get("date_iso") or erp_item.get("date"))
    )


def onec_kind_compatible(erp_kind: str, onec_kind: str) -> bool:
    if erp_kind == "invoice":
        return onec_kind in {"invoice"}
    if erp_kind == "payment":
        return onec_kind in {"bank_receipt"}
    if erp_kind == "act":
        return onec_kind in {"sale_act", "purchase_act"}
    return True


def score_onec_candidate(erp_item: dict[str, object], onec_doc: dict[str, object]) -> tuple[int, list[str], list[str]]:
    score = 0
    evidence: list[str] = []
    mismatch: list[str] = []

    if not onec_kind_compatible(normalize_text(erp_item.get("kind")), normalize_text(onec_doc.get("kind"))):
        return -1000, ["kind"], ["kind"]
    erp_date = parse_any_date_to_iso(erp_item.get("date_iso") or erp_item.get("date"))
    onec_date = parse_any_date_to_iso(onec_doc.get("date_iso") or onec_doc.get("date"))
    dates_match = bool(erp_date and onec_date and erp_date == onec_date)
    dates_mismatch = bool(erp_date and onec_date and erp_date != onec_date)

    erp_code_keys = strong_doc_number_keys(erp_item.get("code1c"))
    erp_num_keys = strong_doc_number_keys(erp_item.get("number"))
    onec_code_keys = strong_doc_number_keys(onec_doc.get("code1c"))
    onec_num_keys = strong_doc_number_keys(onec_doc.get("number"))
    if erp_code_keys and onec_code_keys and erp_code_keys.intersection(onec_code_keys):
        if dates_match:
            score += 180
            evidence.append("code1c")
            evidence.append("date")
        else:
            mismatch.append("date")
    elif erp_code_keys and onec_code_keys:
        mismatch.append("code1c")
    elif erp_num_keys and (onec_code_keys | onec_num_keys) and erp_num_keys.intersection(onec_code_keys | onec_num_keys):
        score += 30
        evidence.append("document_number")

    erp_invoice_keys = strong_doc_number_keys(erp_item.get("invoice_number"))
    onec_invoice_keys = strong_doc_number_keys(onec_doc.get("invoice_number"))
    if normalize_text(erp_item.get("kind")) != "invoice":
        if erp_invoice_keys and onec_invoice_keys and erp_invoice_keys.intersection(onec_invoice_keys):
            score += 70
            evidence.append("invoice_number")
        elif erp_invoice_keys and onec_invoice_keys:
            mismatch.append("invoice_number")

    erp_contract_keys = {key for key in {compact_key(erp_item.get("dog_number")), compact_key(erp_item.get("dog_code1c"))} if key}
    onec_contract_keys = {key for key in {compact_key(onec_doc.get("base_contract")), compact_key(onec_doc.get("contract_number"))} if key}
    if erp_contract_keys.intersection(onec_contract_keys):
        score += 35
        evidence.append("contract")
    elif erp_contract_keys and onec_contract_keys:
        mismatch.append("contract")

    if onec_doc.get("match_reasons"):
        score += 15
        evidence.extend([f"context:{reason}" for reason in onec_doc.get("match_reasons") if reason])

    amount_matches = money_equal(erp_item.get("sum"), onec_doc.get("sum"))
    if (
        not amount_matches
        and normalize_text(erp_item.get("kind")) == "payment"
        and normalize_text(onec_doc.get("kind")) == "bank_receipt"
        and money_equal(erp_item.get("payment_full_sum"), onec_doc.get("sum"))
    ):
        amount_matches = True

    if amount_matches:
        score += 25
        evidence.append("sum")
    else:
        if normalize_sum(erp_item.get("sum")) is not None and normalize_sum(onec_doc.get("sum")) is not None:
            mismatch.append("sum")

    if dates_match and "date" not in evidence:
        score += 10
        evidence.append("date")
    elif dates_mismatch:
        mismatch.append("date")

    primary_match = "code1c" in evidence
    if not primary_match:
        return -1000, evidence, list(dict.fromkeys(mismatch or ["kod1c", "date"]))

    if score < 40:
        return score, evidence, list(dict.fromkeys(mismatch or ["no_strong_key"]))
    return score, evidence, list(dict.fromkeys(mismatch))


def account_movement_side_matches_delivery(movement: dict[str, object], side: str, snapshot: dict[str, object]) -> bool:
    delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
    spec_number = normalize_text(delivery.get("spec_number"))
    base_contract_key = compact_key(delivery.get("main_dog_number"))
    if not spec_number:
        return False

    side_spec = normalize_text(movement.get(f"{side}_spec_number"))
    side_base_key = compact_key(movement.get(f"{side}_base_contract") or movement.get(f"{side}_contract_number"))
    if side_spec == spec_number and (not base_contract_key or side_base_key == base_contract_key):
        return True

    side_contract_key = compact_key(movement.get(f"{side}_contract"))
    spec_contract_key = compact_key(f"{delivery.get('main_dog_number')}/{spec_number}")
    return bool(spec_contract_key and side_contract_key and spec_contract_key in side_contract_key)


def account_movement_registrar_matches_erp(erp_item: dict[str, object], movement: dict[str, object]) -> bool:
    erp_date = parse_any_date_to_iso(erp_item.get("date_iso") or erp_item.get("date"))
    movement_date = parse_any_date_to_iso(movement.get("registrar_date_iso") or movement.get("period_iso") or movement.get("period"))
    if erp_date and movement_date and erp_date != movement_date:
        return False

    erp_code_keys = strong_doc_number_keys(erp_item.get("code1c"))
    movement_code_keys = strong_doc_number_keys(movement.get("registrar_number"))
    return bool(erp_code_keys and movement_code_keys and erp_code_keys.intersection(movement_code_keys))


def opposite_account_side(side: str) -> str:
    return "credit" if side == "debit" else "debit"


def account_movement_is_usable_for_erp_kind(
    erp_item: dict[str, object],
    movement: dict[str, object],
    side: str,
    snapshot: dict[str, object],
) -> bool:
    erp_kind = normalize_text(erp_item.get("kind"))
    if erp_kind == "payment":
        return True
    if erp_kind != "act":
        return False

    other_side = opposite_account_side(side)
    if account_movement_side_matches_delivery(movement, other_side, snapshot):
        return False
    side_account = normalize_text(movement.get(f"{side}_account"))
    other_account = normalize_text(movement.get(f"{other_side}_account"))
    if side_account and other_account and side_account == other_account:
        return False
    return True


def build_onec_doc_from_account_movements(
    erp_item: dict[str, object],
    movements: list[dict[str, object]],
    side: str,
    amount: float,
) -> dict[str, object]:
    first = movements[0] if movements else {}
    source_files = compact_join({normalize_text(movement.get("source_file")) for movement in movements}, limit=2)
    source_rows = sorted({pg_int_or_zero(movement.get("source_row")) for movement in movements if pg_int_or_zero(movement.get("source_row"))})
    contract_texts = {
        normalize_text(movement.get(f"{side}_contract"))
        for movement in movements
        if normalize_text(movement.get(f"{side}_contract"))
    }
    return {
        "row_num": source_rows[0] if source_rows else 0,
        "source_file": source_files or normalize_text(first.get("source_file")),
        "source_sheet": normalize_text(first.get("source_sheet")),
        "source_row": source_rows[0] if source_rows else pg_int_or_zero(first.get("source_row")),
        "kind": "account_movement",
        "type": f"Карточка счета {normalize_text(first.get('account_scope'))}".strip(),
        "code1c": normalize_text(first.get("registrar_number")) or normalize_text(erp_item.get("code1c")),
        "ref_id": "",
        "number": normalize_text(first.get("registrar_number")) or normalize_text(erp_item.get("number")),
        "number_norm": normalize_number(first.get("registrar_number") or erp_item.get("number")),
        "date": display_date_from_iso(first.get("registrar_date_iso") or first.get("period_iso") or erp_item.get("date_iso")),
        "date_iso": parse_any_date_to_iso(first.get("registrar_date_iso") or first.get("period_iso") or erp_item.get("date_iso")),
        "sum": round(amount, 2),
        "organization": "",
        "counterparty": normalize_text(first.get(f"{side}_counterparty")),
        "contract": compact_join(contract_texts, limit=2) or normalize_text(first.get(f"{side}_contract")),
        "contract_key": compact_key(first.get(f"{side}_contract")),
        "request_type": "Заявка" if normalize_text(first.get(f"{side}_spec_number")) else "",
        "spec_number": normalize_text(first.get(f"{side}_spec_number")),
        "base_contract": normalize_text(first.get(f"{side}_base_contract")),
        "contract_number": normalize_text(first.get(f"{side}_contract_number")),
        "contract_date": "",
        "contract_date_iso": normalize_text(first.get(f"{side}_contract_date_iso")),
        "invoice_number": "",
        "invoice_number_norm": "",
        "comment": (
            f"Подтверждено карточкой счета {normalize_text(first.get('account_scope'))}; "
            f"строки: {', '.join(str(row) for row in source_rows[:8])}"
            + (f"; +{len(source_rows) - 8}" if len(source_rows) > 8 else "")
        ),
        "account_movement_rows": source_rows,
        "account_movement_count": len(movements),
    }


def document_line_matches_erp_act(erp_item: dict[str, object], line: dict[str, object]) -> bool:
    if normalize_text(erp_item.get("kind")) != "act":
        return False
    erp_date = parse_any_date_to_iso(erp_item.get("date_iso") or erp_item.get("date"))
    line_date = parse_any_date_to_iso(line.get("document_date_iso") or line.get("document_date"))
    if erp_date and line_date and erp_date != line_date:
        return False
    erp_code_keys = strong_doc_number_keys(erp_item.get("code1c"))
    line_code_keys = strong_doc_number_keys(line.get("document_code1c")) | strong_doc_number_keys(line.get("document_number"))
    return bool(erp_code_keys and line_code_keys and erp_code_keys.intersection(line_code_keys))


def build_onec_doc_from_document_lines(erp_item: dict[str, object], lines: list[dict[str, object]], amount: float) -> dict[str, object]:
    first = lines[0] if lines else {}
    source_files = compact_join({normalize_text(line.get("source_file")) for line in lines}, limit=2)
    source_rows = sorted({pg_int_or_zero(line.get("source_row")) for line in lines if pg_int_or_zero(line.get("source_row"))})
    contracts = {
        normalize_text(line.get("commissioner_contract"))
        for line in lines
        if normalize_text(line.get("commissioner_contract"))
    }
    commissioners = {
        normalize_text(line.get("commissioner"))
        for line in lines
        if normalize_text(line.get("commissioner"))
    }
    return {
        "row_num": source_rows[0] if source_rows else 0,
        "source_file": source_files or normalize_text(first.get("source_file")),
        "source_sheet": normalize_text(first.get("source_sheet")),
        "source_row": source_rows[0] if source_rows else pg_int_or_zero(first.get("source_row")),
        "kind": "document_line",
        "type": "Строки поступления 1С",
        "code1c": normalize_text(first.get("document_code1c")) or normalize_text(erp_item.get("code1c")),
        "ref_id": "",
        "number": normalize_text(first.get("document_number")) or normalize_text(erp_item.get("number")),
        "number_norm": normalize_number(first.get("document_number") or erp_item.get("number")),
        "date": display_date_from_iso(first.get("document_date_iso") or erp_item.get("date_iso")),
        "date_iso": parse_any_date_to_iso(first.get("document_date_iso") or erp_item.get("date_iso")),
        "sum": round(amount, 2),
        "organization": "",
        "counterparty": compact_join(commissioners, limit=2) or normalize_text(first.get("commissioner")),
        "contract": compact_join(contracts, limit=2) or normalize_text(first.get("commissioner_contract")),
        "contract_key": compact_key(first.get("commissioner_contract")),
        "request_type": "Заявка" if normalize_text(first.get("spec_number")) else "",
        "spec_number": normalize_text(first.get("spec_number")),
        "base_contract": normalize_text(first.get("base_contract")),
        "contract_number": normalize_text(first.get("contract_number")),
        "contract_date": normalize_text(first.get("contract_date")),
        "contract_date_iso": normalize_text(first.get("contract_date_iso")),
        "invoice_number": "",
        "invoice_number_norm": "",
        "comment": (
            "Подтверждено строковой детализацией документа 1С; "
            f"строки: {', '.join(str(row) for row in source_rows[:8])}"
            + (f"; +{len(source_rows) - 8}" if len(source_rows) > 8 else "")
        ),
        "document_line_rows": source_rows,
        "document_line_count": len(lines),
    }


def find_document_line_match(
    erp_item: dict[str, object],
    document_lines: list[dict[str, object]],
    snapshot: dict[str, object],
) -> tuple[dict[str, object] | None, list[str], list[str]]:
    if normalize_text(erp_item.get("kind")) != "act" or not document_lines:
        return None, [], []

    delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
    spec_number = normalize_text(delivery.get("spec_number"))
    base_contract_key = compact_key(delivery.get("main_dog_number"))
    strict_matched = []
    loose_matched = []
    for line in document_lines:
        if line.get("_used"):
            continue
        if not document_line_matches_erp_act(erp_item, line):
            continue
        line_spec = normalize_text(line.get("spec_number"))
        line_base_key = compact_key(line.get("base_contract") or line.get("contract_number"))
        has_contract_context = bool(line_spec or line_base_key)
        if has_contract_context:
            if spec_number and line_spec and line_spec != spec_number:
                continue
            if base_contract_key and line_base_key and line_base_key != base_contract_key:
                continue
            strict_matched.append(line)
        else:
            loose_matched.append(line)

    matched = strict_matched
    missing_contract_context = False
    if not matched and loose_matched:
        target_sum = normalize_sum(erp_item.get("sum"))
        exact_lines = [line for line in loose_matched if target_sum is not None and money_equal(line.get("sum"), target_sum)]
        if exact_lines:
            matched = [exact_lines[0]]
        else:
            loose_total = round(sum(float(normalize_sum(line.get("sum")) or 0.0) for line in loose_matched), 2)
            matched = loose_matched if money_equal(loose_total, target_sum) else loose_matched
        missing_contract_context = True

    if not matched:
        return None, [], []

    total = round(sum(float(normalize_sum(line.get("sum")) or 0.0) for line in matched), 2)
    onec_doc = build_onec_doc_from_document_lines(erp_item, matched, total)
    evidence = ["code1c", "date", "document_line"]
    if missing_contract_context:
        mismatch = [] if money_equal(total, erp_item.get("sum")) else ["sum"]
        if "sum" not in mismatch:
            evidence.append("sum")
        mismatch.append("contract")
    else:
        evidence.extend(["contract", "spec_number"])
        mismatch = [] if money_equal(total, erp_item.get("sum")) else ["sum"]
    if not mismatch:
        evidence.append("sum")
    return onec_doc, evidence, mismatch


def mark_document_lines_used(matched_doc: dict[str, object], document_lines: list[dict[str, object]]) -> None:
    rows = {pg_int_or_zero(row) for row in matched_doc.get("document_line_rows", []) if pg_int_or_zero(row)}
    doc_date = parse_any_date_to_iso(matched_doc.get("date_iso") or matched_doc.get("date"))
    doc_keys = strong_doc_number_keys(matched_doc.get("code1c")) | strong_doc_number_keys(matched_doc.get("number"))
    for line in document_lines:
        if rows and pg_int_or_zero(line.get("source_row")) not in rows:
            continue
        line_date = parse_any_date_to_iso(line.get("document_date_iso") or line.get("document_date"))
        if doc_date and line_date and doc_date != line_date:
            continue
        line_keys = strong_doc_number_keys(line.get("document_code1c")) | strong_doc_number_keys(line.get("document_number"))
        if doc_keys and line_keys and not doc_keys.intersection(line_keys):
            continue
        line["_used"] = True


def find_account_movement_document_match(
    erp_item: dict[str, object],
    account_movements: list[dict[str, object]],
    snapshot: dict[str, object],
) -> tuple[dict[str, object] | None, list[str], list[str]]:
    if normalize_text(erp_item.get("kind")) not in {"payment", "act"}:
        return None, [], []
    if not account_movements:
        return None, [], []

    matched_by_side: dict[str, list[dict[str, object]]] = {"debit": [], "credit": []}
    for movement in account_movements:
        if not account_movement_registrar_matches_erp(erp_item, movement):
            continue
        for side in ["debit", "credit"]:
            if account_movement_side_matches_delivery(movement, side, snapshot):
                if not account_movement_is_usable_for_erp_kind(erp_item, movement, side, snapshot):
                    continue
                amount = normalize_sum(movement.get(f"{side}_amount"))
                if amount is not None and amount != 0:
                    matched_by_side[side].append(movement)

    best_side = ""
    best_movements: list[dict[str, object]] = []
    best_total = 0.0
    for side, movements in matched_by_side.items():
        total = round(sum(float(normalize_sum(movement.get(f"{side}_amount")) or 0.0) for movement in movements), 2)
        if money_equal(total, erp_item.get("sum")):
            best_side = side
            best_movements = movements
            best_total = total
            break
        if movements and (not best_movements or abs(total - float(normalize_sum(erp_item.get("sum")) or 0.0)) < abs(best_total - float(normalize_sum(erp_item.get("sum")) or 0.0))):
            best_side = side
            best_movements = movements
            best_total = total

    if not best_movements:
        return None, [], []

    onec_doc = build_onec_doc_from_account_movements(erp_item, best_movements, best_side, best_total)
    evidence = ["code1c", "date", "account_movement", "contract", "spec_number"]
    mismatch = [] if money_equal(best_total, erp_item.get("sum")) else ["sum"]
    if not mismatch:
        evidence.append("sum")
    return onec_doc, evidence, mismatch


def find_account_movement_payment_match(
    erp_item: dict[str, object],
    account_movements: list[dict[str, object]],
    snapshot: dict[str, object],
) -> tuple[dict[str, object] | None, list[str], list[str]]:
    if normalize_text(erp_item.get("kind")) != "payment":
        return None, [], []
    return find_account_movement_document_match(erp_item, account_movements, snapshot)


def mark_related_onec_docs_used(erp_item: dict[str, object], onec_docs: list[dict[str, object]]) -> None:
    erp_date = parse_any_date_to_iso(erp_item.get("date_iso") or erp_item.get("date"))
    erp_code_keys = strong_doc_number_keys(erp_item.get("code1c"))
    if not erp_date or not erp_code_keys:
        return
    erp_kind = normalize_text(erp_item.get("kind"))
    if erp_kind == "payment":
        compatible_kinds = {"bank_receipt"}
    elif erp_kind == "act":
        compatible_kinds = {"sale_act", "purchase_act"}
    else:
        compatible_kinds = set()
    for onec_doc in onec_docs:
        if normalize_text(onec_doc.get("kind")) not in compatible_kinds:
            continue
        onec_date = parse_any_date_to_iso(onec_doc.get("date_iso") or onec_doc.get("date"))
        if onec_date and onec_date != erp_date:
            continue
        onec_keys = strong_doc_number_keys(onec_doc.get("code1c")) | strong_doc_number_keys(onec_doc.get("number"))
        if erp_code_keys.intersection(onec_keys):
            onec_doc["_used"] = True


def mark_related_bank_receipts_used(erp_item: dict[str, object], onec_docs: list[dict[str, object]]) -> None:
    mark_related_onec_docs_used(erp_item, onec_docs)


def build_reconciliation_row(
    status: str,
    erp_item: dict[str, object] | None,
    onec_doc: dict[str, object] | None,
    note: str,
    mismatch_fields: list[str] | None = None,
    evidence: list[str] | None = None,
) -> dict[str, object]:
    erp_item = erp_item or {}
    onec_doc = onec_doc or {}
    return {
        "status": status,
        "note": note,
        "mismatch_fields": mismatch_fields or [],
        "match_evidence": evidence or [],
        "oper_id": erp_item.get("oper_id") or 0,
        "operation_url": erp_item.get("operation_url") or "",
        "operation_title": erp_item.get("operation_title") or "",
        "erp_doc_id": erp_item.get("erp_doc_id") or 0,
        "erp_doc_ids": erp_item.get("erp_doc_ids") or [],
        "erp_doc_count": erp_item.get("erp_doc_count") or 0,
        "erp_detail_count": erp_item.get("erp_detail_count") or 0,
        "erp_is_aggregated": bool(erp_item.get("is_aggregated")),
        "erp_aggregation_note": erp_item.get("aggregation_note") or "",
        "erp_doc_kind": erp_item.get("kind") or "",
        "erp_code1c": erp_item.get("code1c") or "",
        "erp_number": erp_item.get("number") or "",
        "erp_date_iso": parse_any_date_to_iso(erp_item.get("date_iso") or erp_item.get("date")),
        "erp_sum": erp_item.get("sum"),
        "erp_currency": erp_item.get("currency") or "",
        "erp_type": erp_item.get("type_name") or erp_item.get("type_id") or "",
        "erp_status_id": erp_item.get("status_id") or 0,
        "erp_status_name": erp_item.get("status_name") or "",
        "erp_onec_export_state": erp_item.get("onec_export_state") or "",
        "erp_dog_number": erp_item.get("dog_number") or "",
        "erp_dog_code1c": erp_item.get("dog_code1c") or "",
        "erp_invoice_number": erp_item.get("invoice_number") or "",
        "erp_reimbursement": erp_item.get("reimbursement_name") or "",
        "onec_code1c": onec_doc.get("code1c") or "",
        "onec_ref_id": onec_doc.get("ref_id") or "",
        "onec_number": onec_doc.get("number") or "",
        "onec_date_iso": parse_any_date_to_iso(onec_doc.get("date_iso") or onec_doc.get("date")),
        "onec_sum": onec_doc.get("sum"),
        "onec_type": onec_doc.get("type") or onec_doc.get("kind") or "",
        "onec_doc_kind": onec_doc.get("kind") or "",
        "onec_contract": onec_doc.get("contract") or "",
        "onec_base_contract": onec_doc.get("base_contract") or "",
        "onec_spec_number": onec_doc.get("spec_number") or "",
        "onec_invoice_number": onec_doc.get("invoice_number") or "",
        "onec_source_file": onec_doc.get("source_file") or "",
        "onec_source_row": onec_doc.get("source_row") or 0,
    }


def summarize_compare_rows(rows: list[dict[str, object]]) -> dict[str, int]:
    summary = summarize_results(rows)
    summary["FIELDS_MISMATCH_SUM"] = len([row for row in rows if row.get("status") == STATUS_FIELDS_MISMATCH and "sum" in (row.get("mismatch_fields") or [])])
    summary["FIELDS_MISMATCH_CONTRACT"] = len([row for row in rows if row.get("status") == STATUS_FIELDS_MISMATCH and "contract" in (row.get("mismatch_fields") or [])])
    summary["FIELDS_MISMATCH_INVOICE"] = len([row for row in rows if row.get("status") == STATUS_FIELDS_MISMATCH and "invoice_number" in (row.get("mismatch_fields") or [])])
    return summary


def compare_onec_docs_with_erp_snapshot(spec_id: int, onec_source: dict[str, object], snapshot: dict[str, object]) -> dict[str, object]:
    erp_items = build_erp_compare_items(snapshot)
    onec_docs = onec_source.get("docs") if isinstance(onec_source.get("docs"), list) else []
    account_movements = onec_source.get("account_movements") if isinstance(onec_source.get("account_movements"), list) else []
    document_lines = onec_source.get("document_lines") if isinstance(onec_source.get("document_lines"), list) else []
    rows: list[dict[str, object]] = []

    for erp_item in erp_items:
        if not erp_item_has_primary_key(erp_item):
            rows.append(
                build_reconciliation_row(
                    STATUS_NOT_COMPARABLE,
                    erp_item,
                    None,
                    "ERP-документ не участвует в сверке: нет kod1c или даты для надежной связки с 1С",
                    ["kod1c", "date"],
                )
            )
            continue

        is_partial_payment = (
            normalize_text(erp_item.get("kind")) == "payment"
            and normalize_sum(erp_item.get("payment_full_sum")) is not None
            and normalize_sum(erp_item.get("sum")) is not None
            and not money_equal(erp_item.get("payment_full_sum"), erp_item.get("sum"))
        )
        if is_partial_payment:
            movement_doc, movement_evidence, movement_mismatch = find_account_movement_payment_match(erp_item, account_movements, snapshot)
            if movement_doc and not movement_mismatch:
                mark_related_bank_receipts_used(erp_item, onec_docs)
                rows.append(build_reconciliation_row(STATUS_MATCH, erp_item, movement_doc, "", [], movement_evidence))
                continue

        best_idx: int | None = None
        best_score = -1000
        best_mismatch: list[str] = []
        best_evidence: list[str] = []
        for idx, onec_doc in enumerate(onec_docs):
            if onec_doc.get("_used"):
                continue
            score, evidence, mismatch = score_onec_candidate(erp_item, onec_doc)
            if score > best_score:
                best_idx = idx
                best_score = score
                best_mismatch = mismatch
                best_evidence = evidence

        if best_idx is None or best_score < 40:
            line_doc, line_evidence, line_mismatch = find_document_line_match(erp_item, document_lines, snapshot)
            if line_doc:
                status = STATUS_FIELDS_MISMATCH if line_mismatch else STATUS_MATCH
                if status == STATUS_MATCH:
                    note = ""
                elif line_mismatch == ["contract"] or ("contract" in line_mismatch and "sum" not in line_mismatch):
                    note = "Строка 1С найдена по документу и сумме, но в выгрузке нет договора-заявки"
                else:
                    note = "Документ найден в строковой детализации 1С, но сумма строк отличается"
                if status == STATUS_MATCH or ("contract" in line_mismatch and "sum" not in line_mismatch):
                    mark_document_lines_used(line_doc, document_lines)
                    mark_related_onec_docs_used(erp_item, onec_docs)
                rows.append(build_reconciliation_row(status, erp_item, line_doc, note, line_mismatch, line_evidence))
                continue

            movement_doc, movement_evidence, movement_mismatch = find_account_movement_document_match(erp_item, account_movements, snapshot)
            if movement_doc:
                status = STATUS_FIELDS_MISMATCH if movement_mismatch else STATUS_MATCH
                note = "" if status == STATUS_MATCH else "Документ найден в карточке счета 1С, но сумма движений отличается"
                if status == STATUS_MATCH:
                    mark_related_onec_docs_used(erp_item, onec_docs)
                rows.append(build_reconciliation_row(status, erp_item, movement_doc, note, movement_mismatch, movement_evidence))
                continue

            rows.append(
                build_reconciliation_row(
                    STATUS_NOT_FOUND_IN_1C,
                    erp_item,
                    None,
                    "ERP-документ не найден в выгрузках 1С по коду, договору, счету или сумме",
                    ["match_key"],
                )
            )
            continue

        onec_doc = onec_docs[best_idx]
        if best_mismatch and "sum" in best_mismatch and normalize_text(erp_item.get("kind")) == "act":
            line_doc, line_evidence, line_mismatch = find_document_line_match(erp_item, document_lines, snapshot)
            if line_doc and (not line_mismatch or ("contract" in line_mismatch and "sum" not in line_mismatch)):
                onec_doc["_used"] = True
                mark_document_lines_used(line_doc, document_lines)
                mark_related_onec_docs_used(erp_item, onec_docs)
                if line_mismatch:
                    rows.append(
                        build_reconciliation_row(
                            STATUS_FIELDS_MISMATCH,
                            erp_item,
                            line_doc,
                            "Строка 1С найдена по документу и сумме, но в выгрузке нет договора-заявки",
                            line_mismatch,
                            line_evidence,
                        )
                    )
                else:
                    rows.append(build_reconciliation_row(STATUS_MATCH, erp_item, line_doc, "", [], line_evidence))
                continue

            movement_doc, movement_evidence, movement_mismatch = find_account_movement_document_match(erp_item, account_movements, snapshot)
            if movement_doc and not movement_mismatch:
                onec_doc["_used"] = True
                mark_related_onec_docs_used(erp_item, onec_docs)
                rows.append(build_reconciliation_row(STATUS_MATCH, erp_item, movement_doc, "", [], movement_evidence))
                continue

        onec_doc["_used"] = True
        status = STATUS_FIELDS_MISMATCH if best_mismatch else STATUS_MATCH
        note = "" if status == STATUS_MATCH else "Документ найден, но есть расхождения полей"
        rows.append(build_reconciliation_row(status, erp_item, onec_doc, note, best_mismatch, best_evidence))

    for onec_doc in onec_docs:
        if onec_doc.get("_used"):
            continue
        if normalize_text(onec_doc.get("kind")) == "contract":
            continue
        rows.append(
            build_reconciliation_row(
                STATUS_NOT_FOUND_IN_ERP,
                None,
                onec_doc,
                "Документ из выгрузки 1С не найден в ERP по текущей поставке",
                ["match_key"],
                onec_doc.get("match_reasons") if isinstance(onec_doc.get("match_reasons"), list) else [],
            )
        )

    return {
        "ok": True,
        "spec_id": spec_id,
        "run_id": None,
        "summary": summarize_compare_rows(rows),
        "rows": rows,
        "counts": {
            "erp_docs": len(erp_items),
            "onec_docs": len(onec_docs),
            "onec_all_docs": onec_source.get("all_docs_count", len(onec_docs)),
            "onec_account_movements": len(account_movements),
            "onec_all_account_movements": onec_source.get("all_account_movements_count", len(account_movements)),
            "onec_document_lines": len(document_lines),
            "onec_all_document_lines": onec_source.get("all_document_lines_count", len(document_lines)),
        },
        "source_state": onec_source.get("source_state", "xlsx"),
        "source": {
            "files": onec_source.get("files", []),
            "warnings": onec_source.get("warnings", []),
            "by_kind": onec_source.get("by_kind", {}),
            "contract_pairs": onec_source.get("contract_pairs", []),
            "filter": onec_source.get("filter", {}),
            "cache_hit": onec_source.get("cache_hit", False),
            "storage": onec_source.get("storage", onec_source.get("source_state", "xlsx")),
            "account_movements_count": len(account_movements),
            "all_account_movements_count": onec_source.get("all_account_movements_count", len(account_movements)),
            "document_lines_count": len(document_lines),
            "all_document_lines_count": onec_source.get("all_document_lines_count", len(document_lines)),
        },
    }


def build_not_comparable_rows(erp_docs: list[dict[str, object]], note: str) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for erp_doc in erp_docs:
        rows.append(
            {
                "status": STATUS_NOT_COMPARABLE,
                "note": note,
                "mismatch_fields": [],
                "oper_id": erp_doc.get("oper_id"),
                "operation_url": erp_doc.get("operation_url"),
                "operation_title": erp_doc.get("operation_title"),
                "erp_doc_id": erp_doc.get("erp_doc_id"),
                "erp_code1c": erp_doc.get("code1c"),
                "erp_number": erp_doc.get("number"),
                "erp_date_iso": erp_doc.get("date_iso"),
                "erp_sum": erp_doc.get("sum"),
                "erp_type": erp_doc.get("type_id"),
                "onec_code1c": "",
                "onec_number": "",
                "onec_date_iso": "",
                "onec_sum": None,
                "onec_type": "",
            }
        )
    return rows
    

def build_candidate_indices_by_code(onec_docs: list[dict[str, object]]) -> dict[str, list[int]]:
    indices: dict[str, list[int]] = {}
    for idx, item in enumerate(onec_docs):
        code = normalize_text(item.get("code1c"))
        if not code:
            continue
        indices.setdefault(code, []).append(idx)
    return indices


def compare_erp_to_onec(erp_doc: dict[str, object], onec_doc: dict[str, object]) -> list[str]:
    mismatch: list[str] = []

    erp_num = normalize_number(erp_doc.get("number"))
    onec_num = normalize_number(onec_doc.get("number"))
    if erp_num and onec_num and erp_num != onec_num:
        mismatch.append("number")

    erp_date = parse_any_date_to_iso(erp_doc.get("date_iso") or erp_doc.get("date"))
    onec_date = parse_any_date_to_iso(onec_doc.get("date_iso") or onec_doc.get("date"))
    if erp_date and onec_date and erp_date != onec_date:
        mismatch.append("date")

    erp_sum = normalize_sum(erp_doc.get("sum"))
    onec_sum = normalize_sum(onec_doc.get("sum"))
    if erp_sum is not None and onec_sum is not None and abs(erp_sum - onec_sum) > 0.01:
        mismatch.append("sum")

    erp_type_norm = normalize_number(erp_doc.get("type_id"))
    onec_type_norm = normalize_number(onec_doc.get("type"))
    if erp_type_norm and onec_type_norm and erp_type_norm != onec_type_norm:
        mismatch.append("type")

    return mismatch


def choose_best_onec_candidate(erp_doc: dict[str, object], onec_docs: list[dict[str, object]], candidates: list[int]) -> tuple[int | None, list[str]]:
    best_idx: int | None = None
    best_mismatch: list[str] = ["number", "date", "sum", "type", "fallback"]

    for idx in candidates:
        onec_doc = onec_docs[idx]
        if onec_doc.get("_used"):
            continue
        mismatch = compare_erp_to_onec(erp_doc, onec_doc)
        if best_idx is None or len(mismatch) < len(best_mismatch):
            best_idx = idx
            best_mismatch = mismatch
            if len(mismatch) == 0:
                break

    return best_idx, best_mismatch


def summarize_results(rows: list[dict[str, object]]) -> dict[str, int]:
    summary = {
        "total": len(rows),
        STATUS_MATCH: 0,
        STATUS_NOT_FOUND_IN_1C: 0,
        STATUS_NOT_FOUND_IN_ERP: 0,
        STATUS_FIELDS_MISMATCH: 0,
        STATUS_NOT_COMPARABLE: 0,
    }
    for row in rows:
        status = normalize_text(row.get("status"))
        if status in summary:
            summary[status] += 1
    return summary


def save_report_if_schema_exists(
    spec_id: int,
    onec_docs_count: int,
    report_rows: list[dict[str, object]],
    summary: dict[str, int],
    scope: str = "specification",
    scope_id: int = 0,
    client_id: int = 0,
    source_mode: str = "server-run",
) -> int | None:
    table_check_rows = run_mysql_tsv(
        """
SELECT
  SUM(CASE WHEN TABLE_NAME = 'veda_reconciliation_runs' THEN 1 ELSE 0 END) AS runs_exists,
  SUM(CASE WHEN TABLE_NAME = 'veda_reconciliation_items' THEN 1 ELSE 0 END) AS items_exists
FROM information_schema.TABLES
WHERE TABLE_SCHEMA = DATABASE()
  AND TABLE_NAME IN ('veda_reconciliation_runs', 'veda_reconciliation_items');
"""
    )
    if not table_check_rows:
        return None

    runs_exists = int_at(table_check_rows[0], 0)
    items_exists = int_at(table_check_rows[0], 1)
    if runs_exists <= 0 or items_exists <= 0:
        return None

    summary_json = json.dumps(summary, ensure_ascii=False)
    insert_run = f"""
INSERT INTO veda_reconciliation_runs
    (scope, scope_id, spec_id, client_id, source_mode, onec_docs_count, erp_docs_count, status, summary_json, created_at)
VALUES
    ({sql_quote(scope)}, {sql_int(scope_id)}, {sql_int(spec_id)}, {sql_int(client_id)}, {sql_quote(source_mode)}, {sql_int(onec_docs_count)}, {sql_int(len([r for r in report_rows if r.get('erp_doc_id')]))}, 'COMPLETED', {sql_quote(summary_json)}, NOW());
SELECT LAST_INSERT_ID();
"""
    run_rows = run_mysql_tsv(insert_run)
    run_id = int_at(run_rows[-1], 0) if run_rows else 0
    if run_id <= 0:
        return None

    values_parts: list[str] = []
    for row in report_rows:
        mismatch_json = json.dumps(row.get("mismatch_fields") or [], ensure_ascii=False)
        values_parts.append(
            "(" + ", ".join(
                [
                    str(run_id),
                    str(sql_int(row.get("oper_id"))),
                    str(sql_int(row.get("erp_doc_id"))),
                    sql_quote(row.get("erp_code1c")),
                    sql_quote(row.get("erp_number")),
                    sql_quote(row.get("erp_date_iso")),
                    sql_decimal(row.get("erp_sum")),
                    sql_quote(row.get("erp_type")),
                    sql_quote(row.get("onec_code1c")),
                    sql_quote(row.get("onec_number")),
                    sql_quote(row.get("onec_date_iso")),
                    sql_decimal(row.get("onec_sum")),
                    sql_quote(row.get("onec_type")),
                    sql_quote(row.get("status")),
                    sql_quote(mismatch_json),
                    sql_quote(row.get("note")),
                    "NOW()",
                ]
            )
            + ")"
        )

    if values_parts:
        insert_items = """
INSERT INTO veda_reconciliation_items
    (run_id, oper_id, erp_doc_id, erp_code1c, erp_number, erp_date_iso, erp_sum, erp_type,
     onec_code1c, onec_number, onec_date_iso, onec_sum, onec_type,
     status, mismatch_fields_json, note, created_at)
VALUES
""" + ",\n".join(values_parts) + ";"
        run_mysql_tsv(insert_items)

    return run_id


def compare_documents(
    spec_id: int,
    raw_onec_docs: object,
    source_state: str = "available",
    scope: str = "specification",
    scope_id: int = 0,
    client_id: int = 0,
    source_mode: str = "server-run",
) -> dict[str, object]:
    erp_docs = fetch_erp_docs(spec_id)
    onec_docs = normalize_onec_docs(raw_onec_docs)
    code_map = build_candidate_indices_by_code(onec_docs)

    report_rows: list[dict[str, object]] = []

    if source_state != "available" and len(onec_docs) == 0:
        report_rows = build_not_comparable_rows(erp_docs, "Источник данных 1С не подключен")
        summary = summarize_results(report_rows)
        run_id = None
        try:
            run_id = save_report_if_schema_exists(
                spec_id,
                0,
                report_rows,
                summary,
                scope=scope,
                scope_id=scope_id,
                client_id=client_id,
                source_mode=source_mode,
            )
        except Exception:
            run_id = None
        return {
            "summary": summary,
            "rows": report_rows,
            "run_id": run_id,
            "erp_docs_count": len(erp_docs),
            "onec_docs_count": 0,
            "source_state": source_state,
        }

    for erp_doc in erp_docs:
        erp_code = normalize_text(erp_doc.get("code1c"))
        if not erp_code:
            report_rows.append(
                {
                    "status": STATUS_NOT_COMPARABLE,
                    "note": "У ERP-документа нет кода 1С",
                    "mismatch_fields": ["code1c"],
                    "oper_id": erp_doc.get("oper_id"),
                    "operation_url": erp_doc.get("operation_url"),
                    "operation_title": erp_doc.get("operation_title"),
                    "erp_doc_id": erp_doc.get("erp_doc_id"),
                    "erp_code1c": erp_doc.get("code1c"),
                    "erp_number": erp_doc.get("number"),
                    "erp_date_iso": erp_doc.get("date_iso"),
                    "erp_sum": erp_doc.get("sum"),
                    "erp_type": erp_doc.get("type_id"),
                    "onec_code1c": "",
                    "onec_number": "",
                    "onec_date_iso": "",
                    "onec_sum": None,
                    "onec_type": "",
                }
            )
            continue

        candidate_indices = code_map.get(erp_code, [])
        best_idx, mismatch_fields = choose_best_onec_candidate(erp_doc, onec_docs, candidate_indices)

        if best_idx is None:
            report_rows.append(
                {
                    "status": STATUS_NOT_FOUND_IN_1C,
                    "note": "Документ ЕРП не найден в 1С по коду 1С",
                    "mismatch_fields": ["code1c"],
                    "oper_id": erp_doc.get("oper_id"),
                    "operation_url": erp_doc.get("operation_url"),
                    "operation_title": erp_doc.get("operation_title"),
                    "erp_doc_id": erp_doc.get("erp_doc_id"),
                    "erp_code1c": erp_doc.get("code1c"),
                    "erp_number": erp_doc.get("number"),
                    "erp_date_iso": erp_doc.get("date_iso"),
                    "erp_sum": erp_doc.get("sum"),
                    "erp_type": erp_doc.get("type_id"),
                    "onec_code1c": "",
                    "onec_number": "",
                    "onec_date_iso": "",
                    "onec_sum": None,
                    "onec_type": "",
                }
            )
            continue

        matched_onec = onec_docs[best_idx]
        matched_onec["_used"] = True

        if mismatch_fields:
            status = STATUS_FIELDS_MISMATCH
            note = "Поля документа различаются"
        else:
            status = STATUS_MATCH
            note = ""

        report_rows.append(
            {
                "status": status,
                "note": note,
                "mismatch_fields": mismatch_fields,
                "oper_id": erp_doc.get("oper_id"),
                "operation_url": erp_doc.get("operation_url"),
                "operation_title": erp_doc.get("operation_title"),
                "erp_doc_id": erp_doc.get("erp_doc_id"),
                "erp_code1c": erp_doc.get("code1c"),
                "erp_number": erp_doc.get("number"),
                "erp_date_iso": parse_any_date_to_iso(erp_doc.get("date_iso") or erp_doc.get("date")),
                "erp_sum": erp_doc.get("sum"),
                "erp_type": erp_doc.get("type_id"),
                "onec_code1c": matched_onec.get("code1c"),
                "onec_number": matched_onec.get("number"),
                "onec_date_iso": parse_any_date_to_iso(matched_onec.get("date_iso") or matched_onec.get("date")),
                "onec_sum": matched_onec.get("sum"),
                "onec_type": matched_onec.get("type"),
            }
        )

    for onec_doc in onec_docs:
        if onec_doc.get("_used"):
            continue

        onec_code = normalize_text(onec_doc.get("code1c"))
        status = STATUS_NOT_FOUND_IN_ERP if onec_code else STATUS_NOT_COMPARABLE
        note = "Документ 1С не найден в ЕРП" if onec_code else "У документа 1С нет кода 1С"

        report_rows.append(
            {
                "status": status,
                "note": note,
                "mismatch_fields": ["code1c"] if onec_code else ["code1c"],
                "oper_id": 0,
                "operation_url": "",
                "operation_title": "",
                "erp_doc_id": 0,
                "erp_code1c": "",
                "erp_number": "",
                "erp_date_iso": "",
                "erp_sum": None,
                "erp_type": "",
                "onec_code1c": onec_doc.get("code1c"),
                "onec_number": onec_doc.get("number"),
                "onec_date_iso": parse_any_date_to_iso(onec_doc.get("date_iso") or onec_doc.get("date")),
                "onec_sum": onec_doc.get("sum"),
                "onec_type": onec_doc.get("type"),
            }
        )

    summary = summarize_results(report_rows)
    run_id = None
    try:
        run_id = save_report_if_schema_exists(
            spec_id,
            len(onec_docs),
            report_rows,
            summary,
            scope=scope,
            scope_id=scope_id,
            client_id=client_id,
            source_mode=source_mode,
        )
    except Exception:
        run_id = None

    return {
        "summary": summary,
        "rows": report_rows,
        "run_id": run_id,
        "erp_docs_count": len(erp_docs),
        "onec_docs_count": len(onec_docs),
        "source_state": source_state,
    }


def build_delivery_query(spec_id: int) -> str:
    return f"""
SELECT
    s.f_id,
    COALESCE(s.f_num, '') AS spec_number,
    COALESCE(DATE_FORMAT(s.f_dt, '%d.%m.%Y'), '') AS spec_date,
    COALESCE(DATE_FORMAT(s.f_dt, '%Y-%m-%d'), '') AS spec_date_iso,
    COALESCE(s.f_status, 0) AS spec_status_id,
    COALESCE(s.f_tovar, '') AS goods_name,
    COALESCE(s.f_tovarprice, 0) AS goods_amount,
    COALESCE(s.f_tovarval, 0) AS goods_currency_id,
    COALESCE(NULLIF(goods_val.f_dopprstr, ''), NULLIF(goods_val.f_uslstr, ''), goods_val.f_name, '') AS goods_currency,
    COALESCE(s.f_postid, 0) AS post_id,
    COALESCE(s.f_kod1cb, '') AS spec_buyer_code1c,
    COALESCE(s.f_kod1cp, '') AS spec_principal_code1c,
    COALESCE(s.f_parentspecid, 0) AS parent_spec_id,
    COALESCE(d.f_id, 0) AS main_dog_id,
    COALESCE(d.f_kod1c, '') AS main_dog_code1c,
    COALESCE(d.f_dogname, '') AS main_dog_number,
    COALESCE(DATE_FORMAT(d.f_dogdate, '%d.%m.%Y'), '') AS main_dog_date,
    COALESCE(d.f_dogtype, 0) AS main_dog_type,
    COALESCE(d.f_orgid, 0) AS org_id,
    COALESCE(d.f_contrid, 0) AS client_id,
    COALESCE(org.f_cname, '') AS org_name,
    COALESCE(org.f_abbr, '') AS org_abbr,
    COALESCE(org.f_inn, '') AS org_inn,
    COALESCE(client.f_cname, '') AS client_name,
    COALESCE(client.f_abbr, '') AS client_abbr,
    COALESCE(client.f_inn, '') AS client_inn,
    COALESCE(CONCAT(user.f_name1, ' ', user.f_name2), '') AS manager_name
FROM veda_specs s
LEFT JOIN veda_dogs d
    ON d.f_id = s.f_dogid
LEFT JOIN veda_clients org
    ON org.f_id = d.f_orgid
LEFT JOIN veda_clients client
    ON client.f_id = d.f_contrid
LEFT JOIN veda_users user
    ON user.f_id = s.f_operid
LEFT JOIN veda_spr goods_val
    ON goods_val.f_type = 4 AND goods_val.f_num = s.f_tovarval
WHERE s.f_id = {spec_id}
LIMIT 1;
"""


def fetch_delivery(spec_id: int) -> dict[str, object]:
    row = first_row(build_delivery_query(spec_id))
    if not row:
        return {"spec_id": spec_id, "request_label": f"Заявка № {spec_id}"}

    spec_number = text_at(row, 1)
    dog_number = text_at(row, 15)
    dog_date = text_at(row, 16)
    request_label = f"Заявка № {spec_number}"
    if dog_number:
        request_label += f" по договору {dog_number}"
        if dog_date:
            request_label += f" от {dog_date}"

    return {
        "spec_id": int_at(row, 0),
        "spec_number": spec_number,
        "spec_date": text_at(row, 2),
        "spec_date_iso": text_at(row, 3),
        "spec_status_id": int_at(row, 4),
        "goods_name": text_at(row, 5),
        "goods_amount": float_at(row, 6, 0.0),
        "goods_currency_id": int_at(row, 7, 0),
        "goods_currency": text_at(row, 8),
        "post_id": int_at(row, 9, 0),
        "spec_buyer_code1c": text_at(row, 10),
        "spec_principal_code1c": text_at(row, 11),
        "parent_spec_id": int_at(row, 12, 0),
        "main_dog_id": int_at(row, 13, 0),
        "main_dog_code1c": text_at(row, 14),
        "main_dog_number": dog_number,
        "main_dog_date": dog_date,
        "main_dog_type": int_at(row, 17, 0),
        "org_id": int_at(row, 18, 0),
        "client_id": int_at(row, 19, 0),
        "org_name": text_at(row, 20),
        "org_abbr": text_at(row, 21),
        "org_inn": text_at(row, 22),
        "client_name": text_at(row, 23),
        "client_abbr": text_at(row, 24),
        "client_inn": text_at(row, 25),
        "manager_name": text_at(row, 26).strip(),
        "request_label": request_label,
    }


def build_contracts_query(spec_id: int) -> str:
    oper_ids = build_spec_operations_ids_subquery(spec_id)
    return f"""
SELECT
    d.f_id,
    COALESCE(d.f_kod1c, '') AS dog_code1c,
    COALESCE(d.f_dogname, '') AS dog_number,
    COALESCE(DATE_FORMAT(d.f_dogdate, '%d.%m.%Y'), '') AS dog_date,
    COALESCE(d.f_dogtype, 0) AS dog_type_id,
    COALESCE(dog_type.f_name, '') AS dog_type_name,
    COALESCE(d.f_orgid, 0) AS org_id,
    COALESCE(org.f_cname, '') AS org_name,
    COALESCE(org.f_inn, '') AS org_inn,
    COALESCE(d.f_contrid, 0) AS client_id,
    COALESCE(client.f_cname, '') AS client_name,
    COALESCE(client.f_inn, '') AS client_inn,
    GROUP_CONCAT(DISTINCT src.source_role ORDER BY src.source_role SEPARATOR ',') AS source_roles,
    COUNT(DISTINCT src.oper_id) AS operations_count,
    COALESCE(SUM(src.oper_sum), 0) AS operations_sum
FROM (
    SELECT 'spec' AS source_role, s.f_dogid AS dog_id, 0 AS oper_id, 0 AS oper_sum
    FROM veda_specs s
    WHERE s.f_id = {spec_id} AND s.f_dogid > 0

    UNION ALL

    SELECT 'operation' AS source_role, oper.f_dogid AS dog_id, oper.f_id AS oper_id, COALESCE(oper.f_sum, 0) AS oper_sum
    FROM veda_spec_invoices oper
    LEFT JOIN veda_categs oper4_specs
        ON oper4_specs.f_objectid = oper.f_id
       AND oper4_specs.f_ctgtype = 24
       AND oper4_specs.f_objecttype = 5
    WHERE {build_spec_operations_source(spec_id, "oper")}
      AND oper.f_dogid > 0

    UNION ALL

    SELECT 'invoice' AS source_role, schet.f_dogid AS dog_id, COALESCE(schet.f_operid, 0) AS oper_id, COALESCE(schet.f_sum, 0) AS oper_sum
    FROM veda_schets schet
    WHERE schet.f_dogtype <> 2
      AND schet.f_dogid > 0
      AND (
            schet.f_operid IN ({oper_ids})
            OR schet.f_id IN (SELECT si.f_invoiceid FROM veda_spec_invoices si WHERE si.f_id IN ({oper_ids}) AND si.f_invoiceid > 0)
          )

    UNION ALL

    SELECT 'act' AS source_role, akt.f_dogid AS dog_id, COALESCE(akt.f_operid, 0) AS oper_id, COALESCE(akt.f_sum, 0) AS oper_sum
    FROM veda_akts akt
    WHERE akt.f_dogid > 0
      AND akt.f_operid IN ({oper_ids})
) src
JOIN veda_dogs d
    ON d.f_id = src.dog_id
LEFT JOIN veda_spr dog_type
    ON dog_type.f_type = 59 AND dog_type.f_num = d.f_dogtype
LEFT JOIN veda_clients org
    ON org.f_id = d.f_orgid
LEFT JOIN veda_clients client
    ON client.f_id = d.f_contrid
GROUP BY d.f_id
ORDER BY
    MIN(CASE src.source_role
        WHEN 'spec' THEN 0
        WHEN 'operation' THEN 1
        WHEN 'invoice' THEN 2
        WHEN 'act' THEN 3
        ELSE 9
    END),
    d.f_id;
"""


def infer_contract_role(contract: dict[str, object], delivery: dict[str, object]) -> str:
    roles = set(normalize_text(contract.get("source_roles")).split(","))
    if "spec" in roles:
        return "Договор покупателя / основной договор поставки"
    if sql_int(contract.get("org_id")) == sql_int(delivery.get("client_id")) and sql_int(contract.get("client_id")) == sql_int(delivery.get("org_id")):
        return "Встречный договор / договор с коммитентом"
    if "invoice" in roles:
        return "Договор расчетов по счетам"
    if "act" in roles:
        return "Договор закрывающих документов"
    return "Связанный договор поставки"


def meaningful_code1c(value: object) -> str:
    text = normalize_text(value)
    if text in {"", "_", "-", "0"}:
        return ""
    return text


def fetch_contracts(spec_id: int, delivery: dict[str, object]) -> list[dict[str, object]]:
    contracts: list[dict[str, object]] = []
    main_dog_number = normalize_text(delivery.get("main_dog_number"))
    spec_number = normalize_text(delivery.get("spec_number"))
    onec_contract_number = f"{main_dog_number}/{spec_number}" if main_dog_number and spec_number else main_dog_number
    onec_contract_title = normalize_text(delivery.get("request_label")) or onec_contract_number

    base = {
        "dog_id": sql_int(delivery.get("main_dog_id")),
        "dog_number": onec_contract_number,
        "base_dog_number": main_dog_number,
        "base_dog_code1c": normalize_text(delivery.get("main_dog_code1c")),
        "spec_number": spec_number,
        "dog_date": normalize_text(delivery.get("main_dog_date")),
        "dog_type_id": sql_int(delivery.get("main_dog_type")),
        "org_id": sql_int(delivery.get("org_id")),
        "org_name": normalize_text(delivery.get("org_name")),
        "org_inn": normalize_text(delivery.get("org_inn")),
        "client_id": sql_int(delivery.get("client_id")),
        "client_name": normalize_text(delivery.get("client_name")),
        "client_inn": normalize_text(delivery.get("client_inn")),
        "onec_contract_title": onec_contract_title,
        "operations_count": 0,
        "operations_sum": 0.0,
    }

    buyer_code = meaningful_code1c(delivery.get("spec_buyer_code1c"))
    if buyer_code:
        contracts.append(
            {
                **base,
                "dog_code1c": buyer_code,
                "dog_type_name": "С покупателем",
                "source_roles": "spec_buyer",
                "role_label": "Договор 1С с покупателем",
            }
        )

    principal_code = meaningful_code1c(delivery.get("spec_principal_code1c"))
    if principal_code:
        contracts.append(
            {
                **base,
                "dog_code1c": principal_code,
                "dog_type_name": "С комитентом",
                "source_roles": "spec_principal",
                "role_label": "Договор 1С с комитентом",
            }
        )

    if not contracts and (main_dog_number or meaningful_code1c(delivery.get("main_dog_code1c"))):
        contracts.append(
            {
                **base,
                "dog_code1c": meaningful_code1c(delivery.get("main_dog_code1c")),
                "dog_type_name": "Основной договор",
                "source_roles": "spec",
                "role_label": "Основной договор ERP",
            }
        )
    return contracts


def fetch_client_tree(delivery: dict[str, object]) -> dict[str, object]:
    current_client_id = sql_int(delivery.get("client_id"))
    current_dog_id = sql_int(delivery.get("main_dog_id"))
    if current_client_id <= 0:
        return {}

    contact_row = first_row(
        f"""
SELECT
    COALESCE(cl.f_contactid, 0) AS contact_id,
    COALESCE(NULLIF(NULLIF(contact.f_cname, ''), '_'), NULLIF(NULLIF(contact.f_name, ''), '_'), cl.f_cname, '') AS contact_name,
    COALESCE(contact.f_inn, '') AS contact_inn,
    COALESCE(cl.f_id, 0) AS legal_entity_id,
    COALESCE(cl.f_cname, '') AS legal_entity_name,
    COALESCE(cl.f_abbr, '') AS legal_entity_abbr,
    COALESCE(cl.f_inn, '') AS legal_entity_inn
FROM veda_clients cl
LEFT JOIN veda_contacts contact
    ON contact.f_id = cl.f_contactid
WHERE cl.f_id = {current_client_id}
LIMIT 1;
"""
    )
    if not contact_row:
        return {}

    contact_id = int_at(contact_row, 0)
    contact = {
        "id": contact_id,
        "name": meaningful_code1c(text_at(contact_row, 1)) or normalize_text(delivery.get("client_name")),
        "inn": text_at(contact_row, 2),
    }

    if contact_id > 0:
        clients_where = f"cl.f_contactid = {contact_id}"
    else:
        clients_where = f"cl.f_id = {current_client_id}"

    rows = run_mysql_tsv(
        f"""
SELECT
    COALESCE(cl.f_id, 0) AS legal_entity_id,
    COALESCE(cl.f_cname, '') AS legal_entity_name,
    COALESCE(cl.f_abbr, '') AS legal_entity_abbr,
    COALESCE(cl.f_inn, '') AS legal_entity_inn,
    COALESCE(cl.f_kod1c, '') AS legal_entity_code1c,
    COALESCE(cl.f_status, 0) AS legal_entity_status,
    COALESCE(d.f_id, 0) AS dog_id,
    COALESCE(d.f_kod1c, '') AS dog_code1c,
    COALESCE(d.f_kod1cg, '') AS dog_group_code1c,
    COALESCE(d.f_kod1cgo, '') AS dog_open_group_code1c,
    COALESCE(d.f_kod1cgc, '') AS dog_closed_group_code1c,
    COALESCE(d.f_dogname, '') AS dog_number,
    COALESCE(DATE_FORMAT(d.f_dogdate, '%d.%m.%Y'), '') AS dog_date,
    COALESCE(d.f_dogtype, 0) AS dog_type_id,
    COALESCE(dog_type.f_name, '') AS dog_type_name,
    COALESCE(d.f_orgid, 0) AS org_id,
    COALESCE(org.f_abbr, '') AS org_abbr,
    COALESCE(org.f_cname, '') AS org_name,
    COALESCE(org.f_inn, '') AS org_inn,
    COALESCE(d.f_status, 0) AS dog_status,
    COALESCE(d.f_parentdogid, 0) AS parent_dog_id
FROM veda_clients cl
LEFT JOIN veda_dogs d
    ON d.f_contrid = cl.f_id
LEFT JOIN veda_spr dog_type
    ON dog_type.f_type = 59 AND dog_type.f_num = d.f_dogtype
LEFT JOIN veda_clients org
    ON org.f_id = d.f_orgid
WHERE {clients_where}
ORDER BY
    cl.f_cname,
    d.f_dogdate DESC,
    d.f_id DESC
LIMIT 1000;
"""
    )

    entities: dict[int, dict[str, object]] = {}
    for row in rows:
        legal_id = int_at(row, 0)
        if legal_id <= 0:
            continue
        entity = entities.setdefault(
            legal_id,
            {
                "id": legal_id,
                "name": text_at(row, 1),
                "abbr": text_at(row, 2),
                "inn": text_at(row, 3),
                "code1c": text_at(row, 4),
                "status_id": int_at(row, 5),
                "is_current": legal_id == current_client_id,
                "contracts": [],
            },
        )
        dog_id = int_at(row, 6)
        if dog_id <= 0:
            continue
        entity["contracts"].append(
            {
                "id": dog_id,
                "code1c": text_at(row, 7),
                "group_code1c": text_at(row, 8),
                "open_group_code1c": text_at(row, 9),
                "closed_group_code1c": text_at(row, 10),
                "number": text_at(row, 11),
                "date": text_at(row, 12),
                "type_id": int_at(row, 13),
                "type_name": text_at(row, 14),
                "org_id": int_at(row, 15),
                "org_abbr": text_at(row, 16),
                "org_name": text_at(row, 17),
                "org_inn": text_at(row, 18),
                "status_id": int_at(row, 19),
                "parent_dog_id": int_at(row, 20),
                "is_current": dog_id == current_dog_id,
            }
        )

    if not entities:
        legal_id = int_at(contact_row, 3, current_client_id)
        entities[legal_id] = {
            "id": legal_id,
            "name": text_at(contact_row, 4) or normalize_text(delivery.get("client_name")),
            "abbr": text_at(contact_row, 5),
            "inn": text_at(contact_row, 6) or normalize_text(delivery.get("client_inn")),
            "code1c": "",
            "status_id": 0,
            "is_current": True,
            "contracts": [],
        }

    return {
        "contact": contact,
        "legal_entities": list(entities.values()),
        "current": {
            "legal_entity_id": current_client_id,
            "dog_id": current_dog_id,
        },
    }


def build_payments_query(spec_id: int) -> str:
    return f"""
SELECT
    h.f_id AS payment_id,
    COALESCE(h.f_kod1C, '') AS code1c,
    d.f_docid AS oper_id,
    COALESCE(oper.f_num_oper, 999999) AS oper_num,
    COALESCE(oper_type.f_name, '') AS oper_type_name,
    COALESCE(h.f_type, 0) AS payment_type,
    COALESCE(h.f_ahtype, 0) AS acchist_type,
    COALESCE(h.f_ppnum, '') AS pp_number,
    COALESCE(
        CASE
            WHEN h.f_dt1C <> '0000-00-00 00:00:00' THEN DATE_FORMAT(h.f_dt1C, '%Y-%m-%d')
            ELSE DATE_FORMAT(h.f_ppdt, '%Y-%m-%d')
        END,
        ''
    ) AS payment_date_iso,
    COALESCE(
        CASE
            WHEN h.f_dt1C <> '0000-00-00 00:00:00' THEN DATE_FORMAT(h.f_dt1C, '%d.%m.%Y')
            ELSE DATE_FORMAT(h.f_ppdt, '%d.%m.%Y')
        END,
        ''
    ) AS payment_date,
    COALESCE(h.f_sum, 0) AS payment_sum,
    COALESCE(d.f_clssum, 0) AS classified_sum,
    COALESCE(h.f_val, 0) AS currency_id,
    COALESCE(NULLIF(val.f_dopprstr, ''), NULLIF(val.f_uslstr, ''), val.f_name, '') AS currency,
    COALESCE(h.f_name, '') AS payer_name,
    COALESCE(h.f_contrid, 0) AS contr_id,
    COALESCE(oper.f_invoiceid, 0) AS invoice_id,
    COALESCE(invoice.f_num, '') AS invoice_number,
    COALESCE(oper.f_isvozm, 0) AS reimbursement_id,
    COALESCE(vozm.f_name, '') AS reimbursement_name,
    CASE
        WHEN h.f_type = 0 THEN 'incoming'
        WHEN h.f_type = 1 THEN 'outgoing'
        ELSE 'other'
    END AS direction
FROM veda_acchist_docs d
JOIN veda_acchist h
    ON h.f_id = d.f_acchistid
JOIN veda_spec_invoices oper
    ON oper.f_id = d.f_docid
LEFT JOIN veda_typeopers oper_type
    ON oper_type.f_id = oper.f_idoper
LEFT JOIN veda_categs oper4_specs
    ON oper4_specs.f_objectid = oper.f_id
   AND oper4_specs.f_ctgtype = 24
   AND oper4_specs.f_objecttype = 5
LEFT JOIN veda_spr val
    ON val.f_type = 4 AND val.f_num = h.f_val
LEFT JOIN veda_schets invoice
    ON invoice.f_id = oper.f_invoiceid
LEFT JOIN veda_spr vozm
    ON vozm.f_type = 2 AND vozm.f_num = oper.f_isvozm
WHERE d.f_doctype = 3
  AND {build_spec_operations_source(spec_id, "oper")}
ORDER BY payment_date_iso, h.f_id;
"""


def fetch_payments(spec_id: int) -> list[dict[str, object]]:
    payments: list[dict[str, object]] = []
    for row in run_mysql_tsv(build_payments_query(spec_id)):
        payments.append(
            {
                "payment_id": int_at(row, 0),
                "code1c": text_at(row, 1),
                "oper_id": int_at(row, 2),
                "oper_num": int_at(row, 3, 999999),
                "oper_type_name": text_at(row, 4),
                "payment_type": int_at(row, 5),
                "acchist_type": int_at(row, 6),
                "pp_number": text_at(row, 7),
                "payment_date_iso": text_at(row, 8),
                "payment_date": text_at(row, 9),
                "payment_sum": float_at(row, 10, 0.0),
                "classified_sum": float_at(row, 11, 0.0),
                "currency_id": int_at(row, 12, 0),
                "currency": text_at(row, 13),
                "payer_name": text_at(row, 14),
                "contr_id": int_at(row, 15),
                "invoice_id": int_at(row, 16),
                "invoice_number": text_at(row, 17),
                "reimbursement_id": int_at(row, 18, 0),
                "reimbursement_name": text_at(row, 19),
                "direction": text_at(row, 20),
            }
        )
    return payments


def build_settlements(
    operations: list[dict[str, object]],
    erp_docs: list[dict[str, object]],
    payments: list[dict[str, object]],
) -> dict[str, object]:
    schets = [doc for doc in erp_docs if doc.get("doc_kind") == "schet"]
    acts = [doc for doc in erp_docs if doc.get("doc_kind") == "act"]
    buyer_schets = [doc for doc in schets if sql_int(doc.get("type_id")) == 2]
    if not buyer_schets:
        buyer_schets = schets

    paid_incoming = [p for p in payments if p.get("direction") == "incoming"]
    paid_outgoing = [p for p in payments if p.get("direction") == "outgoing"]
    reimbursable_ops = [op for op in operations if sql_int(op.get("reimbursement_id")) == 1]
    non_reimbursable_ops = [op for op in operations if sql_int(op.get("reimbursement_id")) != 1]
    add_nds_ops = [op for op in operations if sql_int(op.get("add_nds_flag")) == 1]
    control_excluded_cost_ops = [
        op for op in non_reimbursable_ops
        if (normalize_sum(op.get("rp_expenses_sum")) or 0.0) > 0 and (normalize_sum(op.get("rp_profit_sum")) or 0.0) < 0
    ]
    control_reimbursable_ops = [
        op for op in reimbursable_ops
        if sql_int(op.get("add_nds_flag")) != 1 and normalize_sum(op.get("rp_expenses_sum")) not in (None, 0.0)
    ]
    control_non_reimbursable_ops = [
        op for op in non_reimbursable_ops
        if op not in control_excluded_cost_ops
    ]

    buyer_invoice_sum = sum_values(buyer_schets, "sum")
    classified_paid_sum = sum_values(paid_incoming, "classified_sum")
    rp_paid_sum = sum_values(operations, "rp_paid_sum")
    rp_expenses_sum = sum_values(operations, "rp_expenses_sum")
    rp_realiz_sum = sum_values(operations, "rp_realiz_sum")
    rp_profit_sum = sum_values(operations, "rp_profit_sum")
    reimbursable_sum = sum_values(reimbursable_ops, "oper_sum")
    non_reimbursable_sum = sum_values(non_reimbursable_ops, "oper_sum")
    control_reimbursable_sum = sum_values(control_reimbursable_ops, "rp_expenses_sum")
    control_non_reimbursable_sum = sum_values(control_non_reimbursable_ops, "oper_sum")
    add_nds_oper_sum = sum_values(add_nds_ops, "oper_sum")
    add_nds_expenses_sum = sum_values(add_nds_ops, "rp_expenses_sum")
    add_nds_profit_sum = sum_values(add_nds_ops, "rp_profit_sum")
    control_excluded_cost_oper_sum = sum_values(control_excluded_cost_ops, "oper_sum")
    control_excluded_cost_expenses_sum = sum_values(control_excluded_cost_ops, "rp_expenses_sum")
    control_excluded_cost_profit_sum = sum_values(control_excluded_cost_ops, "rp_profit_sum")
    expenses_paid_sum = sum_values(paid_outgoing, "classified_sum")
    profit_by_period: dict[str, float] = {}
    realiz_by_period: dict[str, float] = {}
    for operation in operations:
        period = str(sql_int(operation.get("out_buh_period")))
        profit_by_period[period] = round(profit_by_period.get(period, 0.0) + float(operation.get("rp_profit_sum") or 0.0), 2)
        realiz_by_period[period] = round(realiz_by_period.get(period, 0.0) + float(operation.get("rp_realiz_sum") or 0.0), 2)

    return {
        "invoice_total": sum_values(schets, "sum"),
        "buyer_invoice_total": buyer_invoice_sum,
        "paid_total": rp_paid_sum,
        "classified_paid_total": classified_paid_sum,
        "expenses_paid_total": expenses_paid_sum,
        "expenses_accrued_total": rp_expenses_sum,
        "realization_total": rp_realiz_sum,
        "profit_total": rp_profit_sum,
        "realization_minus_expenses_total": round(rp_realiz_sum - rp_expenses_sum, 2),
        "profit_by_out_buh_period": profit_by_period,
        "realization_by_out_buh_period": realiz_by_period,
        "reimbursable_total": reimbursable_sum,
        "non_reimbursable_total": non_reimbursable_sum,
        "control_reimbursable_total": control_reimbursable_sum,
        "control_non_reimbursable_total": control_non_reimbursable_sum,
        "control_sheet_delta": round(rp_paid_sum - control_reimbursable_sum - control_non_reimbursable_sum, 2),
        "add_nds_operations_count": len(add_nds_ops),
        "add_nds_oper_total": add_nds_oper_sum,
        "add_nds_expenses_total": add_nds_expenses_sum,
        "add_nds_profit_total": add_nds_profit_sum,
        "control_excluded_cost_operations_count": len(control_excluded_cost_ops),
        "control_excluded_cost_oper_total": control_excluded_cost_oper_sum,
        "control_excluded_cost_expenses_total": control_excluded_cost_expenses_sum,
        "control_excluded_cost_profit_total": control_excluded_cost_profit_sum,
        "acts_total": sum_values(acts, "sum"),
        "balance_to_buyer": round(buyer_invoice_sum - rp_paid_sum, 2),
        "customer_settlement_balance_rub": round(rp_paid_sum - rp_realiz_sum, 2),
        "settlement_delta": round(buyer_invoice_sum - rp_paid_sum + reimbursable_sum + non_reimbursable_sum, 2),
        "settlement_source": "mariadb_routines:get_paidsum/get_expensessum/get_realizsum/get_profit",
        "operations_count": len(operations),
        "schets_count": len(schets),
        "acts_count": len(acts),
        "payments_count": len(payments),
    }


def delivery_balance_signature(values: dict[str, object]) -> str:
    keys = [
        "client_invoice_total",
        "client_invoice_count",
        "client_payment_total",
        "client_payment_count",
        "closing_document_total",
        "closing_document_count",
        "money_balance",
        "documents_balance",
        "client_invoice_mixed_currency",
    ]
    return "|".join(str(values.get(key, "")) for key in keys)


def build_delivery_balance(
    operations: list[dict[str, object]],
    erp_docs: list[dict[str, object]],
    payments: list[dict[str, object]],
    settlements: dict[str, object],
) -> dict[str, object]:
    schets = [doc for doc in erp_docs if doc.get("doc_kind") == "schet"]
    acts = [doc for doc in erp_docs if doc.get("doc_kind") == "act"]

    client_invoice_docs = [
        doc for doc in schets
        if normalize_text(doc.get("code1c"))
        and not normalize_text(doc.get("invoice_number"))
        and (
            compact_key(doc.get("detail_name")) == compact_key("Агрегирующий счет")
            or sql_int(doc.get("status_id")) > 0
        )
    ]
    if not client_invoice_docs:
        client_invoice_docs = [
            doc for doc in schets
            if normalize_text(doc.get("code1c")) and not normalize_text(doc.get("invoice_number"))
        ]

    incoming_payment_ids = {
        sql_int(payment.get("payment_id"))
        for payment in payments
        if normalize_text(payment.get("direction")) == "incoming" and sql_int(payment.get("payment_id")) > 0
    }

    realiz_by_oper = {
        sql_int(operation.get("oper_id")): normalize_sum(operation.get("rp_realiz_sum")) or 0.0
        for operation in operations
    }
    closing_act_rows = [
        row for row in acts
        if (realiz_by_oper.get(sql_int(row.get("oper_id")), 0.0) > 0)
        or normalize_text(row.get("main_code1c"))
        or normalize_text(row.get("code1c"))
    ]
    closing_items = [
        item for item in aggregate_act_compare_items(closing_act_rows)
        if normalize_text(item.get("code1c"))
    ]
    closing_keys = {
        (
            compact_key(item.get("code1c")),
            parse_any_date_to_iso(item.get("date_iso") or item.get("date")),
            round(normalize_sum(item.get("sum")) or 0.0, 2),
        )
        for item in closing_items
    }

    client_invoice_total = sum_values(client_invoice_docs, "sum")
    invoice_by_currency: dict[str, dict[str, object]] = {}
    for doc in client_invoice_docs:
        currency = normalize_text(doc.get("currency")) or "RUB"
        bucket = invoice_by_currency.setdefault(currency, {"currency": currency, "total": 0.0, "count": 0})
        bucket["total"] = round(float(bucket["total"] or 0.0) + float(normalize_sum(doc.get("sum")) or 0.0), 2)
        bucket["count"] = int(bucket["count"] or 0) + 1
    invoice_currency_breakdown = sorted(invoice_by_currency.values(), key=lambda item: normalize_text(item.get("currency")))
    mixed_invoice_currency = len(invoice_currency_breakdown) > 1

    client_payment_total = normalize_sum(settlements.get("paid_total")) or 0.0
    closing_document_total = normalize_sum(settlements.get("realization_total")) or 0.0
    invoice_payment_balance = None if mixed_invoice_currency else round(client_invoice_total - client_payment_total, 2)
    documents_balance = None if mixed_invoice_currency else round(client_invoice_total - closing_document_total, 2)
    money_balance = round(client_payment_total - closing_document_total, 2)
    money_closed = abs(money_balance) <= 0.01
    documents_closed = True if documents_balance is None else abs(documents_balance) <= 0.01

    if money_closed and documents_closed:
        status = "closed"
        status_label = "По суммам закрыто"
    elif money_closed:
        status = "money_closed"
        status_label = "По суммам закрыто"
    else:
        status = "open_balance"
        status_label = "Есть переплата" if money_balance > 0.01 else "Есть долг"

    warnings = []
    if not documents_closed:
        warnings.append("Есть незакрытые документы")
    if mixed_invoice_currency:
        warnings.append("Счета клиенту в разных валютах")

    result = {
        "client_invoice_total": client_invoice_total,
        "client_invoice_count": len(client_invoice_docs),
        "client_invoice_currency_breakdown": invoice_currency_breakdown,
        "client_invoice_mixed_currency": mixed_invoice_currency,
        "client_payment_total": round(client_payment_total, 2),
        "client_payment_count": len(incoming_payment_ids),
        "closing_document_total": round(closing_document_total, 2),
        "closing_document_count": len(closing_keys),
        "money_balance": money_balance,
        "documents_balance": documents_balance,
        "invoice_payment_balance": invoice_payment_balance,
        "money_closed": money_closed,
        "documents_closed": documents_closed,
        "balance_formula": "client_payments_minus_closing_documents" if mixed_invoice_currency else "client_payments_minus_closing_documents",
        "balance_formula_label": "оплаты - закрывающие документы",
        "status": status,
        "status_label": status_label,
        "warnings": warnings,
        "next_action": "Запустить сверку с 1С",
    }
    result["signature"] = delivery_balance_signature(result)
    return result


def build_erp_snapshot(spec_id: int) -> dict[str, object]:
    delivery = fetch_delivery(spec_id)
    operations = fetch_operations(spec_id)
    erp_docs = fetch_erp_docs(spec_id)
    payments = fetch_payments(spec_id)
    contracts = fetch_contracts(spec_id, delivery)
    client_tree = fetch_client_tree(delivery)
    contact = client_tree.get("contact") if isinstance(client_tree, dict) else {}
    if isinstance(contact, dict) and contact:
        delivery["contact_id"] = sql_int(contact.get("id"))
        delivery["contact_name"] = normalize_text(contact.get("name"))
        delivery["contact_inn"] = normalize_text(contact.get("inn"))
    settlements = build_settlements(operations, erp_docs, payments)
    delivery_balance = build_delivery_balance(operations, erp_docs, payments, settlements)

    return {
        "ok": True,
        "spec_id": spec_id,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "delivery": delivery,
        "contracts": contracts,
        "client_tree": client_tree,
        "operations": operations,
        "erp_docs": erp_docs,
        "schets": [doc for doc in erp_docs if doc.get("doc_kind") == "schet"],
        "akts": [doc for doc in erp_docs if doc.get("doc_kind") == "act"],
        "payments": payments,
        "settlements": settlements,
        "delivery_balance": delivery_balance,
        "xlsx_url": f"/api/reconciliation/erp-export.xlsx?spec_id={spec_id}",
    }


def live_money(value: object) -> float:
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def one_line(value: object) -> str:
    return " ".join(str(value or "").replace("\t", " ").replace("\r", " ").replace("\n", " ").split())


def meaningful_text(value: object) -> str:
    text = one_line(value)
    if text in {"", "_", "-", "0"}:
        return ""
    return text


def joined_unique_lines(values: list[str]) -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = one_line(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return "\n".join(result) if result else "—"


def balance_badge_label(delta: float) -> str:
    if delta > 0.01:
        return "Переплата"
    if delta < -0.01:
        return "Долг"
    return ""


def matrix_badges(issues: list[str], delta: float) -> list[dict[str, str]]:
    badges: list[dict[str, str]] = []
    if "NO_CUSTOMER_INVOICE" in issues:
        badges.append({"key": "noerp", "label": "Нет счета покупателю"})
    if "NO_CLOSING_DOC" in issues:
        badges.append({"key": "no1c", "label": "Нет закрывающего документа"})
    if "PAYMENT_ROUTINE_VS_DOCS_MISMATCH" in issues:
        badges.append({"key": "fields", "label": "Оплаты расходятся"})
    if "UNCLASSIFIED_REALIZATION" in issues:
        badges.append({"key": "fields", "label": "Не определена возмещаемость"})
    balance_label = balance_badge_label(delta)
    if balance_label:
        badges.append({"key": "sum", "label": balance_label})
    if not badges:
        badges.append({"key": "ok", "label": "ОК"})
    return badges


def aggregate_matrix_rows(rows: list[dict[str, object]]) -> dict[str, object]:
    totals = {
        "invoiceSum": live_money(sum(live_money(row.get("invoiceSum")) for row in rows)),
        "paymentSum": live_money(sum(live_money(row.get("paymentSum")) for row in rows)),
        "reimbursableSum": live_money(sum(live_money(row.get("reimbursableSum")) for row in rows)),
        "nonReimbursableSum": live_money(sum(live_money(row.get("nonReimbursableSum")) for row in rows)),
        "delta": live_money(sum(live_money(row.get("delta")) for row in rows)),
    }
    issues: list[dict[str, str]] = []
    for row in rows:
        for badge in row.get("badges", []) if isinstance(row.get("badges"), list) else []:
            if isinstance(badge, dict) and badge.get("key") != "ok":
                label = one_line(badge.get("label")) or "Отклонение"
                label = re.sub(r"(?::\s*\d+)+$", "", label).strip()
                issues.append({"key": one_line(badge.get("key")) or "sum", "label": label or "Отклонение"})
    if not issues:
        totals["badges"] = [{"key": "ok", "label": "ОК"}]
    else:
        grouped: dict[tuple[str, str], int] = {}
        for issue in issues:
            grouped[(issue["key"], issue["label"])] = grouped.get((issue["key"], issue["label"]), 0) + 1
        if len(grouped) == 1:
            (key, label), count = next(iter(grouped.items()))
            totals["badges"] = [{"key": key, "label": f"{label}: {count}"}]
        else:
            totals["badges"] = [{"key": "sum", "label": f"Отклонений: {len(issues)}"}]
    totals["showAmounts"] = True
    return totals


def operation_bucket(operation: dict[str, object]) -> str:
    marker = sql_int(operation.get("reimbursement_id"))
    if marker == 1:
        return "reimbursable"
    if marker == 2:
        return "non_reimbursable"
    return "unclassified"


def customer_invoice_docs(erp_docs: list[dict[str, object]]) -> list[dict[str, object]]:
    invoices: list[dict[str, object]] = []
    seen: set[int] = set()
    for doc in erp_docs:
        if doc.get("doc_kind") != "schet":
            continue
        if sql_int(doc.get("invoice_id")) != 0:
            continue
        if one_line(doc.get("type_name")) != "Счет покупателю":
            continue
        doc_id = sql_int(doc.get("erp_doc_id"))
        if doc_id <= 0 or doc_id in seen:
            continue
        seen.add(doc_id)
        invoices.append(doc)
    invoices.sort(key=lambda item: (str(item.get("date_iso") or ""), str(item.get("number") or ""), sql_int(item.get("erp_doc_id"))))
    return invoices


def closing_docs(erp_docs: list[dict[str, object]]) -> list[dict[str, object]]:
    docs: list[dict[str, object]] = []
    seen: set[tuple[str, str, str, float]] = set()
    for doc in erp_docs:
        if doc.get("doc_kind") != "act":
            continue
        key = (
            one_line(doc.get("code1c")),
            one_line(doc.get("number")),
            one_line(doc.get("date")),
            live_money(doc.get("sum")),
        )
        if key in seen:
            continue
        seen.add(key)
        docs.append(doc)
    docs.sort(key=lambda item: (str(item.get("date_iso") or ""), str(item.get("code1c") or ""), str(item.get("number") or "")))
    return docs


def document_line(doc: dict[str, object]) -> str:
    number = one_line(doc.get("code1c")) or one_line(doc.get("number"))
    date = one_line(doc.get("date"))
    if number and date:
        return f"{number} от {date}"
    return number


def detail_doc(
    doc_id: str,
    doc_type: str,
    erp: str,
    onec: str,
    date: str,
    amount: float,
    status_key: str = "pending",
    status_label: str = "ERP",
) -> dict[str, object]:
    return {
        "id": doc_id,
        "type": doc_type,
        "erp": erp or "—",
        "onec": onec or "—",
        "date": date or "—",
        "amount": live_money(amount),
        "status": {"key": status_key, "label": status_label},
    }


def compact_unique(values: list[str], empty: str = "—") -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = one_line(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return ", ".join(result) if result else empty


def build_spec_detail_docs(
    spec_id: int,
    operations: list[dict[str, object]],
    invoices: list[dict[str, object]],
    acts: list[dict[str, object]],
    payments: list[dict[str, object]],
    reimbursable: float,
    non_reimbursable: float,
) -> list[dict[str, object]]:
    details: list[dict[str, object]] = []

    for doc in invoices:
        details.append(
            detail_doc(
                f"spec-{spec_id}-invoice-{sql_int(doc.get('erp_doc_id'))}",
                "Счет на оплату",
                one_line(doc.get("number")) or one_line(doc.get("code1c")),
                one_line(doc.get("code1c")),
                one_line(doc.get("date")),
                live_money(doc.get("sum")),
                "pending",
                "ERP-данные",
            )
        )

    for payment in payments:
        if payment.get("direction") != "incoming":
            continue
        payment_id = sql_int(payment.get("payment_id"))
        details.append(
            detail_doc(
                f"spec-{spec_id}-payment-{payment_id}-{sql_int(payment.get('oper_id'))}",
                "Оплата покупателя",
                one_line(payment.get("pp_number")) or f"ПП {payment_id}",
                one_line(payment.get("code1c")),
                one_line(payment.get("payment_date")),
                live_money(payment.get("classified_sum") or payment.get("payment_sum")),
                "pending",
                "ERP-данные",
            )
        )

    for doc in acts:
        details.append(
            detail_doc(
                f"spec-{spec_id}-act-{sql_int(doc.get('erp_doc_id'))}",
                one_line(doc.get("type_name")) or "Закрывающий документ",
                one_line(doc.get("number")) or one_line(doc.get("code1c")),
                one_line(doc.get("code1c")),
                one_line(doc.get("date")),
                live_money(doc.get("sum")),
                "pending",
                "ERP-данные",
            )
        )

    return details


def build_spec_snapshot_from_parts(
    spec_id: int,
    operations: list[dict[str, object]],
    erp_docs: list[dict[str, object]],
    payments: list[dict[str, object]],
) -> dict[str, object]:
    delivery = fetch_delivery(spec_id)
    contracts = fetch_contracts(spec_id, delivery)
    settlements = build_settlements(operations, erp_docs, payments)
    delivery_balance = build_delivery_balance(operations, erp_docs, payments, settlements)
    return {
        "ok": True,
        "spec_id": spec_id,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "delivery": delivery,
        "contracts": contracts,
        "client_tree": {},
        "operations": operations,
        "erp_docs": erp_docs,
        "schets": [doc for doc in erp_docs if doc.get("doc_kind") == "schet"],
        "akts": [doc for doc in erp_docs if doc.get("doc_kind") == "act"],
        "payments": payments,
        "settlements": settlements,
        "delivery_balance": delivery_balance,
        "xlsx_url": f"/api/reconciliation/erp-export.xlsx?spec_id={spec_id}",
    }


def compare_status_badge(row: dict[str, object]) -> dict[str, str]:
    status = normalize_text(row.get("status"))
    mismatch_fields = row.get("mismatch_fields") if isinstance(row.get("mismatch_fields"), list) else []
    if status == STATUS_MATCH:
        return {"key": "ok", "label": "Совпало"}
    if status == STATUS_NOT_FOUND_IN_1C:
        return {"key": "no1c", "label": "Нет в 1С"}
    if status == STATUS_NOT_FOUND_IN_ERP:
        return {"key": "noerp", "label": "Нет в ERP"}
    if status == STATUS_FIELDS_MISMATCH:
        if "sum" in mismatch_fields or "amount" in mismatch_fields:
            return {"key": "sum", "label": "Сумма расходится"}
        if "vat" in mismatch_fields or "nds" in mismatch_fields or "vat_rate" in mismatch_fields:
            return {"key": "nosf", "label": "Вопрос по НДС"}
        if "contract" in mismatch_fields or "dog" in mismatch_fields or "dog_code1c" in mismatch_fields:
            return {"key": "fields", "label": "Договор расходится"}
        if "date" in mismatch_fields or "doc_date" in mismatch_fields:
            return {"key": "fields", "label": "Дата расходится"}
        if "invoice_number" in mismatch_fields or "number" in mismatch_fields or "code1c" in mismatch_fields:
            return {"key": "fields", "label": "Номер расходится"}
        return {"key": "fields", "label": "Поля документа расходятся"}
    if status == STATUS_NOT_COMPARABLE:
        return {"key": "nokey", "label": "Нет ключа 1С"}
    return {"key": "pending", "label": "Не сверено"}


def compare_matrix_badges(report: dict[str, object] | None, delta: float, fallback_issues: list[str]) -> list[dict[str, str]]:
    if not report:
        return matrix_badges(fallback_issues, delta)
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    badges: list[dict[str, str]] = []
    not_found_1c = sql_int(summary.get(STATUS_NOT_FOUND_IN_1C))
    not_found_erp = sql_int(summary.get(STATUS_NOT_FOUND_IN_ERP))
    mismatch_sum = sql_int(summary.get("FIELDS_MISMATCH_SUM"))
    mismatch_total = sql_int(summary.get(STATUS_FIELDS_MISMATCH))
    not_comparable = sql_int(summary.get(STATUS_NOT_COMPARABLE))
    total = sql_int(summary.get("total"))
    match_count = sql_int(summary.get(STATUS_MATCH))
    has_compare_problem = bool(not_found_1c or not_found_erp or mismatch_total or not_comparable)
    if total and not has_compare_problem:
        badges.append({"key": "ok", "label": f"1С: {match_count}/{total} совпало"})
    if not_found_1c:
        badges.append({"key": "no1c", "label": f"Нет в 1С {not_found_1c}"})
    if not_found_erp:
        badges.append({"key": "noerp", "label": f"Нет в ERP {not_found_erp}"})
    if mismatch_sum:
        badges.append({"key": "sum", "label": f"Сумма расходится {mismatch_sum}"})
    other_mismatch = max(mismatch_total - mismatch_sum, 0)
    if other_mismatch:
        badges.append({"key": "fields", "label": f"Поля расходятся {other_mismatch}"})
    if not_comparable:
        badges.append({"key": "nokey", "label": f"Нет ключа 1С {not_comparable}"})
    balance_label = balance_badge_label(delta)
    if balance_label:
        badges.append({"key": "sum", "label": balance_label})
    if not badges:
        badges.append({"key": "ok", "label": "ОК" if total else "Нет строк сверки"})
    return badges


def build_compare_detail_docs(report: dict[str, object] | None, spec_id: int) -> list[dict[str, object]]:
    if not report:
        return []
    rows = report.get("rows") if isinstance(report.get("rows"), list) else []
    details: list[dict[str, object]] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        badge = compare_status_badge(row)
        erp_num = one_line(row.get("erp_number")) or one_line(row.get("erp_code1c")) or (f"ERP-{sql_int(row.get('erp_doc_id'))}" if sql_int(row.get("erp_doc_id")) else "")
        onec_num = one_line(row.get("onec_number")) or one_line(row.get("onec_code1c")) or one_line(row.get("onec_ref_id"))
        erp_kind = one_line(row.get("erp_doc_kind"))
        doc_type = (
            one_line(row.get("operation_title")) if erp_kind == "act" else ""
        ) or one_line(row.get("erp_type")) or one_line(row.get("onec_type")) or erp_kind or one_line(row.get("onec_doc_kind")) or "Документ"
        details.append(
            detail_doc(
                f"spec-{spec_id}-compare-{idx}",
                f"Сверка 1С: {doc_type}",
                erp_num,
                onec_num,
                one_line(row.get("erp_date_iso")) or one_line(row.get("onec_date_iso")),
                live_money(row.get("erp_sum") if row.get("erp_sum") is not None else row.get("onec_sum")),
                badge["key"],
                badge["label"],
            )
        )
    return details


def build_client_specs_query(client_id: int, dog_id: int, limit: int, scope: str) -> str:
    scope = normalize_text(scope).lower()
    if scope == "contact":
        client_filter = f"(cl.f_contactid = {client_id} OR contact.f_id = {client_id})"
    elif scope == "legal":
        client_filter = f"cl.f_id = {client_id}"
    else:
        client_filter = f"(cl.f_id = {client_id} OR cl.f_contactid = {client_id} OR contact.f_id = {client_id})"
    dog_filter = f"AND d.f_id = {dog_id}" if dog_id > 0 else ""
    return f"""
SELECT
    s.f_id,
    COALESCE(s.f_num, '') AS spec_num,
    COALESCE(s.f_num, '') AS spec_num_short,
    COALESCE(NULLIF(spec_type.f_dopprstr, ''), NULLIF(spec_type.f_name, ''), '') AS spec_type,
    COALESCE(NULLIF(spec_subtype.f_dopprstr, ''), NULLIF(spec_subtype.f_name, ''), '') AS spec_subtype,
    COALESCE(s.f_subtype, 0) AS spec_subtype_id,
    COALESCE(DATE_FORMAT(s.f_dt, '%Y-%m-%d'), '') AS spec_date,
    COALESCE(s.f_tovar, '') AS spec_name,
    COALESCE(d.f_id, 0) AS dog_id,
    COALESCE(d.f_dogname, '') AS dog_number,
    COALESCE(d.f_kod1c, '') AS dog_code1c,
    COALESCE(DATE_FORMAT(d.f_dogdate, '%d.%m.%Y'), '') AS dog_date,
    COALESCE(cl.f_id, 0) AS legal_id,
    COALESCE(cl.f_cname, '') AS legal_name,
    COALESCE(cl.f_abbr, '') AS legal_abbr,
    COALESCE(cl.f_inn, '') AS legal_inn,
    COALESCE(contact.f_id, 0) AS contact_id,
    COALESCE(NULLIF(contact.f_cname, ''), NULLIF(contact.f_name, ''), cl.f_cname, '') AS contact_name,
    COALESCE(contact.f_inn, '') AS contact_inn
FROM veda_specs s
JOIN veda_dogs d
    ON d.f_id = s.f_dogid
JOIN veda_clients cl
    ON cl.f_id = d.f_contrid
LEFT JOIN veda_contacts contact
    ON contact.f_id = cl.f_contactid
LEFT JOIN veda_spr spec_type
    ON spec_type.f_type = 33
   AND spec_type.f_num = s.f_typez
LEFT JOIN veda_spr spec_subtype
    ON spec_subtype.f_type = 130
   AND spec_subtype.f_num = s.f_subtype
WHERE {client_filter}
  {dog_filter}
ORDER BY
    s.f_dt DESC,
    s.f_id DESC
LIMIT {limit};
"""


def fetch_client_specs(client_id: int, dog_id: int, limit: int, scope: str) -> list[dict[str, object]]:
    specs: list[dict[str, object]] = []
    for row in run_mysql_tsv(build_client_specs_query(client_id, dog_id, limit, scope)):
        specs.append(
            {
                "spec_id": int_at(row, 0),
                "spec_num": text_at(row, 1),
                "spec_num_short": text_at(row, 2),
                "spec_type": text_at(row, 3),
                "spec_subtype": text_at(row, 4),
                "spec_subtype_id": int_at(row, 5, 0),
                "spec_date": text_at(row, 6),
                "spec_name": text_at(row, 7),
                "dog_id": int_at(row, 8, 0),
                "dog_number": text_at(row, 9),
                "dog_code1c": text_at(row, 10),
                "dog_date": text_at(row, 11),
                "legal_id": int_at(row, 12, 0),
                "legal_name": text_at(row, 13),
                "legal_abbr": text_at(row, 14),
                "legal_inn": text_at(row, 15),
                "contact_id": int_at(row, 16, 0),
                "contact_name": text_at(row, 17),
                "contact_inn": text_at(row, 18),
            }
        )
    return specs


def build_client_spec_row(
    spec: dict[str, object],
    level: int,
    onec_base_source: dict[str, object] | None = None,
    compare_1c: bool = False,
    source_mode: str = "rest",
) -> dict[str, object]:
    spec_id = sql_int(spec.get("spec_id"))
    operations = fetch_operations(spec_id)
    erp_docs = fetch_erp_docs(spec_id)
    payments = fetch_payments(spec_id)
    invoices = customer_invoice_docs(erp_docs)
    acts = closing_docs(erp_docs)

    paid_by_routine = live_money(sum(live_money(op.get("rp_paid_sum")) for op in operations))
    paid_by_docs = live_money(sum(live_money(pay.get("classified_sum")) for pay in payments if pay.get("direction") == "incoming"))
    reimbursable = live_money(sum(live_money(op.get("rp_realiz_sum")) for op in operations if operation_bucket(op) == "reimbursable"))
    non_reimbursable = live_money(sum(live_money(op.get("rp_realiz_sum")) for op in operations if operation_bucket(op) == "non_reimbursable"))
    unclassified = live_money(sum(live_money(op.get("rp_realiz_sum")) for op in operations if operation_bucket(op) == "unclassified"))
    realization_total = live_money(reimbursable + non_reimbursable + unclassified)
    delta = live_money(paid_by_routine - realization_total)
    invoice_total = live_money(sum(live_money(doc.get("sum")) for doc in invoices))

    issues: list[str] = []
    if not invoices:
        issues.append("NO_CUSTOMER_INVOICE")
    if not acts:
        issues.append("NO_CLOSING_DOC")
    if abs(paid_by_routine - paid_by_docs) > 0.01:
        issues.append("PAYMENT_ROUTINE_VS_DOCS_MISMATCH")
    if abs(delta) > 0.01:
        issues.append("SETTLEMENT_BALANCE_NONZERO")
    if abs(unclassified) > 0.01:
        issues.append("UNCLASSIFIED_REALIZATION")
    detail_docs = build_spec_detail_docs(
        spec_id=spec_id,
        operations=operations,
        invoices=invoices,
        acts=acts,
        payments=payments,
        reimbursable=reimbursable,
        non_reimbursable=non_reimbursable,
    )
    compare_report: dict[str, object] | None = None
    compare_error = ""
    if compare_1c:
        try:
            compare_snapshot = build_spec_snapshot_from_parts(spec_id, operations, erp_docs, payments)
            if source_mode in {"rest", "1c-rest", "onec-rest"}:
                onec_source = load_onec_sources_from_rest(compare_snapshot)
            elif onec_base_source is not None:
                onec_source = filter_onec_postgres_base_source_for_snapshot(onec_base_source, compare_snapshot)
            else:
                onec_source = {}
            if not onec_source:
                raise RuntimeError("Источник 1С не настроен или недоступен")
            compare_report = compare_onec_docs_with_erp_snapshot(spec_id, onec_source, compare_snapshot)
            detail_docs = build_compare_detail_docs(compare_report, spec_id)
        except Exception as exc:
            compare_error = str(exc)
            detail_docs = []
            detail_docs.append(
                detail_doc(
                    f"spec-{spec_id}-compare-error",
                    "Сверка 1С",
                    "ERP",
                    "1С",
                    "",
                    0,
                    "source-error",
                    "Ошибка 1С",
                )
            )

    spec_type = one_line(spec.get("spec_type")) or "Поставка"
    spec_num = one_line(spec.get("spec_num")) or str(spec_id)
    spec_note_bits = [one_line(spec.get("spec_date")), f"spec_id {spec_id}"]
    if spec.get("spec_name"):
        spec_note_bits.insert(1, one_line(spec.get("spec_name")))

    return {
        "id": f"spec-{spec_id}",
        "kind": "spec",
        "level": level,
        "name": f"{spec_type} №{spec_num}",
        "note": " · ".join([item for item in spec_note_bits if item]),
        "specNo": one_line(spec.get("spec_num_short")) or spec_num,
        "isParent": False,
        "defaultExpanded": False,
        "children": [],
        "detailDocs": detail_docs,
        "invoiceLabel": joined_unique_lines([one_line(doc.get("number")) or one_line(doc.get("code1c")) for doc in invoices]),
        "invoiceSum": invoice_total,
        "paymentSum": paid_by_routine,
        "reimbursableSum": reimbursable,
        "nonReimbursableSum": non_reimbursable,
        "sfLabel": joined_unique_lines([document_line(doc) for doc in acts]),
        "delta": delta,
        "badges": [{"key": "source-error", "label": "Ошибка 1С"}] if compare_error else compare_matrix_badges(compare_report, delta, issues),
        "showAmounts": True,
        "meta": {
            "spec_id": spec_id,
            "paid_total_acchist_docs": paid_by_docs,
            "realization_total_get_realizsum": realization_total,
            "unclassified_realization_get_realizsum": unclassified,
            "operations_count": len(operations),
            "customer_invoice_count": len(invoices),
            "closing_docs_count": len(acts),
            "payments_count": len(payments),
            "issues": issues,
            "compare_1c": bool(compare_report),
            "compare_1c_error": compare_error,
        },
        "reconciliation": {
            "source_state": compare_report.get("source_state") if compare_report else ("error" if compare_error else "not_run"),
            "summary": compare_report.get("summary") if compare_report else {},
            "counts": compare_report.get("counts") if compare_report else {},
            "error": compare_error,
        },
    }


def build_client_matrix_snapshot(
    client_id: int,
    dog_id: int = 0,
    limit: int = 25,
    scope: str = "auto",
    compare_1c: bool = False,
    source_mode: str = "rest",
    live_mode: bool = True,
) -> dict[str, object]:
    try:
        requested_limit = int(limit)
    except Exception:
        requested_limit = 25
    source_mode = normalize_text(source_mode).lower() or "rest"
    hard_limit = 100 if live_mode else max(RECON_BACKGROUND_MATRIX_MAX_LIMIT, 1)
    limit = min(max(requested_limit, 1), hard_limit)
    if live_mode and source_mode in {"rest", "1c-rest", "onec-rest"} and ONEC_REST_CLIENT_MATRIX_MAX_LIMIT > 0:
        limit = min(limit, ONEC_REST_CLIENT_MATRIX_MAX_LIMIT)
    specs = fetch_client_specs(client_id, dog_id, limit, scope)
    onec_base_source: dict[str, object] | None = None
    onec_source_error = ""
    if compare_1c:
        if source_mode in {"postgres", "postgresql", "pg", "auto"}:
            try:
                onec_base_source = load_onec_sources_from_postgres(snapshot=None)
            except Exception as exc:
                onec_source_error = str(exc)
                if source_mode in {"postgres", "postgresql", "pg"}:
                    onec_base_source = None
        elif source_mode in {"rest", "1c-rest", "onec-rest"}:
            onec_base_source = None
        else:
            onec_source_error = f"Unsupported client matrix 1C source mode: {source_mode}"
    if not specs:
        return {
            "ok": True,
            "source": "live_client_mariadb",
            "sourceState": "loaded",
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "parameters": {
                "client_id": client_id,
                "dog_id": dog_id,
                "requested_limit": requested_limit,
                "limit": limit,
                "scope": scope,
                "compare_1c": compare_1c,
                "source_mode": source_mode,
                "live_mode": live_mode,
                "limit_reason": "rest_live_limit" if requested_limit != limit and live_mode and source_mode in {"rest", "1c-rest", "onec-rest"} else ("background_hard_limit" if requested_limit != limit and not live_mode else ""),
            },
            "delivery": {
                "request_label": f"client_id {client_id}",
                "client_name": f"client_id {client_id}",
                "client_inn": "",
                "spec_id": 0,
            },
            "matrix_model": {
                "clients": [],
                "totals": {"invoiceSum": 0, "paymentSum": 0, "reimbursableSum": 0, "nonReimbursableSum": 0, "delta": 0},
                "rowsCount": 0,
                "sourceLabel": f"ERP{' + 1С' if compare_1c else ''} · client_id {client_id} · нет поставок",
            },
            "onec_source": {
                "enabled": compare_1c,
                "state": onec_base_source.get("source_state") if onec_base_source else ("error" if onec_source_error else "not_run"),
                "error": onec_source_error,
            },
        }

    contact_groups: dict[int, dict[str, object]] = {}
    for spec in specs:
        contact_id = sql_int(spec.get("contact_id")) or sql_int(spec.get("legal_id"))
        legal_id = sql_int(spec.get("legal_id"))
        dog_key = sql_int(spec.get("dog_id"))
        contact = contact_groups.setdefault(
            contact_id,
            {
                "id": contact_id,
                "name": meaningful_text(spec.get("contact_name")) or meaningful_text(spec.get("legal_name")) or f"client_id {client_id}",
                "inn": one_line(spec.get("contact_inn")),
                "legals": {},
            },
        )
        legals = contact["legals"] if isinstance(contact.get("legals"), dict) else {}
        legal = legals.setdefault(
            legal_id,
            {
                "id": legal_id,
                "name": meaningful_text(spec.get("legal_abbr")) or meaningful_text(spec.get("legal_name")) or f"ЮЛ {legal_id}",
                "inn": one_line(spec.get("legal_inn")),
                "dogs": {},
            },
        )
        dogs = legal["dogs"] if isinstance(legal.get("dogs"), dict) else {}
        dog = dogs.setdefault(
            dog_key,
            {
                "id": dog_key,
                "number": one_line(spec.get("dog_number")) or f"dog_id {dog_key}",
                "code1c": one_line(spec.get("dog_code1c")),
                "date": one_line(spec.get("dog_date")),
                "specs": [],
            },
        )
        dog_specs = dog["specs"] if isinstance(dog.get("specs"), list) else []
        dog_specs.append(spec)

    clients: list[dict[str, object]] = []
    all_spec_rows: list[dict[str, object]] = []
    for contact in contact_groups.values():
        legal_rows: list[dict[str, object]] = []
        legals = contact.get("legals") if isinstance(contact.get("legals"), dict) else {}
        for legal in legals.values():
            dog_rows: list[dict[str, object]] = []
            dogs = legal.get("dogs") if isinstance(legal.get("dogs"), dict) else {}
            for dog in dogs.values():
                spec_rows = [
                    build_client_spec_row(
                        spec,
                        3,
                        onec_base_source=onec_base_source,
                        compare_1c=compare_1c and (onec_base_source is not None or source_mode in {"rest", "1c-rest", "onec-rest"}),
                        source_mode=source_mode,
                    )
                    for spec in (dog.get("specs") if isinstance(dog.get("specs"), list) else [])
                ]
                all_spec_rows.extend(spec_rows)
                dog_totals = aggregate_matrix_rows(spec_rows)
                dog_rows.append(
                    {
                        "id": f"dog-{legal.get('id')}-{dog.get('id')}",
                        "kind": "contract",
                        "level": 2,
                        "name": f"Договор {dog.get('number')}",
                        "note": " · ".join([item for item in [f"dog_id {dog.get('id')}", f"код 1С {dog.get('code1c')}" if dog.get("code1c") else "", one_line(dog.get("date"))] if item]),
                        "specNo": f"{len(spec_rows)} поставок",
                        "isParent": True,
                        "defaultExpanded": True,
                        "children": spec_rows,
                        **dog_totals,
                    }
                )
            legal_totals = aggregate_matrix_rows(dog_rows)
            legal_rows.append(
                {
                    "id": f"legal-{legal.get('id')}",
                    "kind": "legal",
                    "level": 1,
                    "name": f"ЮЛ: {legal.get('name')}",
                    "note": " · ".join([item for item in [f"client_id {legal.get('id')}", f"ИНН {legal.get('inn')}" if legal.get("inn") else ""] if item]),
                    "specNo": f"{len(dog_rows)} договоров",
                    "isParent": True,
                    "defaultExpanded": True,
                    "children": dog_rows,
                    **legal_totals,
                }
            )
        client_totals = aggregate_matrix_rows(legal_rows)
        root = {
            "id": f"client-{contact.get('id')}",
            "kind": "client",
            "level": 0,
            "name": f"Клиент: {contact.get('name')}",
            "note": " · ".join([item for item in [f"contact_id {contact.get('id')}", f"ИНН {contact.get('inn')}" if contact.get("inn") else ""] if item]),
            "specNo": f"{len(legal_rows)} ЮЛ",
            "isParent": True,
            "defaultExpanded": True,
            "children": legal_rows,
            **client_totals,
        }
        clients.append({"id": root["id"], "name": contact.get("name"), "inn": contact.get("inn"), "root": root})

    matrix_totals = aggregate_matrix_rows(all_spec_rows)
    first_spec = specs[0]
    return {
        "ok": True,
        "source": "live_client_mariadb",
        "sourceState": "loaded",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "parameters": {
            "client_id": client_id,
            "dog_id": dog_id,
            "requested_limit": requested_limit,
            "limit": limit,
            "scope": scope,
            "compare_1c": compare_1c,
            "source_mode": source_mode,
            "live_mode": live_mode,
            "limit_reason": "rest_live_limit" if requested_limit != limit and live_mode and source_mode in {"rest", "1c-rest", "onec-rest"} else ("background_hard_limit" if requested_limit != limit and not live_mode else ""),
        },
        "delivery": {
            "request_label": f"Матрица client_id {client_id}",
            "spec_id": 0,
            "client_id": sql_int(first_spec.get("legal_id")),
            "client_name": meaningful_text(first_spec.get("contact_name")) or meaningful_text(first_spec.get("legal_name")),
            "client_inn": meaningful_text(first_spec.get("contact_inn")) or meaningful_text(first_spec.get("legal_inn")),
            "main_dog_id": dog_id or sql_int(first_spec.get("dog_id")),
            "main_dog_number": one_line(first_spec.get("dog_number")),
            "main_dog_code1c": one_line(first_spec.get("dog_code1c")),
        },
        "matrix_model": {
            "clients": clients,
            "totals": matrix_totals,
            "rowsCount": len(all_spec_rows),
            "sourceLabel": f"ERP{' + 1С' if compare_1c else ''} · client_id {client_id} · {len(all_spec_rows)} поставок · limit {limit}",
        },
        "onec_source": {
            "enabled": compare_1c,
            "state": "rest" if source_mode in {"rest", "1c-rest", "onec-rest"} and compare_1c else (onec_base_source.get("source_state") if onec_base_source else ("error" if onec_source_error else "not_run")),
            "error": onec_source_error,
            "docs_count": onec_base_source.get("all_docs_count", 0) if onec_base_source else 0,
            "account_movements_count": onec_base_source.get("all_account_movements_count", 0) if onec_base_source else 0,
            "document_lines_count": onec_base_source.get("all_document_lines_count", 0) if onec_base_source else 0,
        },
        "summary": {
            "specs": len(all_spec_rows),
            "invoice_total": matrix_totals.get("invoiceSum", 0),
            "paid_total_get_paidsum": matrix_totals.get("paymentSum", 0),
            "reimbursable_get_realizsum": matrix_totals.get("reimbursableSum", 0),
            "non_reimbursable_get_realizsum": matrix_totals.get("nonReimbursableSum", 0),
            "delta_paid_minus_realization": matrix_totals.get("delta", 0),
        },
    }


def xlsx_safe(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value)


def setup_sheet(ws, title: str):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24


def style_table(ws, header_row: int, max_col: int, max_row: int):
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    header_font = Font(bold=True, color="374151")
    border = Border(bottom=Side(style="thin", color="D1D5DB"))
    for col in range(1, max_col + 1):
        cell = ws.cell(header_row, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row in range(header_row + 1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, (int, float)) and col > 1:
                cell.number_format = '# ##0.00'
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"


def autosize_columns(ws, max_width: int = 42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def add_table_sheet(wb, sheet_name: str, title: str, headers: list[tuple[str, str]], rows: list[dict[str, object]]):
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, title)
    header_row = 3
    for col_idx, (_, label) in enumerate(headers, start=1):
        ws.cell(header_row, col_idx, label)
    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, (key, _) in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, xlsx_safe(row.get(key)))
    style_table(ws, header_row, len(headers), max(header_row, header_row + len(rows)))
    autosize_columns(ws)
    return ws


def snapshot_to_xlsx(snapshot: dict[str, object]) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed; XLSX export is unavailable")

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    setup_sheet(ws, "Сводка по заявке и взаиморасчетам")

    delivery = snapshot.get("delivery") if isinstance(snapshot.get("delivery"), dict) else {}
    settlements = snapshot.get("settlements") if isinstance(snapshot.get("settlements"), dict) else {}

    summary_rows = [
        ("Заявка", delivery.get("request_label")),
        ("spec_id", snapshot.get("spec_id")),
        ("Дата заявки", delivery.get("spec_date")),
        ("Организация", delivery.get("org_name")),
        ("Покупатель", delivery.get("client_name")),
        ("ИНН покупателя", delivery.get("client_inn")),
        ("Основной договор", delivery.get("main_dog_number")),
        ("Код договора 1С", delivery.get("main_dog_code1c")),
        ("Менеджер", delivery.get("manager_name")),
        ("Сумма счетов покупателю", settlements.get("buyer_invoice_total")),
        ("Сумма оплат покупателя (get_paidsum)", settlements.get("paid_total")),
        ("Сумма оплат по связям выписки", settlements.get("classified_paid_total")),
        ("Начисленные расходы (get_expensessum)", settlements.get("expenses_accrued_total")),
        ("Реализация / отчет агента (get_realizsum)", settlements.get("realization_total")),
        ("Прибыль / убыток (get_profit)", settlements.get("profit_total")),
        ("Реализация - расходы", settlements.get("realization_minus_expenses_total")),
        ("Оплаченные расходы по выпискам", settlements.get("expenses_paid_total")),
        ("Возмещаемые расходы", settlements.get("reimbursable_total")),
        ("Невозмещаемые расходы", settlements.get("non_reimbursable_total")),
        ("Контроль Sheet: возмещаемые расходы", settlements.get("control_reimbursable_total")),
        ("Контроль Sheet: невозмещаемые услуги", settlements.get("control_non_reimbursable_total")),
        ("Контроль Sheet: (+/-)", settlements.get("control_sheet_delta")),
        ("Операций f_addnds", settlements.get("add_nds_operations_count")),
        ("Сумма операций f_addnds", settlements.get("add_nds_oper_total")),
        ("Начислено расходов f_addnds", settlements.get("add_nds_expenses_total")),
        ("Прибыль/убыток f_addnds", settlements.get("add_nds_profit_total")),
        ("Контроль Sheet: исключено расходных операций", settlements.get("control_excluded_cost_operations_count")),
        ("Контроль Sheet: сумма исключенных операций", settlements.get("control_excluded_cost_oper_total")),
        ("Контроль Sheet: расходы исключенных операций", settlements.get("control_excluded_cost_expenses_total")),
        ("Контроль Sheet: прибыль исключенных операций", settlements.get("control_excluded_cost_profit_total")),
        ("Сальдо взаиморасчетов в руб. (get_paidsum - get_realizsum)", settlements.get("customer_settlement_balance_rub")),
        ("Тех. баланс по счетам: счета - оплаты", settlements.get("balance_to_buyer")),
        ("Итог взаиморасчетов", settlements.get("settlement_delta")),
    ]
    ws["A3"] = "Показатель"
    ws["B3"] = "Значение"
    for idx, (name, value) in enumerate(summary_rows, start=4):
        ws.cell(idx, 1, name)
        ws.cell(idx, 2, xlsx_safe(value))
        if isinstance(value, (int, float)):
            ws.cell(idx, 2).number_format = '# ##0.00'
    style_table(ws, 3, 2, 3 + len(summary_rows))
    autosize_columns(ws, max_width=56)

    add_table_sheet(
        wb,
        "Договоры",
        "Договоры поставки: покупатель и коммитент",
        [
            ("role_label", "Роль"),
            ("dog_id", "ID договора"),
            ("dog_number", "Номер договора"),
            ("dog_date", "Дата договора"),
            ("dog_code1c", "Код 1С договора"),
            ("dog_type_name", "Тип договора"),
            ("org_name", "Организация"),
            ("org_inn", "ИНН организации"),
            ("client_name", "Контрагент"),
            ("client_inn", "ИНН контрагента"),
            ("source_roles", "Источники связи"),
            ("operations_count", "Операций"),
            ("operations_sum", "Сумма операций"),
        ],
        snapshot.get("contracts") if isinstance(snapshot.get("contracts"), list) else [],
    )

    add_table_sheet(
        wb,
        "Счета",
        "Счета по заявке",
        [
            ("erp_doc_id", "ID счета"),
            ("number", "Номер счета"),
            ("date", "Дата"),
            ("sum", "Сумма"),
            ("currency", "Валюта"),
            ("type_name", "Тип счета"),
            ("status_name", "Статус ERP"),
            ("onec_export_state", "Состояние 1С"),
            ("code1c", "Код 1С счета"),
            ("dog_number", "Договор"),
            ("dog_code1c", "Код 1С договора"),
            ("oper_id", "ID операции"),
            ("oper_type_name", "Операция"),
            ("reimbursement_name", "Возмещаемость"),
            ("nds_name", "НДС"),
            ("detail_name", "Основание"),
        ],
        snapshot.get("schets") if isinstance(snapshot.get("schets"), list) else [],
    )

    add_table_sheet(
        wb,
        "Оплаты",
        "Оплаты и банковские связи",
        [
            ("payment_id", "ID платежа"),
            ("direction", "Направление"),
            ("pp_number", "Номер ПП"),
            ("payment_date", "Дата"),
            ("payment_sum", "Сумма платежа"),
            ("classified_sum", "Сумма по операции"),
            ("currency", "Валюта"),
            ("payer_name", "Контрагент/банк"),
            ("invoice_number", "Счет"),
            ("oper_id", "ID операции"),
            ("oper_type_name", "Операция"),
            ("reimbursement_name", "Возмещаемость"),
        ],
        snapshot.get("payments") if isinstance(snapshot.get("payments"), list) else [],
    )

    add_table_sheet(
        wb,
        "Акты",
        "Акты и детализация",
        [
            ("erp_doc_id", "ID акта"),
            ("number", "Номер"),
            ("date", "Дата"),
            ("sum", "Сумма документа"),
            ("detail_sum", "Сумма детализации"),
            ("currency", "Валюта"),
            ("type_name", "Тип документа"),
            ("status_name", "Статус ERP"),
            ("onec_export_state", "Состояние 1С"),
            ("code1c", "Код 1С"),
            ("dog_number", "Договор"),
            ("dog_code1c", "Код 1С договора"),
            ("invoice_number", "Счет/тип возмещаемости"),
            ("oper_id", "ID операции"),
            ("oper_type_name", "Операция"),
            ("reimbursement_name", "Возмещаемость"),
            ("detail_name", "Детализация"),
        ],
        snapshot.get("akts") if isinstance(snapshot.get("akts"), list) else [],
    )

    add_table_sheet(
        wb,
        "Операции",
        "Операции veda_spec_invoices",
        [
            ("oper_id", "ID операции"),
            ("oper_num", "№"),
            ("oper_type_name", "Операция"),
            ("oper_sum", "Сумма"),
            ("rp_paid_sum", "Оплачено клиентом (get_paidsum)"),
            ("rp_expenses_sum", "Начислено расходов (get_expensessum)"),
            ("rp_realiz_sum", "Реализация/отчет агента (get_realizsum)"),
            ("rp_profit_sum", "Прибыль/убыток (get_profit)"),
            ("out_buh_period", "Период вывода в БУ"),
            ("add_nds_flag", "f_addnds"),
            ("hnds_flag", "f_hnds"),
            ("oper_currency", "Валюта"),
            ("invoice_id", "ID счета"),
            ("dog_number", "Договор"),
            ("dog_code1c", "Код 1С договора"),
            ("reimbursement_name", "Возмещаемость"),
            ("nds_name", "НДС"),
            ("bdr_article_name", "БДР"),
            ("c1_doc_type_id", "Тип 1С"),
            ("parent_oper_id", "Родительская операция"),
        ],
        snapshot.get("operations") if isinstance(snapshot.get("operations"), list) else [],
    )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def split_matrix_lines(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text or text == "—":
        return []
    return [one_line(part) for part in text.splitlines() if one_line(part)]


def collect_matrix_spec_rows(node: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if node.get("kind") == "spec":
        rows.append(node)
    children = node.get("children")
    if isinstance(children, list):
        for child in children:
            if isinstance(child, dict):
                rows.extend(collect_matrix_spec_rows(child))
    return rows


def matrix_snapshot_to_accounting_xlsx(snapshot: dict[str, object]) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed; XLSX export is unavailable")

    wb = Workbook()
    ws = wb.active
    ws.title = "Выгрузка"

    headers = [
        "№ спецификации",
        "",
        "Счет",
        "Сумма по счету",
        "Сумма оплаты",
        "",
        "Возмещаемые расходы",
        "Невозмещаемые расходы",
        "",
        "№ счф",
        "(+/-)",
    ]
    ws.append(headers)

    fills = {
        "header": PatternFill("solid", fgColor="D9D7D2"),
        "spec": PatternFill("solid", fgColor="F3F1ED"),
        "invoice": PatternFill("solid", fgColor="EEF6FF"),
        "payment": PatternFill("solid", fgColor="EEF8F0"),
        "expense": PatternFill("solid", fgColor="FFF5DF"),
        "sf": PatternFill("solid", fgColor="F7F4EF"),
        "delta_ok": PatternFill("solid", fgColor="E2F0D9"),
        "delta_bad": PatternFill("solid", fgColor="FCE4D6"),
        "sep": PatternFill("solid", fgColor="FAFAF9"),
    }
    border = Border(
        left=Side(style="thin", color="D6D3CD"),
        right=Side(style="thin", color="D6D3CD"),
        top=Side(style="thin", color="D6D3CD"),
        bottom=Side(style="thin", color="D6D3CD"),
    )
    money_format = '#,##0.00" р.";[Red]-#,##0.00" р.";0.00" р."'

    for cell in ws[1]:
        cell.fill = fills["header"]
        cell.font = Font(bold=True, color="1F2933")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    model = snapshot.get("matrix_model") if isinstance(snapshot.get("matrix_model"), dict) else {}
    clients = model.get("clients") if isinstance(model.get("clients"), list) else []
    spec_rows: list[dict[str, object]] = []
    for client in clients:
        if isinstance(client, dict) and isinstance(client.get("root"), dict):
            spec_rows.extend(collect_matrix_spec_rows(client["root"]))

    row_idx = 2
    merge_columns = [1, 5, 7, 8, 10, 11]
    for spec in spec_rows:
        detail_docs = spec.get("detailDocs") if isinstance(spec.get("detailDocs"), list) else []
        invoice_docs = [
            doc for doc in detail_docs
            if isinstance(doc, dict) and "Счет покупателю" in one_line(doc.get("type"))
        ]
        if invoice_docs:
            invoice_lines = [
                {
                    "number": one_line(doc.get("erp")) or one_line(doc.get("onec")),
                    "amount": live_money(doc.get("amount")),
                }
                for doc in invoice_docs
            ]
        else:
            labels = split_matrix_lines(spec.get("invoiceLabel"))
            invoice_lines = [{"number": label, "amount": None} for label in labels]
        if not invoice_lines:
            invoice_lines = [{"number": "—", "amount": None}]

        start = row_idx
        for idx, invoice in enumerate(invoice_lines):
            ws.cell(row_idx, 3, invoice.get("number") or "—")
            amount = invoice.get("amount")
            if isinstance(amount, (int, float)) and abs(float(amount)) > 0.01:
                ws.cell(row_idx, 4, amount)
            elif idx == 0:
                ws.cell(row_idx, 4, live_money(spec.get("invoiceSum")))
            row_idx += 1
        end = row_idx - 1

        ws.cell(start, 1, spec.get("name") or spec.get("specNo") or "—")
        ws.cell(start, 5, live_money(spec.get("paymentSum")))
        ws.cell(start, 7, live_money(spec.get("reimbursableSum")))
        ws.cell(start, 8, live_money(spec.get("nonReimbursableSum")))
        ws.cell(start, 10, "\n".join(split_matrix_lines(spec.get("sfLabel"))) or "—")
        ws.cell(start, 11, live_money(spec.get("delta")))

        if end > start:
            for col in merge_columns:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)

        for r in range(start, end + 1):
            for c in range(1, 12):
                cell = ws.cell(r, c)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(r, 1).fill = fills["spec"]
            ws.cell(r, 2).fill = fills["sep"]
            ws.cell(r, 3).fill = fills["invoice"]
            ws.cell(r, 4).fill = fills["invoice"]
            ws.cell(r, 5).fill = fills["payment"]
            ws.cell(r, 6).fill = fills["sep"]
            ws.cell(r, 7).fill = fills["expense"]
            ws.cell(r, 8).fill = fills["expense"]
            ws.cell(r, 9).fill = fills["sep"]
            ws.cell(r, 10).fill = fills["sf"]
            ws.cell(r, 11).fill = fills["delta_ok"] if abs(live_money(spec.get("delta"))) <= 0.01 else fills["delta_bad"]
        for col in [4, 5, 7, 8, 11]:
            ws.cell(start, col).number_format = money_format
            ws.cell(start, col).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.cell(start, 1).font = Font(bold=True)
        sf_lines = max(1, len(split_matrix_lines(spec.get("sfLabel"))))
        ws.row_dimensions[start].height = max(28, 16 * sf_lines)

    totals = model.get("totals") if isinstance(model.get("totals"), dict) else {}
    total_row = row_idx + 1
    ws.cell(total_row, 1, "ИТОГО")
    ws.cell(total_row, 4, live_money(totals.get("invoiceSum")))
    ws.cell(total_row, 5, live_money(totals.get("paymentSum")))
    ws.cell(total_row, 7, live_money(totals.get("reimbursableSum")))
    ws.cell(total_row, 8, live_money(totals.get("nonReimbursableSum")))
    ws.cell(total_row, 11, live_money(totals.get("delta")))
    for c in range(1, 12):
        cell = ws.cell(total_row, c)
        cell.fill = fills["header"]
        cell.border = border
        cell.font = Font(bold=True)
    for col in [4, 5, 7, 8, 11]:
        ws.cell(total_row, col).number_format = money_format

    rules = wb.create_sheet("Правила")
    rules_rows = [
        ("Поле", "Источник", "Правило"),
        ("№ спецификации", "ERP veda_specs + veda_spr(f_type=33/130)", "Тип и номер поставки из исходных таблиц ERP; view_specinv/view_specs не использовать."),
        ("Счет", "ERP veda_schets, f_type=1", "Только счета покупателю; счета поставщиков не попадают в эту колонку."),
        ("Сумма по счету", "ERP veda_schets.f_sum", "По каждому счету отдельная строка; разные валюты не суммировать без курса."),
        ("Сумма оплаты", "ERP get_paidsum / veda_acchist_docs.f_clssum", "Сумма оплат клиента по поставке."),
        ("Возмещаемые расходы", "ERP get_realizsum по операциям f_isvozm=1", "Итог по поставке, объединяется по строкам счетов."),
        ("Невозмещаемые расходы", "ERP get_realizsum по операциям f_isvozm=2", "Итог по поставке, объединяется по строкам счетов."),
        ("№ счф", "ERP/1С закрывающие документы", "Каждый номер документа внутри объединенной ячейки с переносом строки."),
        ("(+/-)", "Сумма оплаты - возмещаемые - невозмещаемые", "Сальдо взаиморасчетов по поставке."),
    ]
    for row in rules_rows:
        rules.append(row)
    for cell in rules[1]:
        cell.fill = fills["header"]
        cell.font = Font(bold=True)
    for row_cells in rules.iter_rows():
        for cell in row_cells:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 24,
        "B": 3,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 3,
        "G": 20,
        "H": 22,
        "I": 3,
        "J": 48,
        "K": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    for sheet in [rules]:
        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 26 if col_idx > 1 else 20

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


class ReconciliationApiHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_ROOT), **kwargs)

    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.NO_CONTENT)
        self.end_headers()

    def write_json(self, status: int, payload: dict[str, object]):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def write_binary(self, status: int, body: bytes, content_type: str, filename: str | None = None):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(body)

    def read_json_body(self) -> dict[str, object]:
        length = int(self.headers.get("Content-Length") or "0")
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        try:
            payload = json.loads(raw.decode("utf-8"))
            return payload if isinstance(payload, dict) else {}
        except Exception:
            return {}

    def bearer_token(self) -> str:
        header = self.headers.get("Authorization") or ""
        if header.lower().startswith("bearer "):
            return header[7:].strip()
        return ""

    def current_auth_session(self) -> dict[str, object] | None:
        return get_auth_session(self.bearer_token())

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/auth/config":
            self.handle_auth_config()
            return
        if parsed.path == "/api/auth/me":
            self.handle_auth_me()
            return
        if parsed.path == "/api/reconciliation/erp-snapshot":
            self.handle_erp_snapshot(parsed.query)
            return
        if parsed.path == "/api/reconciliation/client-matrix":
            self.handle_client_matrix(parsed.query)
            return
        if parsed.path == "/api/reconciliation/client-matrix.xlsx":
            self.handle_client_matrix_xlsx(parsed.query)
            return
        if parsed.path == "/api/reconciliation/erp-export.xlsx":
            self.handle_erp_export_xlsx(parsed.query)
            return
        if parsed.path == "/api/reconciliation/1c-rest-status":
            self.handle_1c_rest_status()
            return
        if parsed.path == "/api/reconciliation/job-status":
            self.handle_job_status(parsed.query, include_result=False)
            return
        if parsed.path == "/api/reconciliation/job-result":
            self.handle_job_status(parsed.query, include_result=True)
            return
        if parsed.path == "/api/reconciliation/erp-docs":
            self.handle_erp_docs(parsed.query)
            return
        super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/auth/erp-launch":
            self.handle_auth_erp_launch()
            return
        if parsed.path == "/api/auth/logout":
            self.handle_auth_logout()
            return
        if parsed.path == "/api/reconciliation/run":
            self.handle_run()
            return
        if parsed.path == "/api/reconciliation/compare":
            self.handle_compare()
            return
        if parsed.path == "/api/reconciliation/client-matrix-job":
            self.handle_client_matrix_job(parsed.query)
            return
        self.write_json(HTTPStatus.NOT_FOUND, {"ok": False, "error": "Not found"})

    def handle_auth_config(self):
        self.write_json(HTTPStatus.OK, {"ok": True, "auth": auth_config_payload()})

    def handle_auth_erp_launch(self):
        payload = self.read_json_body()
        launch_token = normalize_text(payload.get("launch_token") or payload.get("token"))
        if not launch_token:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "ERP launch token is required"})
            return
        try:
            identity_profile = validate_erp_launch_token(launch_token)
            user = fetch_erp_user_by_identity(identity_profile)
            session = create_auth_session(user, erp_token="", identity_profile=identity_profile)
            self.write_json(HTTPStatus.OK, {"ok": True, **public_auth_session(session, include_erp_token=False)})
        except Exception as exc:
            self.write_json(HTTPStatus.UNAUTHORIZED, {"ok": False, "error": str(exc), "auth": auth_config_payload()})

    def handle_auth_me(self):
        session = self.current_auth_session()
        if not session:
            self.write_json(HTTPStatus.OK, {"ok": True, "authenticated": False, "auth": auth_config_payload()})
            return
        self.write_json(HTTPStatus.OK, {"ok": True, **public_auth_session(session, include_erp_token=False), "auth": auth_config_payload()})

    def handle_auth_logout(self):
        token = self.bearer_token()
        if token:
            with AUTH_SESSIONS_LOCK:
                AUTH_SESSIONS.pop(token, None)
        self.write_json(HTTPStatus.OK, {"ok": True, "authenticated": False})

    def handle_erp_docs(self, query_str: str):
        params = parse_qs(query_str)
        spec_raw = (params.get("spec_id") or params.get("specId") or [None])[0]
        try:
            spec_id = int(str(spec_raw))
            if spec_id <= 0:
                raise ValueError("spec_id must be positive")
        except Exception:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return

        try:
            operations = fetch_operations(spec_id)
            docs = fetch_erp_docs(spec_id)
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(
            HTTPStatus.OK,
            {
                "ok": True,
                "spec_id": spec_id,
                "operations": operations,
                "erp_docs": docs,
                "operation_url_template": OPER_URL_TEMPLATE,
            },
        )

    def parse_spec_id_from_query(self, query_str: str) -> int | None:
        params = parse_qs(query_str)
        spec_raw = (params.get("spec_id") or params.get("specId") or [None])[0]
        try:
            spec_id = int(str(spec_raw))
            if spec_id <= 0:
                raise ValueError("spec_id must be positive")
            return spec_id
        except Exception:
            return None

    def parse_positive_int(self, params: dict[str, list[str]], keys: list[str], default: int | None = None) -> int | None:
        raw = None
        for key in keys:
            values = params.get(key)
            if values:
                raw = values[0]
                break
        if raw is None:
            return default
        try:
            value = int(str(raw))
            if value <= 0:
                raise ValueError("value must be positive")
            return value
        except Exception:
            return default

    def handle_erp_snapshot(self, query_str: str):
        spec_id = self.parse_spec_id_from_query(query_str)
        if not spec_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return

        try:
            snapshot = build_erp_snapshot(spec_id)
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(HTTPStatus.OK, snapshot)

    def handle_client_matrix(self, query_str: str):
        params = parse_qs(query_str)
        client_id = self.parse_positive_int(params, ["client_id", "clientId", "legal_id", "contact_id"])
        if not client_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "client_id is required and must be positive integer"})
            return
        dog_id = self.parse_positive_int(params, ["dog_id", "dogId"], 0) or 0
        limit = self.parse_positive_int(params, ["limit"], 25) or 25
        scope = normalize_text((params.get("scope") or [params.get("client_scope", ["auto"])[0]])[0]).lower() or "auto"
        if scope not in {"auto", "legal", "contact"}:
            scope = "auto"
        compare_raw = normalize_text((params.get("compare_1c") or params.get("compare1c") or params.get("with_1c") or ["0"])[0]).lower()
        compare_1c = compare_raw not in {"0", "false", "no", "off"}
        source_mode = normalize_text((params.get("source") or params.get("source_mode") or ["rest"])[0]).lower() or "rest"

        try:
            snapshot = build_client_matrix_snapshot(
                client_id=client_id,
                dog_id=dog_id,
                limit=limit,
                scope=scope,
                compare_1c=compare_1c,
                source_mode=source_mode,
            )
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(HTTPStatus.OK, snapshot)

    def handle_client_matrix_job(self, query_str: str):
        params = parse_qs(query_str)
        payload = self.read_json_body()

        def pick(keys: list[str], default: object = None) -> object:
            for key in keys:
                if isinstance(payload, dict) and payload.get(key) not in (None, ""):
                    return payload.get(key)
                if params.get(key) and params[key][0] not in (None, ""):
                    return params[key][0]
            return default

        client_id = sql_int(pick(["client_id", "clientId", "legal_id", "contact_id"], 0))
        if client_id <= 0:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "client_id is required and must be positive integer"})
            return
        dog_id = sql_int(pick(["dog_id", "dogId"], 0))
        limit = sql_int(pick(["limit"], 100))
        scope = normalize_text(pick(["scope", "client_scope"], "auto")).lower() or "auto"
        if scope not in {"auto", "legal", "contact"}:
            scope = "auto"
        compare_raw = normalize_text(pick(["compare_1c", "compare1c", "with_1c"], "0")).lower()
        compare_1c = compare_raw not in {"0", "false", "no", "off"}
        source_mode = normalize_text(pick(["source", "source_mode"], "rest")).lower() or "rest"

        job_params = {
            "client_id": client_id,
            "dog_id": dog_id,
            "limit": limit,
            "scope": scope,
            "compare_1c": compare_1c,
            "source_mode": source_mode,
            "live_mode": False,
        }

        def runner():
            return build_client_matrix_snapshot(
                client_id=client_id,
                dog_id=dog_id,
                limit=limit,
                scope=scope,
                compare_1c=compare_1c,
                source_mode=source_mode,
                live_mode=False,
            )

        job = start_recon_job("client_matrix", job_params, runner)
        self.write_json(HTTPStatus.ACCEPTED, {"ok": True, "job": job, "jobs": recon_jobs_summary()})

    def handle_job_status(self, query_str: str, include_result: bool = False):
        params = parse_qs(query_str)
        job_id = normalize_text((params.get("job_id") or params.get("id") or [""])[0])
        if not job_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "job_id is required"})
            return
        job = get_recon_job(job_id, include_result=include_result)
        if job is None:
            self.write_json(HTTPStatus.NOT_FOUND, {"ok": False, "error": "job not found or expired", "jobs": recon_jobs_summary()})
            return
        self.write_json(HTTPStatus.OK, {"ok": True, "job": job, "jobs": recon_jobs_summary()})

    def handle_client_matrix_xlsx(self, query_str: str):
        params = parse_qs(query_str)
        client_id = self.parse_positive_int(params, ["client_id", "clientId", "legal_id", "contact_id"])
        if not client_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "client_id is required and must be positive integer"})
            return
        dog_id = self.parse_positive_int(params, ["dog_id", "dogId"], 0) or 0
        limit = self.parse_positive_int(params, ["limit"], 25) or 25
        scope = normalize_text((params.get("scope") or [params.get("client_scope", ["auto"])[0]])[0]).lower() or "auto"
        if scope not in {"auto", "legal", "contact"}:
            scope = "auto"
        compare_raw = normalize_text((params.get("compare_1c") or params.get("compare1c") or params.get("with_1c") or ["0"])[0]).lower()
        compare_1c = compare_raw not in {"0", "false", "no", "off"}
        source_mode = normalize_text((params.get("source") or params.get("source_mode") or ["rest"])[0]).lower() or "rest"

        try:
            snapshot = build_client_matrix_snapshot(
                client_id=client_id,
                dog_id=dog_id,
                limit=limit,
                scope=scope,
                compare_1c=compare_1c,
                source_mode=source_mode,
            )
            body = matrix_snapshot_to_accounting_xlsx(snapshot)
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        filename_bits = [f"client_{client_id}"]
        if dog_id:
            filename_bits.append(f"dog_{dog_id}")
        filename_bits.append("matrix")
        filename = "akt_sverki_" + "_".join(filename_bits) + ".xlsx"
        self.write_binary(
            HTTPStatus.OK,
            body,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename,
        )

    def handle_erp_export_xlsx(self, query_str: str):
        spec_id = self.parse_spec_id_from_query(query_str)
        if not spec_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return

        try:
            snapshot = build_erp_snapshot(spec_id)
            body = snapshot_to_xlsx(snapshot)
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        filename = f"akt_sverki_erp_spec_{spec_id}.xlsx"
        self.write_binary(
            HTTPStatus.OK,
            body,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename,
        )

    def handle_control_sheet(self, query_str: str):
        spec_id = self.parse_spec_id_from_query(query_str)
        if not spec_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return
        params = parse_qs(query_str)
        spreadsheet_id = normalize_text((params.get("spreadsheet_id") or params.get("sheet_id") or [CONTROL_SHEET_ID])[0])
        gid = normalize_text((params.get("gid") or [CONTROL_SHEET_GID])[0])
        sheet_title = normalize_text((params.get("sheet") or params.get("sheet_title") or [""])[0])
        try:
            payload = build_three_way_control(spec_id, spreadsheet_id, gid, sheet_title)
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return
        self.write_json(HTTPStatus.OK, payload)

    def build_xlsx_reconciliation_from_paths(self, spec_id: int, paths: list[Path]) -> dict[str, object]:
        snapshot = build_erp_snapshot(spec_id)
        onec_source = parse_onec_sources(paths, snapshot=snapshot)
        return compare_onec_docs_with_erp_snapshot(spec_id, onec_source, snapshot)

    def build_default_reconciliation(self, spec_id: int, source_mode: str = "auto") -> dict[str, object]:
        snapshot = build_erp_snapshot(spec_id)
        source_mode = normalize_text(source_mode).lower() or "auto"
        pg_warning = ""

        if source_mode in {"rest", "1c-rest", "onec-rest"}:
            onec_source = load_onec_sources_from_rest(snapshot)
            return compare_onec_docs_with_erp_snapshot(spec_id, onec_source, snapshot)

        if source_mode in {"auto", "postgres", "postgresql", "pg"}:
            try:
                pg_source = load_onec_sources_from_postgres(snapshot=snapshot)
                if pg_source.get("all_docs_count", 0) or source_mode != "auto":
                    return compare_onec_docs_with_erp_snapshot(spec_id, pg_source, snapshot)
            except Exception as exc:
                pg_warning = f"PostgreSQL source unavailable: {exc}"
                if source_mode in {"postgres", "postgresql", "pg"}:
                    raise

        if not ONEC_DEFAULT_DIR.exists():
            raise FileNotFoundError(f"1C directory not found: {ONEC_DEFAULT_DIR}")

        onec_source = parse_onec_sources([ONEC_DEFAULT_DIR], snapshot=snapshot)
        if pg_warning:
            warnings = onec_source.get("warnings") if isinstance(onec_source.get("warnings"), list) else []
            warnings.append(pg_warning)
            onec_source["warnings"] = warnings
        return compare_onec_docs_with_erp_snapshot(spec_id, onec_source, snapshot)

    def handle_compare_1c_folder(self, query_str: str):
        spec_id = self.parse_spec_id_from_query(query_str)
        if not spec_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return

        try:
            params = parse_qs(query_str)
            source_mode = (params.get("source") or ["auto"])[0]
            report = self.build_default_reconciliation(spec_id, source_mode=source_mode)
        except FileNotFoundError as exc:
            self.write_json(HTTPStatus.NOT_FOUND, {"ok": False, "error": str(exc)})
            return
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(HTTPStatus.OK, report)

    def handle_1c_source_summary(self):
        pg_summary = onec_postgres_summary()
        xlsx_payload: dict[str, object] = {
            "ok": False,
            "all_docs_count": 0,
            "docs_count": 0,
            "all_account_movements_count": 0,
            "account_movements_count": 0,
            "cache_hit": False,
            "by_kind": {},
            "files": [],
            "warnings": [],
        }
        if not ONEC_DEFAULT_DIR.exists():
            xlsx_payload["error"] = f"1C directory not found: {ONEC_DEFAULT_DIR}"
        else:
            try:
                source = parse_onec_sources([ONEC_DEFAULT_DIR])
                xlsx_payload = {
                    "ok": True,
                    "all_docs_count": source.get("all_docs_count", 0),
                    "docs_count": source.get("docs_count", 0),
                    "all_account_movements_count": source.get("all_account_movements_count", 0),
                    "account_movements_count": source.get("account_movements_count", 0),
                    "cache_hit": source.get("cache_hit", False),
                    "by_kind": source.get("by_kind", {}),
                    "files": source.get("files", []),
                    "warnings": source.get("warnings", []),
                    "contract_pairs_sample": source.get("contract_pairs", [])[:50],
                }
            except Exception as exc:
                xlsx_payload["error"] = str(exc)

        source_state = "postgresql" if pg_summary.get("docs_count", 0) or pg_summary.get("account_movements_count", 0) else "xlsx"
        self.write_json(
            HTTPStatus.OK,
            {
                "ok": True,
                "source_state": source_state,
                "postgresql": pg_summary,
                "xlsx": xlsx_payload,
                "rest": onec_rest_status(),
                "jobs": recon_jobs_summary(),
            },
        )

    def handle_1c_rest_status(self):
        status = onec_rest_status()
        self.write_json(HTTPStatus.OK, {"ok": True, "source": "1c_rest", **status, "jobs": recon_jobs_summary()})

    def handle_sync_1c_drive(self, query_str: str):
        params = parse_qs(query_str)
        folder_id = (params.get("folder_id") or params.get("folderId") or [None])[0]
        if not folder_id:
            payload = self.read_json_body()
            folder_id = payload.get("folder_id") or payload.get("folderId")
        try:
            result = sync_onec_drive_to_postgres(folder_id=normalize_text(folder_id))
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(
            HTTPStatus.OK,
            {
                "ok": True,
                "source_state": "postgresql",
                "sync": result,
                "postgresql": onec_postgres_summary(),
            },
        )

    def read_multipart_xlsx_files(self) -> list[tuple[str, bytes]]:
        content_type = self.headers.get("Content-Type") or ""
        if "multipart/form-data" not in content_type:
            raise ValueError("multipart/form-data is required")
        length = int(self.headers.get("Content-Length") or "0")
        if length <= 0:
            raise ValueError("empty request body")

        body = self.rfile.read(length)
        raw_message = (
            f"Content-Type: {content_type}\r\n"
            "MIME-Version: 1.0\r\n\r\n"
        ).encode("utf-8") + body
        message = BytesParser(policy=EMAIL_POLICY).parsebytes(raw_message)

        files: list[tuple[str, bytes]] = []
        for part in message.iter_parts():
            filename = normalize_text(part.get_filename())
            if not filename:
                continue
            payload = part.get_payload(decode=True) or b""
            if not payload:
                continue
            lower = filename.lower()
            if lower.endswith(".xlsx"):
                files.append((Path(filename).name, payload))
        return files

    def handle_upload_1c(self, query_str: str):
        spec_id = self.parse_spec_id_from_query(query_str)
        if not spec_id:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return

        try:
            upload_files = self.read_multipart_xlsx_files()
            if not upload_files:
                self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "no .xlsx files uploaded"})
                return

            with tempfile.TemporaryDirectory(prefix="onec_upload_") as tmp_dir:
                paths: list[Path] = []
                for filename, payload in upload_files:
                    safe_name = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._() -]+", "_", filename).strip(" .") or "upload.xlsx"
                    path = Path(tmp_dir) / safe_name
                    path.write_bytes(payload)
                    paths.append(path)

                report = self.build_xlsx_reconciliation_from_paths(spec_id, paths)
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(HTTPStatus.OK, report)

    def handle_compare(self):
        payload = self.read_json_body()
        spec_raw = payload.get("spec_id")
        try:
            spec_id = int(str(spec_raw))
            if spec_id <= 0:
                raise ValueError("spec_id must be positive")
        except Exception:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return

        try:
            report = compare_documents(
                spec_id,
                payload.get("onec_docs"),
                source_state="available",
                scope="specification",
                scope_id=spec_id,
                source_mode="manual-import",
            )
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(
            HTTPStatus.OK,
            {
                "ok": True,
                "spec_id": spec_id,
                "run_id": report.get("run_id"),
                "summary": report.get("summary"),
                "rows": report.get("rows"),
                "counts": {
                    "erp_docs": report.get("erp_docs_count", 0),
                    "onec_docs": report.get("onec_docs_count", 0),
                },
                "source_state": report.get("source_state", "manual"),
            },
        )

    def handle_run(self):
        payload = self.read_json_body()
        spec_raw = payload.get("spec_id")
        scope = normalize_text(payload.get("scope") or "specification").lower()
        scope_id = sql_int(payload.get("scope_id")) if payload.get("scope_id") is not None else None
        if scope not in {"specification", "client"}:
            scope = "specification"
        if scope == "specification":
            scope_id = scope_id if scope_id is not None and scope_id > 0 else None

        try:
            spec_id = int(str(spec_raw))
            if spec_id <= 0:
                raise ValueError("spec_id must be positive")
        except Exception:
            self.write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "spec_id is required and must be positive integer"})
            return

        try:
            onec_docs, source_state = load_onec_docs_from_connector(spec_id, scope=scope, scope_id=scope_id)
            report = compare_documents(
                spec_id,
                onec_docs,
                source_state=source_state,
                scope=scope,
                scope_id=(scope_id if scope_id is not None else spec_id),
                client_id=(scope_id if scope == "client" and scope_id is not None else 0),
                source_mode="server-run",
            )
            operations = fetch_operations(spec_id)
            erp_docs = fetch_erp_docs(spec_id)
        except Exception as exc:
            self.write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
            return

        self.write_json(
            HTTPStatus.OK,
            {
                "ok": True,
                "spec_id": spec_id,
                "scope": scope,
                "scope_id": scope_id,
                "run_id": report.get("run_id"),
                "summary": report.get("summary"),
                "rows": report.get("rows"),
                "counts": {
                    "erp_docs": report.get("erp_docs_count", 0),
                    "onec_docs": report.get("onec_docs_count", 0),
                },
                "operations": operations,
                "erp_docs": erp_docs,
                "source_state": report.get("source_state", "unknown"),
            },
        )


def main():
    server = ThreadingHTTPServer((LISTEN_HOST, LISTEN_PORT), ReconciliationApiHandler)
    print(f"Reconciliation API server started on http://{LISTEN_HOST}:{LISTEN_PORT} (static root: {STATIC_ROOT})")
    server.serve_forever()


if __name__ == "__main__":
    main()
