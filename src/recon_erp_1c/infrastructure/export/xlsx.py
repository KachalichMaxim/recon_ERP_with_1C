from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


def reconciliation_run_xlsx(run: dict[str, Any]) -> bytes:
    issues = run.get("issues") if isinstance(run.get("issues"), list) else []
    delivery = run.get("delivery") if isinstance(run.get("delivery"), dict) else {}
    summary = run.get("summary") if isinstance(run.get("summary"), dict) else {}
    matched = [issue for issue in issues if issue.get("status") == "match"]
    problems = [issue for issue in issues if issue.get("status") != "match"]
    feedback = run.get("review_feedback") if isinstance(run.get("review_feedback"), dict) else {}
    balance = run.get("balance_comparison") if isinstance(run.get("balance_comparison"), dict) else {}

    wb = Workbook()
    result_ws = wb.active
    result_ws.title = "Итог"
    _build_reconciliation_summary_sheet(result_ws, run, delivery, balance, issues, matched, problems)

    review_ws = wb.create_sheet("К разбору")
    _build_review_sheet(review_ws, problems, feedback)

    matched_ws = wb.create_sheet("Совпало")
    _build_matched_sheet(matched_ws, matched)

    technical_ws = wb.create_sheet("Технические данные")
    _build_technical_sheet(technical_ws, run, summary, issues, feedback)

    return _save_workbook(wb)


def _build_reconciliation_summary_sheet(
    ws: Worksheet,
    run: dict[str, Any],
    delivery: dict[str, Any],
    balance: dict[str, Any],
    issues: list[dict[str, Any]],
    matched: list[dict[str, Any]],
    problems: list[dict[str, Any]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A14"
    ws.merge_cells("A1:H1")
    ws["A1"] = "Проверка документов по поставке ERP ↔ 1С"
    ws.merge_cells("A2:H2")
    ws["A2"] = _delivery_caption(delivery)
    ws.merge_cells("A3:H3")
    ws["A3"] = _run_summary_text(balance, problems)

    result_label = _run_result_label(balance, problems)
    contract_codes = delivery.get("contract_codes") if isinstance(delivery.get("contract_codes"), dict) else {}
    period = run.get("period") if isinstance(run.get("period"), dict) else {}
    ws["A5"], ws["B5"] = "Результат проверки", result_label
    ws["A6"], ws["B6"] = "Проверено строк", len(issues)
    ws["A7"], ws["B7"] = "Совпало", len(matched)
    ws["A8"], ws["B8"] = "Требует разбора", len(problems)
    ws["D5"], ws["E5"] = "Дата запуска", _excel_datetime(run.get("created_at"))
    ws["D6"], ws["E6"] = "Договор покупателя 1С", contract_codes.get("buyer_contract_code") or "—"
    ws["D7"], ws["E7"] = "Договор комитента 1С", contract_codes.get("committent_contract_code") or "—"
    ws["D8"], ws["E8"] = "Период документов", _period_label(period)

    balance_values = [
        ("Сальдо ERP", _nested_money_number(balance.get("erp_balance")), _nested_money_currency(balance.get("erp_balance"))),
        ("Сальдо 1С", _nested_money_number(balance.get("onec_balance")), _nested_money_currency(balance.get("onec_balance"))),
        ("Разница ERP − 1С", _nested_money_number(balance.get("difference")), _nested_money_currency(balance.get("difference"))),
        ("Сальдо сопоставимо", "Да" if balance.get("comparable", True) else "Нет", ""),
    ]
    for col, (label, value, currency) in zip((1, 3, 5, 7), balance_values):
        ws.cell(10, col, label)
        ws.cell(10, col + 1, value)
        if isinstance(value, (int, float, Decimal)):
            ws.cell(10, col + 1).number_format = _currency_format(currency)

    ws.merge_cells("A11:H11")
    ws["A11"] = "Совпадение сальдо не означает, что все документы оформлены и сопоставлены корректно."
    ws.merge_cells("A13:H13")
    ws["A13"] = "Что требует внимания"
    headers = [
        "Проблема",
        "Документ ERP",
        "Документ 1С",
        "ERP дата",
        "1С дата",
        "ERP сумма",
        "1С сумма",
        "Что сделать",
    ]
    _append_rows(ws, [headers])
    header_row = ws.max_row
    if problems:
        for issue in problems:
            erp = _document(issue.get("erp_document"))
            onec = _document(issue.get("onec_document"))
            row_idx = ws.max_row + 1
            ws.append(
                [
                    _issue_label(issue),
                    _document_title(erp),
                    _document_title(onec),
                    _excel_date(erp.get("date")),
                    _excel_date(onec.get("date")),
                    _money_number(erp),
                    _money_number(onec),
                    _issue_action(issue),
                ]
            )
            _set_document_hyperlink(ws.cell(row_idx, 2), erp)
            _format_money_cell(ws.cell(row_idx, 6), _money_currency(erp))
            _format_money_cell(ws.cell(row_idx, 7), _money_currency(onec))
    else:
        ws.append(["Проблем не найдено. Все проверенные строки совпали."] + [""] * 7)

    _style_reconciliation_summary_sheet(ws, header_row, ws.max_row, result_label, bool(problems))


def _build_review_sheet(
    ws: Worksheet,
    problems: list[dict[str, Any]],
    feedback: dict[str, Any],
) -> None:
    headers = [
        "Проблема",
        "Что произошло",
        "Что сделать",
        "Тип документа",
        "ERP документ",
        "ERP дата",
        "ERP сумма",
        "1С документ",
        "1С дата",
        "1С сумма",
        "Причина разбора",
        "Комментарий пользователя",
    ]
    ws.append(headers)
    for issue in problems:
        erp = _document(issue.get("erp_document"))
        onec = _document(issue.get("onec_document"))
        issue_feedback = _issue_feedback(issue, feedback)
        row_idx = ws.max_row + 1
        ws.append(
            [
                _issue_label(issue),
                _issue_explanation(issue),
                _issue_action(issue),
                _document_kind_label(erp.get("kind") or onec.get("kind")),
                _document_title(erp),
                _excel_date(erp.get("date")),
                _money_number(erp),
                _document_title(onec),
                _excel_date(onec.get("date")),
                _money_number(onec),
                issue_feedback.get("reason_label") or issue_feedback.get("reason") or "",
                issue_feedback.get("comment") or "",
            ]
        )
        _set_document_hyperlink(ws.cell(row_idx, 5), erp)
        _format_money_cell(ws.cell(row_idx, 7), _money_currency(erp))
        _format_money_cell(ws.cell(row_idx, 10), _money_currency(onec))
    if not problems:
        ws.append(["Проблем не найдено"] + [""] * (len(headers) - 1))
    _style_issue_table(ws, "ReviewTable", problem_sheet=True)


def _build_matched_sheet(ws: Worksheet, matched: list[dict[str, Any]]) -> None:
    headers = [
        "Тип документа",
        "Документ ERP",
        "ERP дата",
        "ERP сумма",
        "Документ 1С",
        "1С дата",
        "1С сумма",
        "Как сопоставлено",
    ]
    ws.append(headers)
    for issue in matched:
        erp = _document(issue.get("erp_document"))
        onec = _document(issue.get("onec_document"))
        row_idx = ws.max_row + 1
        ws.append(
            [
                _document_kind_label(erp.get("kind") or onec.get("kind")),
                _document_title(erp),
                _excel_date(erp.get("date")),
                _money_number(erp),
                _document_title(onec),
                _excel_date(onec.get("date")),
                _money_number(onec),
                _match_basis_label(issue.get("match_basis")),
            ]
        )
        _set_document_hyperlink(ws.cell(row_idx, 2), erp)
        _format_money_cell(ws.cell(row_idx, 4), _money_currency(erp))
        _format_money_cell(ws.cell(row_idx, 7), _money_currency(onec))
    if not matched:
        ws.append(["Совпавших строк нет"] + [""] * (len(headers) - 1))
    _style_issue_table(ws, "MatchedTable", problem_sheet=False)


def _build_technical_sheet(
    ws: Worksheet,
    run: dict[str, Any],
    summary: dict[str, Any],
    issues: list[dict[str, Any]],
    feedback: dict[str, Any],
) -> None:
    headers = [
        "Ключ строки",
        "Статус",
        "Основная причина",
        "Критичность",
        "Уверенность",
        "Основание сопоставления",
        "ID строки/распределения 1С",
        "Сообщение",
        "Поля",
        "ERP тип",
        "ERP код 1С",
        "ERP номер",
        "ERP дата",
        "ERP сумма",
        "ERP валюта",
        "ERP договор 1С",
        "ERP source_id",
        "ERP operation_id",
        "ERP URL",
        "1С тип",
        "1С код",
        "1С номер",
        "1С дата",
        "1С сумма",
        "1С валюта",
        "1С договор",
        "1С source_id",
        "Причина разбора",
        "Комментарий пользователя",
    ]
    ws.append(headers)
    for issue in issues:
        erp = _document(issue.get("erp_document"))
        onec = _document(issue.get("onec_document"))
        issue_feedback = _issue_feedback(issue, feedback)
        row_idx = ws.max_row + 1
        ws.append(
            [
                issue.get("issue_key") or "",
                issue.get("status") or "",
                issue.get("primary_reason") or "",
                issue.get("severity") or "",
                issue.get("match_confidence") or "",
                issue.get("match_basis") or "",
                issue.get("matched_detail_id") or "",
                issue.get("message") or "",
                ", ".join(issue.get("fields") or []),
                erp.get("kind") or "",
                erp.get("code1c") or "",
                erp.get("number") or "",
                _excel_date(erp.get("date")),
                _money_number(erp),
                _money_currency(erp),
                erp.get("contract_code1c") or "",
                erp.get("source_id") or "",
                erp.get("operation_id") or "",
                erp.get("erp_url") or erp.get("operation_url") or "",
                onec.get("kind") or "",
                onec.get("code1c") or "",
                onec.get("number") or "",
                _excel_date(onec.get("date")),
                _money_number(onec),
                _money_currency(onec),
                onec.get("contract_code1c") or "",
                onec.get("source_id") or "",
                issue_feedback.get("reason_label") or issue_feedback.get("reason") or "",
                issue_feedback.get("comment") or "",
            ]
        )
        _format_money_cell(ws.cell(row_idx, 14), _money_currency(erp))
        _format_money_cell(ws.cell(row_idx, 24), _money_currency(onec))

    _style_issue_table(ws, "TechnicalTable", technical=True)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = (
        f"Run ID: {run.get('run_id') or '—'} · Проверено: {summary.get('issues_total') or len(issues)}"
    )


def reconciliation_matrix_xlsx(matrix: dict[str, Any]) -> bytes:
    items = matrix.get("items") if isinstance(matrix.get("items"), list) else []
    summary = matrix.get("summary") if isinstance(matrix.get("summary"), dict) else {}
    rows = [
        [
            "№ спецификации",
            "",
            "Счет",
            "Сумма по счету",
            "Сумма оплаты",
            "",
            "Возмещаемые расходы",
            "Невозмещаемые расходы",
            "",
            "№ счф",
            "(+/-)",
        ]
    ]
    merges: list[str] = []
    balances_by_row: dict[int, Decimal] = {}
    links_by_row: dict[int, str] = {}

    current_row = 2
    for item in items:
        invoice_rows = _invoice_rows(item)
        start_row = current_row
        balance = _to_decimal(item.get("balance"))
        for idx, invoice in enumerate(invoice_rows):
            if idx == 0 and item.get("erp_url"):
                links_by_row[current_row] = str(item.get("erp_url"))
            rows.append(
                [
                    _spec_label(item) if idx == 0 else "",
                    "",
                    invoice.get("number") or "—",
                    _to_decimal(invoice.get("amount")) if invoice.get("amount") not in (None, "") else "",
                    _to_decimal(invoice.get("paid_amount")) if invoice.get("paid_amount") not in (None, "") else "",
                    "",
                    _to_decimal(item.get("reimbursable_sum")) if idx == 0 else "",
                    _to_decimal(item.get("non_reimbursable_sum")) if idx == 0 else "",
                    "",
                    "\n".join(item.get("sf_numbers") or []) if idx == 0 else "",
                    _to_decimal(item.get("balance")) if idx == 0 else "",
                ]
            )
            balances_by_row[current_row] = balance
            current_row += 1
        end_row = current_row - 1
        if end_row > start_row:
            for col_idx in [1, 7, 8, 10, 11]:
                col = _col(col_idx)
                merges.append(f"{col}{start_row}:{col}{end_row}")

    totals = summary if summary else _matrix_totals(items)
    total_row = current_row + 1
    rows.append([""] * 11)
    rows.append(
        [
            "ИТОГО",
            "",
            "",
            _to_decimal(totals.get("invoice_sum")),
            _to_decimal(totals.get("payment_sum")),
            "",
            _to_decimal(totals.get("reimbursable_sum")),
            _to_decimal(totals.get("non_reimbursable_sum")),
            "",
            "",
            _to_decimal(totals.get("balance")),
        ]
    )

    rules = [
        ["Поле", "Источник", "Правило"],
        [
            "№ спецификации",
            "ERP veda_specs + veda_spr(f_type=33/130)",
            "Тип и номер поставки из исходных таблиц ERP; view_specinv/view_specs не использовать.",
        ],
        [
            "Счет",
            "ERP veda_schets, f_type=1",
            "Только счета покупателю; по каждому счету отдельная строка. Если backend не передал invoice_rows, номер берется из invoice_numbers, а агрегатная сумма ставится в первую строку блока.",
        ],
        ["Сумма по счету", "ERP veda_schets.f_sum", "Сумма конкретного счета покупателю."],
        ["Сумма оплаты", "ERP get_paidsum(veda_schets.f_operid)", "Сумма оплаты назначается операции один раз. Для агрегированного счета складываются оплаты операций его дочерних счетов. Только положительный остаток без доказанной связи со счетом выводится отдельной строкой; отрицательные технические строки запрещены."],
        ["Возмещаемые расходы", "ERP get_realizsum по операциям f_isvozm=1", "Итог по поставке; объединяется по строкам счетов."],
        ["Невозмещаемые расходы", "ERP get_realizsum по операциям f_isvozm=2", "Итог по поставке; объединяется по строкам счетов."],
        ["№ счф", "ERP/1С закрывающие документы", "Номер и дата каждого документа выводятся с переносом строки. Если есть агрегированный документ, дочерние документы не выводятся."],
        ["(+/-)", "Сумма оплаты - возмещаемые - невозмещаемые", "Сальдо взаиморасчетов по поставке."],
        ["Дата формирования", "Сервис сверки", datetime.now().isoformat(timespec="seconds")],
        ["Режим", "Сервис сверки", matrix.get("mode") or ""],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Выгрузка"
    _append_rows(ws, rows)
    _style_matrix_sheet(ws, merges, balances_by_row, total_row)
    for row_idx, url in links_by_row.items():
        cell = ws.cell(row_idx, 1)
        cell.hyperlink = url
        cell.font = Font(bold=True, color="0563C1", underline="single")

    rules_ws = wb.create_sheet("Правила")
    _append_rows(rules_ws, rules)
    _style_rules_sheet(rules_ws)

    return _save_workbook(wb)


def _save_workbook(wb: Workbook) -> bytes:
    with BytesIO() as buffer:
        wb.save(buffer)
        return buffer.getvalue()


def _append_rows(ws: Worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append([_cell_value(value) for value in row])


def _cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def _style_matrix_sheet(ws: Worksheet, merges: list[str], balances_by_row: dict[int, Decimal], total_row: int) -> None:
    widths = {1: 24, 2: 3, 3: 18, 4: 18, 5: 18, 6: 3, 7: 20, 8: 22, 9: 3, 10: 48, 11: 18}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="D9D7D2")
    spec_fill = PatternFill("solid", fgColor="F3F1ED")
    invoice_fill = PatternFill("solid", fgColor="EEF6FF")
    payment_fill = PatternFill("solid", fgColor="EEF8F0")
    expense_fill = PatternFill("solid", fgColor="FFF5DF")
    sf_fill = PatternFill("solid", fgColor="F7F4EF")
    positive_fill = PatternFill("solid", fgColor="E2F0D9")
    negative_fill = PatternFill("solid", fgColor="FCE4D6")
    muted_fill = PatternFill("solid", fgColor="FAF9F6")
    border = _thin_border()

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    money_cols = {4, 5, 7, 8, 11}
    for row_idx in range(2, ws.max_row + 1):
        if row_idx == total_row:
            fills = {4: header_fill, 5: header_fill, 7: header_fill, 8: header_fill, 11: header_fill}
            for col_idx in range(1, 12):
                cell = ws.cell(row_idx, col_idx)
                cell.font = Font(bold=True)
                cell.fill = fills.get(col_idx, header_fill)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right" if col_idx in money_cols else "left", vertical="center", wrap_text=True
                )
                if col_idx in money_cols:
                    cell.number_format = _money_format()
            continue

        balance = balances_by_row.get(row_idx, Decimal("0"))
        for col_idx in range(1, 12):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if col_idx in money_cols else "left", vertical="center", wrap_text=True
            )
            if col_idx == 1:
                cell.font = Font(bold=True if cell.value else False)
                cell.fill = spec_fill
            elif col_idx in {2, 6, 9}:
                cell.fill = muted_fill
            elif col_idx in {3, 4}:
                cell.fill = invoice_fill
            elif col_idx == 5:
                cell.fill = payment_fill
            elif col_idx in {7, 8}:
                cell.fill = expense_fill
            elif col_idx == 10:
                cell.fill = sf_fill
            elif col_idx == 11:
                cell.fill = positive_fill if balance >= 0 else negative_fill
                cell.font = Font(color="9C0006" if balance < 0 else "006100")
            if col_idx in money_cols:
                cell.number_format = _money_format()

    for ref in merges:
        ws.merge_cells(ref)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 28


def _style_rules_sheet(ws: Worksheet) -> None:
    widths = {1: 24, 2: 38, 3: 92}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _style_simple_sheet(ws)


def _style_simple_sheet(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    border = _thin_border()
    header_fill = PatternFill("solid", fgColor="D9D7D2")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(18, col_idx * 8), 42)


def _style_reconciliation_summary_sheet(
    ws: Worksheet,
    header_row: int,
    last_row: int,
    result_label: str,
    has_problems: bool,
) -> None:
    widths = {1: 24, 2: 25, 3: 25, 4: 15, 5: 15, 6: 18, 7: 18, 8: 58}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[11].height = 30

    dark_blue = PatternFill("solid", fgColor="244F73")
    light_blue = PatternFill("solid", fgColor="DCEAF5")
    pale_blue = PatternFill("solid", fgColor="EEF5FA")
    pale_green = PatternFill("solid", fgColor="E2F0D9")
    pale_red = PatternFill("solid", fgColor="FCE8E6")
    pale_yellow = PatternFill("solid", fgColor="FFF2CC")
    section_fill = PatternFill("solid", fgColor="E7F1E3" if has_problems else "E2F0D9")
    border = _thin_border()

    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = dark_blue
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"].font = Font(name="Arial", size=11, bold=True, color="1F2937")
    ws["A2"].fill = light_blue
    ws["A3"].font = Font(name="Arial", size=11, bold=True, color="9C0006" if has_problems else "006100")
    ws["A3"].fill = pale_red if has_problems else pale_green
    ws["A3"].alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx in range(5, 9):
        for col_idx in (1, 4):
            cell = ws.cell(row_idx, col_idx)
            cell.font = Font(name="Arial", bold=True, color="374151")
            cell.fill = pale_blue
            cell.border = border
        for col_idx in (2, 5):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws["B5"].font = Font(
        name="Arial",
        bold=True,
        color="9C0006" if result_label == "Требуется разбор" else "006100",
    )
    ws["E5"].number_format = 'dd"."mm"."yyyy hh:mm'

    for col_idx in range(1, 9):
        cell = ws.cell(10, col_idx)
        cell.border = border
        cell.fill = pale_blue if col_idx % 2 else PatternFill("solid", fgColor="FFFFFF")
        cell.font = Font(name="Arial", bold=col_idx % 2 == 1, color="374151")
        cell.alignment = Alignment(horizontal="right" if col_idx % 2 == 0 else "left", vertical="center")
    difference = ws["F10"].value
    if isinstance(difference, (int, float, Decimal)):
        ws["F10"].font = Font(name="Arial", bold=True, color="006100" if abs(float(difference)) < 0.005 else "9C0006")
    ws["A11"].fill = pale_yellow
    ws["A11"].font = Font(name="Arial", italic=True, color="7F6000")
    ws["A11"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["A13"].fill = section_fill
    ws["A13"].font = Font(name="Arial", size=12, bold=True, color="385723")

    _style_table_header(ws, header_row, 8)
    for row_idx in range(header_row + 1, last_row + 1):
        fill = PatternFill("solid", fgColor="FCE8E6" if row_idx % 2 else "FFF7F6") if has_problems else pale_green
        for col_idx in range(1, 9):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if col_idx in {6, 7} else "left",
                vertical="top",
                wrap_text=True,
            )
            if col_idx in {4, 5} and cell.value is not None:
                cell.number_format = 'dd"."mm"."yyyy'
        ws.row_dimensions[row_idx].height = 38
    _add_table(ws, f"A{header_row}:H{last_row}", "SummaryIssuesTable")
    ws.auto_filter.ref = f"A{header_row}:H{last_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _style_issue_table(
    ws: Worksheet,
    table_name: str,
    *,
    problem_sheet: bool = False,
    technical: bool = False,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    _style_table_header(ws, 1, ws.max_column)
    border = _thin_border()
    if technical:
        default_width = 18
        wide_cols = {1: 22, 8: 55, 9: 24, 19: 42, 28: 24, 29: 42}
    elif problem_sheet:
        default_width = 18
        wide_cols = {1: 22, 2: 52, 3: 52, 4: 26, 5: 24, 8: 24, 11: 24, 12: 42}
    else:
        default_width = 20
        wide_cols = {1: 34, 2: 24, 5: 24, 8: 32}
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = wide_cols.get(col_idx, default_width)

    for row_idx in range(2, ws.max_row + 1):
        if technical:
            fill_color = "DCEEF8" if row_idx % 2 == 0 else "F3F8FC"
        elif problem_sheet:
            fill_color = "FCE8E6" if row_idx % 2 == 0 else "FFF7F6"
        else:
            fill_color = "E2F0D9" if row_idx % 2 == 0 else "F3F8EF"
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if _is_money_column(ws, col_idx) else "left",
                vertical="top",
                wrap_text=True,
            )
            if "дата" in str(ws.cell(1, col_idx).value or "").lower() and cell.value is not None:
                cell.number_format = 'dd"."mm"."yyyy'
        if problem_sheet:
            ws.cell(row_idx, 1).font = Font(
                name="Arial",
                bold=True,
                color="9C0006" if str(ws.cell(row_idx, 1).value).startswith("Нет") else "9C6500",
            )
        ws.row_dimensions[row_idx].height = 42 if problem_sheet else 28
    _add_table(ws, f"A1:{get_column_letter(ws.max_column)}{ws.max_row}", table_name)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _style_table_header(ws: Worksheet, row_idx: int, max_column: int) -> None:
    fill = PatternFill("solid", fgColor="4472C4")
    border = _thin_border()
    for col_idx in range(1, max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.fill = fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_idx].height = 32


def _add_table(ws: Worksheet, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _is_money_column(ws: Worksheet, col_idx: int) -> bool:
    value = str(ws.cell(1, col_idx).value or "").lower()
    return "сумма" in value


def _document(value: object) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _document_title(document: dict[str, Any]) -> str:
    operation_id = document.get("operation_id")
    return str(
        document.get("code1c")
        or document.get("number")
        or document.get("incoming_number")
        or document.get("source_number")
        or (f"Операция ERP {operation_id}" if operation_id else "")
        or "—"
    )


def _document_kind_label(kind: object) -> str:
    labels = {
        "customer_invoice": "Счет покупателю",
        "payment": "Оплата покупателя",
        "sale": "Реализация / закрывающий документ",
        "purchase": "Поступление поставщика",
        "supplier_invoice": "Счет от поставщика",
    }
    return labels.get(str(kind or ""), str(kind or "Документ"))


def _issue_label(issue: dict[str, Any]) -> str:
    labels = {
        "not_found_in_1c": "Нет в 1С",
        "erp_code1c_missing": "Нет кода 1С в ERP",
        "not_found_in_erp": "Нет в ERP",
        "not_linked_to_delivery_in_erp": "Нет связи с поставкой",
        "amount_mismatch": "Расходится сумма или валюта",
        "date_mismatch": "Расходится дата",
        "contract_mismatch": "Документ на другом договоре",
        "number_mismatch": "Расходится номер",
        "vat_mismatch": "Расходится НДС",
        "duplicate_in_1c": "Дубли в 1С",
        "ambiguous_match": "Неоднозначное сопоставление",
        "aggregation_conflict": "Конфликт агрегации",
        "not_comparable": "Нельзя сравнить",
        "contract_context_missing": "Нет аналитики поставки",
        "missing_erp_invoice": "Нет счета в ERP",
        "erp_invoice_link_missing": "Счет не связан с операцией",
        "missing_erp_closing_document": "Нет закрывающего документа",
    }
    status = str(issue.get("status") or "")
    return labels.get(status, status or "Требует разбора")


def _issue_action(issue: dict[str, Any]) -> str:
    actions = {
        "not_found_in_1c": "Проверить выгрузку документа из ERP в 1С и повторить сверку.",
        "erp_code1c_missing": "Проверить статус выгрузки и заполнение кода 1С в документе ERP.",
        "not_found_in_erp": "Проверить основание создания документа 1С и его связь с ERP.",
        "not_linked_to_delivery_in_erp": "Привязать документ или операцию ERP к выбранной поставке.",
        "amount_mismatch": "Сверить сумму и валюту документа в ERP и 1С.",
        "date_mismatch": "Сверить дату документа в ERP и 1С и определить корректную.",
        "contract_mismatch": "Проверить договор проведения; если другой договор допустим, зафиксировать причину.",
        "number_mismatch": "Сверить номер документа в ERP и 1С.",
        "vat_mismatch": "Сверить ставку НДС в ERP и 1С.",
        "duplicate_in_1c": "Устранить дубль либо указать корректный документ 1С.",
        "ambiguous_match": "Выбрать корректную пару документов и уточнить правило сопоставления.",
        "aggregation_conflict": "Проверить состав агрегированного документа и его строки.",
        "not_comparable": "Заполнить недостающие реквизиты, необходимые для сравнения.",
        "contract_context_missing": "Заполнить в 1С аналитику связи документа с поставкой.",
        "missing_erp_invoice": "Проверить операцию ERP и создать либо привязать счет покупателю.",
        "erp_invoice_link_missing": "Привязать существующий счет покупателю к операции ERP.",
        "missing_erp_closing_document": "Создать либо привязать исходящий закрывающий документ к операции ERP.",
    }
    return actions.get(str(issue.get("status") or ""), "Проверить документы и зафиксировать результат разбора.")


def _issue_explanation(issue: dict[str, Any]) -> str:
    erp = _document(issue.get("erp_document"))
    onec = _document(issue.get("onec_document"))
    status = str(issue.get("status") or "")
    if status == "date_mismatch":
        return f"Документ найден в ERP и 1С, но даты отличаются: {_display_date(erp.get('date'))} / {_display_date(onec.get('date'))}."
    if status == "amount_mismatch":
        return f"Документ найден в ERP и 1С, но сумма или валюта отличаются: {_display_money(erp)} / {_display_money(onec)}."
    if status == "vat_mismatch":
        return f"Документ найден в ERP и 1С, но ставка НДС отличается: {erp.get('vat_rate') or '—'} / {onec.get('vat_rate') or '—'}."
    if status == "number_mismatch":
        return f"Документ найден в обеих системах, но номера отличаются: {_document_title(erp)} / {_document_title(onec)}."
    return str(issue.get("message") or _issue_label(issue))


def _match_basis_label(value: object) -> str:
    labels = {
        "document_header": "По документу",
        "payment_header": "По платежу",
        "payment_header_allocations": "По распределению платежа",
        "payment_allocation": "По распределению платежа",
        "document_allocation": "По распределению документа",
        "document_line": "По строке документа",
    }
    return labels.get(str(value or ""), str(value or "По реквизитам документа"))


def _issue_feedback(issue: dict[str, Any], feedback: dict[str, Any]) -> dict[str, Any]:
    value = feedback.get(str(issue.get("issue_key") or ""))
    return value if isinstance(value, dict) else {}


def _delivery_caption(delivery: dict[str, Any]) -> str:
    counterparty = delivery.get("counterparty") if isinstance(delivery.get("counterparty"), dict) else {}
    spec_number = delivery.get("spec_number") or delivery.get("erp_spec_id") or "—"
    spec_id = delivery.get("erp_spec_id") or "—"
    contract = delivery.get("base_contract_number") or "—"
    client = counterparty.get("name") or "Клиент не указан"
    return f"{client} · Поставка №{spec_number} · ERP ID {spec_id} · договор {contract}"


def _run_summary_text(balance: dict[str, Any], problems: list[dict[str, Any]]) -> str:
    balance_status = str(balance.get("status") or "")
    comparable = bool(balance.get("comparable", True))
    if not comparable:
        balance_text = "Сальдо нельзя сопоставить."
    elif balance_status == "match":
        balance_text = "Сальдо совпало."
    else:
        balance_text = f"Сальдо не совпало: разница {_display_nested_money(balance.get('difference'))}."
    if not problems:
        return f"{balance_text} Все проверенные документы и строки совпали."
    grouped: dict[str, int] = {}
    for issue in problems:
        label = _issue_label(issue)
        grouped[label] = grouped.get(label, 0) + 1
    details = ", ".join(f"{count} — {label}" for label, count in grouped.items())
    return f"{balance_text} Требует разбора: {len(problems)} ({details})."


def _run_result_label(balance: dict[str, Any], problems: list[dict[str, Any]]) -> str:
    if problems or balance.get("status") not in (None, "", "match") or not balance.get("comparable", True):
        return "Требуется разбор"
    return "Проверка пройдена"


def _period_label(period: dict[str, Any]) -> str:
    date_from = _display_date(period.get("date_from"))
    date_to = _display_date(period.get("date_to"))
    if date_from == "—" and date_to == "—":
        return "Автоматически по документам поставки"
    return f"{date_from} — {date_to}"


def _excel_date(value: object) -> object:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return text


def _excel_datetime(value: object) -> object:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        result = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return result.replace(tzinfo=None)
    except ValueError:
        return text


def _display_date(value: object) -> str:
    parsed = _excel_date(value)
    return parsed.strftime("%d.%m.%Y") if hasattr(parsed, "strftime") else str(parsed or "—")


def _money_number(document: dict[str, Any]) -> float | None:
    amount = document.get("amount") if isinstance(document.get("amount"), dict) else {}
    value = amount.get("amount")
    if value in (None, ""):
        return None
    try:
        return float(Decimal(str(value)))
    except (ArithmeticError, ValueError):
        return None


def _nested_money_number(value: object) -> float | None:
    if not isinstance(value, dict) or value.get("amount") in (None, ""):
        return None
    try:
        return float(Decimal(str(value.get("amount"))))
    except (ArithmeticError, ValueError):
        return None


def _nested_money_currency(value: object) -> str:
    return str(value.get("currency") or "RUB") if isinstance(value, dict) else "RUB"


def _display_money(document: dict[str, Any]) -> str:
    amount = _money_number(document)
    if amount is None:
        return "—"
    return f"{amount:,.2f} {_money_currency(document) or 'RUB'}".replace(",", " ")


def _display_nested_money(value: object) -> str:
    amount = _nested_money_number(value)
    if amount is None:
        return "—"
    return f"{amount:,.2f} {_nested_money_currency(value)}".replace(",", " ")


def _currency_format(currency: object) -> str:
    code = str(currency or "RUB").upper()
    return f'#,##0.00" {code}";[Red]-#,##0.00" {code}";0.00" {code}"'


def _format_money_cell(cell: Any, currency: object) -> None:
    if cell.value is not None:
        cell.number_format = _currency_format(currency)


def _set_document_hyperlink(cell: Any, document: dict[str, Any]) -> None:
    url = str(document.get("erp_url") or document.get("operation_url") or "").strip()
    if not url:
        return
    cell.hyperlink = url
    cell.font = Font(name="Arial", color="0563C1", underline="single")


def _thin_border() -> Border:
    side = Side(style="thin", color="D6D3CD")
    return Border(left=side, right=side, top=side, bottom=side)


def _money_format() -> str:
    return '#,##0.00" р.";[Red]-#,##0.00" р.";0.00" р."'


def _matrix_totals(items: list[dict[str, Any]]) -> dict[str, str]:
    fields = ["invoice_sum", "payment_sum", "reimbursable_sum", "non_reimbursable_sum", "balance"]
    return {field: str(sum((_to_decimal(item.get(field)) for item in items), _to_decimal("0"))) for field in fields}


def _to_decimal(value: Any):
    return Decimal(str(value or "0")).quantize(Decimal("0.01"))


def _spec_label(item: dict[str, Any]) -> str:
    spec_type = str(item.get("spec_type_name") or item.get("type_name") or item.get("delivery_type") or "").strip()
    spec_number = item.get("spec_number") or item.get("spec_id") or ""
    if spec_type:
        return f"{spec_type} №{spec_number}"
    return f"Поставка №{spec_number}"


def _invoice_rows(item: dict[str, Any]) -> list[dict[str, Any]]:
    rows = item.get("invoice_rows")
    if isinstance(rows, list) and rows:
        prepared = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            prepared.append(
                {
                    **row,
                    "paid_amount": row.get("paid_amount")
                    if row.get("operation_id") or row.get("payment_operation_ids")
                    else "",
                }
            )
        return _with_unassigned_payment(prepared or [{"number": "—", "amount": item.get("invoice_sum")}], item)
    numbers = item.get("invoice_numbers") if isinstance(item.get("invoice_numbers"), list) else []
    if not numbers:
        return _with_unassigned_payment([{"number": "—", "amount": item.get("invoice_sum")}], item)
    return _with_unassigned_payment([
        {"number": number, "amount": item.get("invoice_sum") if idx == 0 else ""}
        for idx, number in enumerate(numbers)
    ], item)


def _with_unassigned_payment(rows: list[dict[str, Any]], item: dict[str, Any]) -> list[dict[str, Any]]:
    assigned = sum(
        (_to_decimal(row.get("paid_amount")) for row in rows if row.get("paid_amount") not in (None, "")),
        Decimal("0.00"),
    )
    total = _to_decimal(item.get("payment_sum"))
    unassigned = total - assigned
    if unassigned > Decimal("0.00"):
        rows.append(
            {
                "number": "Оплата по операциям без установленной связи со счетом",
                "amount": "",
                "paid_amount": str(unassigned),
            }
        )
    return rows


def _col(index: int) -> str:
    out = ""
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def _money_amount(document: dict[str, Any]) -> str:
    amount = document.get("amount") if isinstance(document.get("amount"), dict) else {}
    return amount.get("amount") or ""


def _nested_money_amount(value: object) -> str:
    if not isinstance(value, dict):
        return ""
    return str(value.get("amount") or "")


def _money_currency(document: dict[str, Any]) -> str:
    amount = document.get("amount") if isinstance(document.get("amount"), dict) else {}
    return amount.get("currency") or ""
