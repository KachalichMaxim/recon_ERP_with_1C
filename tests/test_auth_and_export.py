from __future__ import annotations

import os
from datetime import date
from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from recon_erp_1c.domain.entities import AccountingDocument, RelatedDocument
from recon_erp_1c.domain.value_objects import DocumentKind, Money, SourceSystem
from recon_erp_1c.infrastructure.export.xlsx import reconciliation_matrix_xlsx, reconciliation_run_xlsx
from recon_erp_1c.interfaces.http.api import _matrix_row
from recon_erp_1c.interfaces.http import api as http_api
from recon_erp_1c.interfaces.http.auth import AuthenticationError
from recon_erp_1c.interfaces.http.auth import create_session_token, verify_session_token


def test_session_token_roundtrip() -> None:
    token = create_session_token({"user_id": 7, "login": "user@example.com", "name": "User Name"})

    payload = verify_session_token(token)

    assert payload is not None
    assert payload["sub"] == "7"
    assert payload["email"] == "user@example.com"
    assert payload["name"] == "User Name"


def test_erp_launch_token_can_be_validated_read_only_in_erp_mariadb(monkeypatch) -> None:
    class Repository:
        def find_user_by_api_token(self, token: str):
            assert token == "launch-token"
            return {"user_id": 7, "login": "user@example.com", "name": "User Name", "structure_code": "FIN"}

    monkeypatch.delenv("RECON_ERP_TOKEN_VALIDATE_URL", raising=False)
    monkeypatch.setattr(http_api, "_erp_repository", lambda: Repository())

    profile = http_api._validate_erp_launch_token("launch-token")

    assert profile["user_id"] == 7
    assert profile["login"] == "user@example.com"


def test_invalid_erp_launch_token_is_rejected(monkeypatch) -> None:
    class Repository:
        def find_user_by_api_token(self, token: str):
            return None

    monkeypatch.delenv("RECON_ERP_TOKEN_VALIDATE_URL", raising=False)
    monkeypatch.setattr(http_api, "_erp_repository", lambda: Repository())

    try:
        http_api._validate_erp_launch_token("invalid")
        assert False, "expected AuthenticationError"
    except AuthenticationError as exc:
        assert "Invalid or expired" in str(exc)


def test_session_secret_required_in_production() -> None:
    old_values = {key: os.environ.get(key) for key in ["RECON_ENV", "RECON_REQUIRE_ERP_TOKEN", "RECON_SESSION_SECRET", "RECON_UI_DEMO", "RECON_DEV_AUTH"]}
    try:
        os.environ["RECON_ENV"] = "production"
        os.environ["RECON_REQUIRE_ERP_TOKEN"] = "1"
        os.environ.pop("RECON_SESSION_SECRET", None)
        os.environ["RECON_UI_DEMO"] = "0"
        os.environ["RECON_DEV_AUTH"] = "0"
        try:
            create_session_token({"user_id": 7, "login": "user@example.com", "name": "User Name"})
            assert False, "expected RuntimeError"
        except RuntimeError as exc:
            assert "RECON_SESSION_SECRET" in str(exc)
    finally:
        for key, value in old_values.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


def test_session_secret_can_be_omitted_in_demo_mode() -> None:
    old_values = {key: os.environ.get(key) for key in ["RECON_ENV", "RECON_REQUIRE_ERP_TOKEN", "RECON_SESSION_SECRET", "RECON_UI_DEMO", "RECON_DEV_AUTH"]}
    try:
        os.environ["RECON_ENV"] = "production"
        os.environ["RECON_REQUIRE_ERP_TOKEN"] = "1"
        os.environ.pop("RECON_SESSION_SECRET", None)
        os.environ["RECON_UI_DEMO"] = "1"
        os.environ["RECON_DEV_AUTH"] = "0"
        token = create_session_token({"user_id": 7, "login": "user@example.com", "name": "User Name"})
        assert verify_session_token(token) is not None
    finally:
        for key, value in old_values.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


def test_reconciliation_xlsx_is_valid_zip_package() -> None:
    raw = reconciliation_run_xlsx(
        {
            "run_id": "test",
            "created_at": "2026-07-02T00:00:00",
            "delivery": {
                "erp_spec_id": 20334,
                "spec_number": "921",
                "base_contract_number": "660/1",
                "contract_codes": {"buyer_contract_code": "БП-051945", "committent_contract_code": "БП-051946"},
            },
            "summary": {"issues_total": 1},
            "issues": [
                {
                    "status": "match",
                    "message": "Документ совпал",
                    "fields": [],
                    "erp_document": {
                        "kind": "payment",
                        "code1c": "00БП-010299",
                        "number": "29195",
                        "date": "2025-07-30",
                        "amount": {"amount": "442296.92", "currency": "RUB"},
                    },
                    "onec_document": {
                        "kind": "payment",
                        "code1c": "00БП-010299",
                        "number": "29195",
                        "date": "2025-07-30",
                        "amount": {"amount": "442296.92", "currency": "RUB"},
                    },
                }
            ],
        }
    )

    with ZipFile(BytesIO(raw)) as workbook:
        names = set(workbook.namelist())

    assert "[Content_Types].xml" in names
    assert "xl/workbook.xml" in names
    assert "xl/worksheets/sheet1.xml" in names


def test_matrix_xlsx_is_valid_zip_package() -> None:
    raw = reconciliation_matrix_xlsx(
        {
            "mode": "test",
            "summary": {"deliveries": 1, "balance": "-100.00"},
            "items": [
                {
                    "spec_id": 20334,
                    "spec_number": "921",
                    "client_id": 221,
                    "client_name": "ООО АЭРО-ТРЕЙД",
                    "dog_id": 88,
                    "base_contract_number": "660/1",
                    "invoice_numbers": ["ВА-015695", "ВА-015696"],
                    "sf_numbers": ["00БП-000198"],
                    "invoice_sum": "100.00",
                    "payment_sum": "50.00",
                    "reimbursable_sum": "100.00",
                    "non_reimbursable_sum": "50.00",
                    "balance": "-100.00",
                    "balance_label": "Долг",
                }
            ],
        }
    )

    with ZipFile(BytesIO(raw)) as workbook:
        names = set(workbook.namelist())
        workbook_xml = workbook.read("xl/workbook.xml").decode("utf-8")
        export_sheet_xml = workbook.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert "[Content_Types].xml" in names
    assert "xl/workbook.xml" in names
    assert "xl/worksheets/sheet1.xml" in names
    assert "Выгрузка" in workbook_xml
    assert "Правила" in workbook_xml
    assert "<mergeCells" in export_sheet_xml
    assert '<mergeCell ref="A2:A' in export_sheet_xml

    parsed = load_workbook(BytesIO(raw))
    ws = parsed["Выгрузка"]
    assert ws.max_row > 1
    assert ws["A1"].value == "№ спецификации"
    assert ws["A2"].value == "Поставка №921"


def test_matrix_row_uses_procedure_calculation_instead_of_document_totals() -> None:
    document = AccountingDocument(
        source=SourceSystem.ERP,
        kind=DocumentKind.SALE,
        code1c="ACT-1",
        number="ACT-1",
        date=date(2025, 7, 1),
        amount=Money.of("10"),
        contract_code1c="B",
        reimbursement_type="non_reimbursable",
    )

    row = _matrix_row(
        None,  # repository is not used when documents are supplied
        {
            "spec_id": 20334,
            "spec_number": "1051",
            "buyer_contract_code": "B",
        },
        [document],
        {
            "payment_sum": "203091.35",
            "reimbursable_sum": "171364.16",
            "non_reimbursable_sum": "38394.01",
            "balance": "-6666.82",
        },
    )

    assert row["payment_sum"] == "203091.35"
    assert row["reimbursable_sum"] == "171364.16"
    assert row["non_reimbursable_sum"] == "38394.01"
    assert row["balance"] == "-6666.82"
    assert row["balance_label"] == "Долг"


def test_matrix_payment_is_assigned_once_and_closing_document_has_date() -> None:
    invoices = [
        AccountingDocument(
            source=SourceSystem.ERP,
            kind=DocumentKind.CUSTOMER_INVOICE,
            code1c=code,
            number=code,
            date=date(2025, 7, 1),
            amount=Money.of(amount),
            contract_code1c="B",
            operation_id=777,
        )
        for code, amount in (("INV-1", "52000"), ("INV-2", "7802"))
    ]
    payment = AccountingDocument(
        source=SourceSystem.ERP,
        kind=DocumentKind.PAYMENT,
        code1c="PAY-1",
        number="PAY-1",
        date=date(2025, 7, 2),
        amount=Money.of("59802"),
        contract_code1c="B",
        operation_id=777,
    )
    sale = AccountingDocument(
        source=SourceSystem.ERP,
        kind=DocumentKind.SALE,
        code1c="SF-1",
        number="SF-1",
        date=date(2025, 7, 3),
        amount=Money.of("59802"),
        contract_code1c="B",
        operation_id=777,
        reimbursement_type="non_reimbursable",
    )

    row = _matrix_row(
        None,
        {"spec_id": 1, "spec_number": "56", "buyer_contract_code": "B"},
        [*invoices, payment, sale],
    )

    assert [invoice["paid_amount"] for invoice in row["invoice_rows"]] == ["59802.00", ""]
    assert row["sf_numbers"] == ["SF-1 от 03.07.2025"]


def test_matrix_xlsx_never_adds_negative_unlinked_payment_row() -> None:
    raw = reconciliation_matrix_xlsx(
        {
            "mode": "test",
            "items": [
                {
                    "spec_number": "56",
                    "invoice_rows": [
                        {"number": "INV-1", "operation_id": 777, "amount": "52000", "paid_amount": "59802"},
                        {"number": "INV-2", "operation_id": 777, "amount": "7802", "paid_amount": "59802"},
                    ],
                    "payment_sum": "59802",
                    "reimbursable_sum": "0",
                    "non_reimbursable_sum": "445776.70",
                    "balance": "4960",
                    "sf_numbers": ["SF-1 от 03.07.2025"],
                }
            ],
        }
    )

    ws = load_workbook(BytesIO(raw))["Выгрузка"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "Оплата без счета" not in values
    assert "Оплата, не связанная со счетом" not in values
    assert all(not (isinstance(value, (int, float)) and value == -59802) for value in values)


def test_aggregate_invoice_collects_payments_from_child_operations() -> None:
    invoice = AccountingDocument(
        source=SourceSystem.ERP,
        kind=DocumentKind.CUSTOMER_INVOICE,
        code1c="VL-001599",
        number="VL-001599",
        date=date(2026, 6, 11),
        amount=Money.of("589160"),
        contract_code1c="BP-093741",
        related_documents=(
            RelatedDocument(source_id="1", number="VL-001903/0", operation_id=477611),
            RelatedDocument(source_id="2", number="VL-001904/1", operation_id=477612),
        ),
    )
    payments = [
        AccountingDocument(
            source=SourceSystem.ERP,
            kind=DocumentKind.PAYMENT,
            code1c="PAY-1",
            number="466",
            date=date(2026, 6, 16),
            amount=Money.of("450260"),
            contract_code1c="BP-093741",
            operation_id=477611,
        ),
        AccountingDocument(
            source=SourceSystem.ERP,
            kind=DocumentKind.PAYMENT,
            code1c="PAY-1",
            number="466",
            date=date(2026, 6, 16),
            amount=Money.of("138900"),
            contract_code1c="BP-093741",
            operation_id=477612,
        ),
    ]

    row = _matrix_row(
        None,
        {"spec_id": 28327, "spec_number": "37/1", "buyer_contract_code": "BP-093741"},
        [invoice, *payments],
    )

    assert row["invoice_rows"][0]["paid_amount"] == "589160.00"
    raw = reconciliation_matrix_xlsx({"items": [row]})
    values = [cell.value for line in load_workbook(BytesIO(raw))["Выгрузка"].iter_rows() for cell in line]
    assert "Оплата по операциям без установленной связи со счетом" not in values
